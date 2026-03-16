#!/usr/bin/env python3
"""
Generate sample Excel files representing common Japanese enterprise document patterns.
These files are designed to test an Excel-to-JSON/HTML converter tool.

Usage:
    python generate_samples.py
"""

import os
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
    numbers,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT_DIR = Path(__file__).parent / "output"

# ---------------------------------------------------------------------------
# Shared style helpers
# ---------------------------------------------------------------------------

THIN = Side(style="thin")
MEDIUM = Side(style="medium")
THICK = Side(style="thick")
HAIR = Side(style="hair")

BORDER_ALL_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_ALL_MEDIUM = Border(left=MEDIUM, right=MEDIUM, top=MEDIUM, bottom=MEDIUM)
BORDER_BOTTOM_THIN = Border(bottom=THIN)
BORDER_BOTTOM_MEDIUM = Border(bottom=MEDIUM)

FILL_LIGHT_BLUE = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
FILL_LIGHT_GREEN = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
FILL_LIGHT_YELLOW = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
FILL_LIGHT_GRAY = PatternFill(start_color="E5E7E9", end_color="E5E7E9", fill_type="solid")
FILL_DARK_BLUE = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")
FILL_DARK_GREEN = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
FILL_RED = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
FILL_ORANGE = PatternFill(start_color="F5B041", end_color="F5B041", fill_type="solid")
FILL_WHITE = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
FILL_GANTT_BLUE = PatternFill(start_color="5DADE2", end_color="5DADE2", fill_type="solid")
FILL_GANTT_GREEN = PatternFill(start_color="58D68D", end_color="58D68D", fill_type="solid")
FILL_GANTT_ORANGE = PatternFill(start_color="F0B27A", end_color="F0B27A", fill_type="solid")
FILL_PASS = PatternFill(start_color="82E0AA", end_color="82E0AA", fill_type="solid")
FILL_FAIL = PatternFill(start_color="F1948A", end_color="F1948A", fill_type="solid")
FILL_PENDING = PatternFill(start_color="F9E79F", end_color="F9E79F", fill_type="solid")

FONT_TITLE = Font(name="Yu Gothic", size=16, bold=True)
FONT_HEADING = Font(name="Yu Gothic", size=12, bold=True)
FONT_SUBHEADING = Font(name="Yu Gothic", size=10, bold=True)
FONT_BODY = Font(name="Yu Gothic", size=10)
FONT_SMALL = Font(name="Yu Gothic", size=9)
FONT_WHITE_BOLD = Font(name="Yu Gothic", size=10, bold=True, color="FFFFFF")
FONT_BOLD = Font(name="Yu Gothic", size=10, bold=True)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)


def _apply_style(cell, font=None, fill=None, border=None, alignment=None, number_format=None):
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if border:
        cell.border = border
    if alignment:
        cell.alignment = alignment
    if number_format:
        cell.number_format = number_format


def _apply_border_range(ws, min_row, max_row, min_col, max_col, border=BORDER_ALL_THIN):
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(row=r, column=c).border = border


# ---------------------------------------------------------------------------
# 1. Excel方眼紙 (Grid paper / Houganshi) - 稟議書
# ---------------------------------------------------------------------------

def generate_houganshi():
    wb = Workbook()
    ws = wb.active
    ws.title = "稟議書"

    # Very narrow columns (grid paper style) - 40 columns of width ~2.5
    for col in range(1, 41):
        ws.column_dimensions[get_column_letter(col)].width = 2.8

    # --- Title ---
    ws.merge_cells("A1:AN1")  # AN = col 40
    ws.row_dimensions[1].height = 36
    c = ws["A1"]
    c.value = "稟 議 書"
    _apply_style(c, font=Font(name="Yu Gothic", size=20, bold=True), alignment=ALIGN_CENTER)

    # --- Document number / date block (right side, within 40 cols) ---
    ws.merge_cells(start_row=2, start_column=25, end_row=2, end_column=40)
    ws.cell(row=2, column=25).value = "文書番号：R-2026-0042"
    _apply_style(ws.cell(row=2, column=25), font=FONT_SMALL, alignment=ALIGN_RIGHT)

    ws.merge_cells(start_row=3, start_column=25, end_row=3, end_column=40)
    ws.cell(row=3, column=25).value = "起案日：2026年3月5日"
    _apply_style(ws.cell(row=3, column=25), font=FONT_SMALL, alignment=ALIGN_RIGHT)

    # --- Approval stamps area (fit within 40 cols: cols 25-40) ---
    row = 4
    labels = ["決裁", "部長", "課長", "起案者"]
    for i, label in enumerate(labels):
        col_start = 25 + i * 4
        col_end = col_start + 3
        ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
        c = ws.cell(row=row, column=col_start)
        c.value = label
        _apply_style(c, font=FONT_SMALL, fill=FILL_LIGHT_GRAY, border=BORDER_ALL_THIN, alignment=ALIGN_CENTER)
        # Stamp box below
        ws.merge_cells(start_row=row + 1, start_column=col_start, end_row=row + 3, end_column=col_end)
        stamp_cell = ws.cell(row=row + 1, column=col_start)
        _apply_style(stamp_cell, border=BORDER_ALL_THIN, alignment=ALIGN_CENTER)
        if label == "起案者":
            stamp_cell.value = "田中"
        # Apply borders to merged region
        for r in range(row, row + 4):
            for cc in range(col_start, col_end + 1):
                ws.cell(row=r, column=cc).border = BORDER_ALL_THIN

    # --- Key-value section ---
    kv_start = 9
    kv_data = [
        ("件　名", "社内業務システム刷新プロジェクトに伴うクラウドサービス契約について"),
        ("起案部署", "情報システム部　基盤グループ"),
        ("起 案 者", "田中　太郎"),
        ("決裁区分", "A区分（1,000万円以上）"),
    ]
    for i, (key, val) in enumerate(kv_data):
        r = kv_start + i
        ws.row_dimensions[r].height = 22
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.merge_cells(start_row=r, start_column=7, end_row=r, end_column=40)
        kc = ws.cell(row=r, column=1)
        kc.value = key
        _apply_style(kc, font=FONT_BOLD, fill=FILL_LIGHT_BLUE, border=BORDER_ALL_THIN, alignment=ALIGN_CENTER)
        vc = ws.cell(row=r, column=7)
        vc.value = val
        _apply_style(vc, font=FONT_BODY, border=BORDER_ALL_THIN, alignment=ALIGN_LEFT)
        for cc in range(1, 41):
            ws.cell(row=r, column=cc).border = BORDER_ALL_THIN

    # --- 稟議内容 (body) ---
    body_start = 14
    ws.merge_cells(start_row=body_start, start_column=1, end_row=body_start, end_column=40)
    hc = ws.cell(row=body_start, column=1)
    hc.value = "稟議内容"
    _apply_style(hc, font=FONT_HEADING, fill=FILL_LIGHT_BLUE, border=BORDER_ALL_THIN, alignment=ALIGN_CENTER)
    for cc in range(1, 41):
        ws.cell(row=body_start, column=cc).border = BORDER_ALL_THIN

    body_text = (
        "1. 目的\n"
        "　現行の社内業務システム（2018年導入）は保守期限を2027年3月に迎えるため、"
        "クラウドベースの新システムへの移行を計画しています。\n\n"
        "2. 概要\n"
        "　AWS上にマイクロサービスアーキテクチャで構築された新システムを導入し、"
        "業務効率の30%向上とランニングコストの20%削減を目指します。\n\n"
        "3. 費用\n"
        "　初期構築費用：15,000,000円（税別）\n"
        "　年間運用費用：3,600,000円（税別）\n"
        "　合計（5年間）：33,000,000円（税別）\n\n"
        "4. スケジュール\n"
        "　2026年4月　要件定義開始\n"
        "　2026年7月　基本設計完了\n"
        "　2026年12月　開発完了・テスト開始\n"
        "　2027年3月　本番稼働"
    )
    ws.merge_cells(start_row=body_start + 1, start_column=1, end_row=body_start + 18, end_column=40)
    bc = ws.cell(row=body_start + 1, column=1)
    bc.value = body_text
    _apply_style(bc, font=FONT_BODY, alignment=ALIGN_LEFT_TOP)
    _apply_border_range(ws, body_start + 1, body_start + 18, 1, 40)

    # --- Cost breakdown table ---
    tbl_start = 34
    ws.merge_cells(start_row=tbl_start, start_column=1, end_row=tbl_start, end_column=40)
    ws.cell(row=tbl_start, column=1).value = "費用内訳"
    _apply_style(ws.cell(row=tbl_start, column=1), font=FONT_HEADING, fill=FILL_LIGHT_BLUE,
                 border=BORDER_ALL_THIN, alignment=ALIGN_CENTER)

    headers = ["No.", "項目", "数量", "単価（円）", "金額（円）", "備考"]
    col_ranges = [(1, 3), (4, 16), (17, 20), (21, 27), (28, 34), (35, 40)]
    tbl_h = tbl_start + 1
    for (cs, ce), h in zip(col_ranges, headers):
        ws.merge_cells(start_row=tbl_h, start_column=cs, end_row=tbl_h, end_column=ce)
        hc = ws.cell(row=tbl_h, column=cs)
        hc.value = h
        _apply_style(hc, font=FONT_WHITE_BOLD, fill=FILL_DARK_BLUE, border=BORDER_ALL_THIN, alignment=ALIGN_CENTER)

    rows_data = [
        ("1", "クラウド基盤構築", "1式", "5,000,000", "5,000,000", "AWS環境"),
        ("2", "アプリケーション開発", "1式", "8,000,000", "8,000,000", ""),
        ("3", "データ移行", "1式", "1,500,000", "1,500,000", ""),
        ("4", "テスト・検証", "1式", "500,000", "500,000", ""),
        ("", "合計", "", "", "15,000,000", ""),
    ]
    for i, row_data in enumerate(rows_data):
        r = tbl_h + 1 + i
        for (cs, ce), val in zip(col_ranges, row_data):
            ws.merge_cells(start_row=r, start_column=cs, end_row=r, end_column=ce)
            c = ws.cell(row=r, column=cs)
            c.value = val
            is_total = (i == len(rows_data) - 1)
            _apply_style(c,
                         font=FONT_BOLD if is_total else FONT_BODY,
                         fill=FILL_LIGHT_YELLOW if is_total else None,
                         border=BORDER_ALL_THIN,
                         alignment=ALIGN_CENTER if cs <= 3 else ALIGN_LEFT)
        _apply_border_range(ws, r, r, 1, 40)

    wb.save(OUTPUT_DIR / "sample_houganshi.xlsx")


# ---------------------------------------------------------------------------
# 2. レビュー資料 (Review document)
# ---------------------------------------------------------------------------

def generate_review():
    wb = Workbook()

    # --- Sheet 1: 概要 ---
    ws1 = wb.active
    ws1.title = "概要"
    ws1.column_dimensions["A"].width = 18
    ws1.column_dimensions["B"].width = 50
    ws1.column_dimensions["C"].width = 20
    ws1.column_dimensions["D"].width = 20

    # Start at row 3 (not A1) — realistic: margin at top
    ws1.merge_cells("A3:D3")
    ws1["A3"].value = "設計レビュー資料"
    _apply_style(ws1["A3"], font=FONT_TITLE, alignment=ALIGN_CENTER)
    ws1.row_dimensions[3].height = 30

    kv = [
        ("プロジェクト名", "顧客管理システム刷新"),
        ("対象工程", "基本設計"),
        ("対象成果物", "BD-010 画面設計書（顧客検索）"),
        ("レビュー日時", "2026年3月8日 14:00-16:00"),
        ("レビュー形式", "対面レビュー（会議室A）"),
        ("レビューア", "佐藤主任、鈴木課長"),
        ("作成者", "高橋　一郎"),
        ("版数", "1.2"),
    ]
    for i, (k, v) in enumerate(kv):
        r = i + 5
        ws1.cell(row=r, column=1).value = k
        _apply_style(ws1.cell(row=r, column=1), font=FONT_BOLD, fill=FILL_LIGHT_BLUE,
                     border=BORDER_ALL_THIN, alignment=ALIGN_CENTER)
        ws1.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        ws1.cell(row=r, column=2).value = v
        _apply_style(ws1.cell(row=r, column=2), font=FONT_BODY, border=BORDER_ALL_THIN, alignment=ALIGN_LEFT)

    r = 14
    ws1.merge_cells(f"A{r}:D{r}")
    ws1.cell(row=r, column=1).value = "レビュー概要"
    _apply_style(ws1.cell(row=r, column=1), font=FONT_HEADING, fill=FILL_LIGHT_GRAY, alignment=ALIGN_CENTER)

    ws1.merge_cells(f"A{r+1}:D{r+4}")
    ws1.cell(row=r + 1, column=1).value = (
        "画面設計書（顧客検索）について基本設計レビューを実施した。\n"
        "検索条件の網羅性、画面遷移の妥当性、非機能要件への対応について重点的に確認を行った。\n"
        "指摘事項は全8件（重大2件、軽微4件、指摘なし改善提案2件）。"
    )
    _apply_style(ws1.cell(row=r + 1, column=1), font=FONT_BODY, alignment=ALIGN_LEFT_TOP,
                 border=BORDER_ALL_THIN)

    # --- Sheet 2: レビュー指摘一覧 ---
    ws2 = wb.create_sheet("レビュー指摘一覧")
    cols = ["No.", "分類", "重要度", "指摘内容", "該当箇所", "対応方針", "ステータス", "対応者"]
    widths = [5, 10, 8, 40, 15, 35, 10, 10]
    for i, (col_name, w) in enumerate(zip(cols, widths)):
        c = i + 1
        ws2.column_dimensions[get_column_letter(c)].width = w
        cell = ws2.cell(row=3, column=c)
        cell.value = col_name
        _apply_style(cell, font=FONT_WHITE_BOLD, fill=FILL_DARK_BLUE, border=BORDER_ALL_THIN, alignment=ALIGN_CENTER)

    issues = [
        (1, "設計", "重大", "検索条件に「顧客ステータス」が含まれていない。業務上必須の絞り込み条件である。",
         "3.2 検索条件", "検索条件に顧客ステータス（活性/非活性/休止）を追加する", "対応済", "高橋"),
        (2, "設計", "重大", "検索結果が1000件を超えた場合のページネーション仕様が未定義。",
         "3.3 検索結果", "20件/ページのページネーションを追加。ソート機能も併せて設計する", "対応中", "高橋"),
        (3, "画面", "軽微", "検索ボタンの配置がガイドラインと異なる（左寄せになっている）。",
         "3.1 画面レイアウト", "右寄せに修正する", "対応済", "高橋"),
        (4, "画面", "軽微", "項目名「TEL」は「電話番号」に統一すべき。",
         "3.2 検索条件", "「電話番号」に統一する", "対応済", "高橋"),
        (5, "仕様", "軽微", "あいまい検索の仕様（前方一致/部分一致）が明記されていない。",
         "3.2 検索条件", "各項目の検索方式を表に追記する", "対応中", "高橋"),
        (6, "仕様", "軽微", "CSVエクスポート時の文字コード指定がない。",
         "3.4 機能仕様", "UTF-8(BOM付き)を標準とし、Shift_JIS選択も可能とする", "未着手", "高橋"),
        (7, "改善", "提案", "検索条件の保存機能があると業務効率が上がる。",
         "3.2 検索条件", "Phase2で検討する旨を備考に記載", "対応済", "高橋"),
        (8, "改善", "提案", "よく使う検索パターンのプリセット機能の検討。",
         "3.2 検索条件", "Phase2検討事項として管理台帳に起票", "対応済", "高橋"),
    ]

    status_fills = {"対応済": FILL_PASS, "対応中": FILL_PENDING, "未着手": FILL_FAIL}
    severity_fills = {"重大": FILL_RED, "軽微": FILL_ORANGE, "提案": FILL_LIGHT_GREEN}

    for i, row_data in enumerate(issues):
        r = i + 4
        for j, val in enumerate(row_data):
            cell = ws2.cell(row=r, column=j + 1)
            cell.value = val
            _apply_style(cell, font=FONT_BODY, border=BORDER_ALL_THIN, alignment=ALIGN_LEFT if j >= 3 else ALIGN_CENTER)
        # Color status
        ws2.cell(row=r, column=7).fill = status_fills.get(row_data[6], FILL_WHITE)
        # Color severity
        ws2.cell(row=r, column=3).fill = severity_fills.get(row_data[2], FILL_WHITE)
        _apply_style(ws2.cell(row=r, column=3), font=Font(name="Yu Gothic", size=10, bold=True, color="FFFFFF"),
                     alignment=ALIGN_CENTER)

    # --- Sheet 3: 修正履歴 ---
    ws3 = wb.create_sheet("修正履歴")
    hist_cols = ["版数", "日付", "修正者", "修正内容"]
    hist_widths = [8, 15, 12, 60]
    for i, (col_name, w) in enumerate(zip(hist_cols, hist_widths)):
        c = i + 1
        ws3.column_dimensions[get_column_letter(c)].width = w
        cell = ws3.cell(row=1, column=c)
        cell.value = col_name
        _apply_style(cell, font=FONT_WHITE_BOLD, fill=FILL_DARK_BLUE, border=BORDER_ALL_THIN, alignment=ALIGN_CENTER)

    history = [
        ("1.0", "2026/02/20", "高橋", "初版作成"),
        ("1.1", "2026/03/01", "高橋", "検索条件の見直し（要件ヒアリング結果反映）"),
        ("1.2", "2026/03/09", "高橋", "レビュー指摘事項（No.1,3,4,7,8）の対応"),
    ]
    for i, row_data in enumerate(history):
        r = i + 2
        for j, val in enumerate(row_data):
            cell = ws3.cell(row=r, column=j + 1)
            cell.value = val
            _apply_style(cell, font=FONT_BODY, border=BORDER_ALL_THIN, alignment=ALIGN_CENTER if j < 3 else ALIGN_LEFT)

    wb.save(OUTPUT_DIR / "sample_review.xlsx")


# ---------------------------------------------------------------------------
# 3. 工程表 (Project schedule / Gantt-like)
# ---------------------------------------------------------------------------

def generate_schedule():
    wb = Workbook()
    ws = wb.active
    ws.title = "マスタスケジュール"

    # Fixed columns: A=No, B=大工程, C=中工程, D=担当, E=開始, F=終了, G=進捗
    fixed_cols = [("No.", 5), ("大工程", 14), ("中工程", 20), ("担当", 8), ("開始日", 11), ("終了日", 11), ("進捗", 7)]
    for i, (_, w) in enumerate(fixed_cols):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    # Date columns: 6 months from 2026-04 to 2026-09, weekly
    base_date = date(2026, 4, 1)
    date_cols = []  # (col_index, date_obj)
    col = 8
    for week in range(26):
        d = base_date + timedelta(weeks=week)
        date_cols.append((col, d))
        ws.column_dimensions[get_column_letter(col)].width = 3
        col += 1

    # --- Header row 1: Year ---
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    ws.cell(row=1, column=1).value = "マスタスケジュール　2026年度"
    _apply_style(ws.cell(row=1, column=1), font=FONT_TITLE, alignment=ALIGN_CENTER)

    # --- Header row 2: Month labels ---
    r2 = 2
    for i, (name, _) in enumerate(fixed_cols):
        c = ws.cell(row=r2, column=i + 1)
        _apply_style(c, fill=FILL_LIGHT_GRAY)

    month_groups = {}
    for col_idx, d in date_cols:
        key = d.month
        if key not in month_groups:
            month_groups[key] = []
        month_groups[key].append(col_idx)

    month_names = {4: "4月", 5: "5月", 6: "6月", 7: "7月", 8: "8月", 9: "9月", 10: "10月"}
    for m, cols_in_month in month_groups.items():
        ws.merge_cells(start_row=r2, start_column=cols_in_month[0], end_row=r2, end_column=cols_in_month[-1])
        c = ws.cell(row=r2, column=cols_in_month[0])
        c.value = month_names.get(m, f"{m}月")
        _apply_style(c, font=FONT_BOLD, fill=FILL_LIGHT_GRAY, border=BORDER_ALL_THIN, alignment=ALIGN_CENTER)

    # --- Header row 3: Week numbers ---
    r3 = 3
    for i, (name, _) in enumerate(fixed_cols):
        c = ws.cell(row=r3, column=i + 1)
        c.value = name
        _apply_style(c, font=FONT_WHITE_BOLD, fill=FILL_DARK_BLUE, border=BORDER_ALL_THIN, alignment=ALIGN_CENTER)

    for col_idx, d in date_cols:
        c = ws.cell(row=r3, column=col_idx)
        c.value = d.day
        _apply_style(c, font=FONT_SMALL, fill=FILL_LIGHT_BLUE, border=BORDER_ALL_THIN, alignment=ALIGN_CENTER)

    # --- Task data ---
    tasks = [
        # (no, phase, task, owner, start_week, end_week, progress, color, is_summary)
        ("1", "要件定義", "", "全員", 0, 7, "100%", FILL_GANTT_BLUE, True),
        ("1.1", "", "業務要件整理", "佐藤", 0, 3, "100%", FILL_GANTT_BLUE, False),
        ("1.2", "", "システム要件定義", "田中", 2, 5, "100%", FILL_GANTT_BLUE, False),
        ("1.3", "", "要件レビュー・承認", "鈴木", 5, 7, "100%", FILL_GANTT_BLUE, False),
        ("2", "基本設計", "", "全員", 6, 14, "60%", FILL_GANTT_GREEN, True),
        ("2.1", "", "画面設計", "高橋", 6, 10, "80%", FILL_GANTT_GREEN, False),
        ("2.2", "", "DB設計", "田中", 7, 11, "50%", FILL_GANTT_GREEN, False),
        ("2.3", "", "API設計", "佐藤", 8, 12, "40%", FILL_GANTT_GREEN, False),
        ("2.4", "", "設計レビュー", "鈴木", 12, 14, "0%", FILL_GANTT_GREEN, False),
        ("3", "詳細設計", "", "全員", 13, 18, "0%", FILL_GANTT_ORANGE, True),
        ("3.1", "", "詳細設計書作成", "高橋", 13, 16, "0%", FILL_GANTT_ORANGE, False),
        ("3.2", "", "詳細設計レビュー", "鈴木", 16, 18, "0%", FILL_GANTT_ORANGE, False),
        ("4", "開発", "", "全員", 17, 24, "0%", FILL_GANTT_BLUE, True),
        ("4.1", "", "フロントエンド開発", "高橋", 17, 22, "0%", FILL_GANTT_BLUE, False),
        ("4.2", "", "バックエンド開発", "田中", 17, 22, "0%", FILL_GANTT_BLUE, False),
        ("4.3", "", "単体テスト", "全員", 20, 24, "0%", FILL_GANTT_BLUE, False),
    ]

    for i, (no, phase, task, owner, sw, ew, progress, color, is_summary) in enumerate(tasks):
        r = 4 + i
        ws.cell(row=r, column=1).value = no
        ws.cell(row=r, column=2).value = phase
        ws.cell(row=r, column=3).value = task
        ws.cell(row=r, column=4).value = owner
        ws.cell(row=r, column=5).value = (base_date + timedelta(weeks=sw)).strftime("%m/%d")
        ws.cell(row=r, column=6).value = (base_date + timedelta(weeks=ew)).strftime("%m/%d")
        ws.cell(row=r, column=7).value = progress

        row_font = FONT_BOLD if is_summary else FONT_BODY
        row_fill = FILL_LIGHT_YELLOW if is_summary else None
        for c in range(1, 8):
            cell = ws.cell(row=r, column=c)
            _apply_style(cell, font=row_font, fill=row_fill, border=BORDER_ALL_THIN,
                         alignment=ALIGN_CENTER if c != 3 else ALIGN_LEFT)

        # Gantt bars
        for col_idx, d in date_cols:
            week_num = (d - base_date).days // 7
            cell = ws.cell(row=r, column=col_idx)
            cell.border = Border(left=HAIR, right=HAIR, top=HAIR, bottom=HAIR)
            if sw <= week_num <= ew:
                cell.fill = color

    wb.save(OUTPUT_DIR / "sample_schedule.xlsx")


# ---------------------------------------------------------------------------
# 4. 仕様書 (Specification document)
# ---------------------------------------------------------------------------

def generate_spec():
    wb = Workbook()
    ws = wb.active
    ws.title = "仕様書"

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 60
    ws.column_dimensions["E"].width = 20

    # Start at row 3 (not A1) — realistic: margin rows at top
    # Title
    ws.merge_cells("A3:E3")
    ws["A3"].value = "顧客管理システム　機能仕様書"
    _apply_style(ws["A3"], font=FONT_TITLE, alignment=ALIGN_CENTER)
    ws.row_dimensions[3].height = 36

    # Document info
    info = [
        ("文書番号", "FS-CRM-001"),
        ("版数", "2.0"),
        ("作成日", "2026年2月15日"),
        ("最終更新日", "2026年3月8日"),
        ("作成者", "システム開発部　田中太郎"),
        ("承認者", "システム開発部長　山田次郎"),
    ]
    for i, (k, v) in enumerate(info):
        r = 5 + i
        ws.cell(row=r, column=1).value = k
        _apply_style(ws.cell(row=r, column=1), font=FONT_BOLD, fill=FILL_LIGHT_GRAY,
                     border=BORDER_ALL_THIN, alignment=ALIGN_CENTER)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
        ws.cell(row=r, column=3).value = v
        _apply_style(ws.cell(row=r, column=3), font=FONT_BODY, border=BORDER_ALL_THIN, alignment=ALIGN_LEFT)

    # Version history table
    r = 12
    ws.merge_cells(f"A{r}:E{r}")
    ws.cell(row=r, column=1).value = "改訂履歴"
    _apply_style(ws.cell(row=r, column=1), font=FONT_HEADING, fill=FILL_LIGHT_BLUE, alignment=ALIGN_CENTER)

    vh_headers = ["版数", "日付", "変更者", "変更内容"]
    vh_cols = [(1, 1), (2, 2), (3, 3), (4, 5)]
    r += 1
    for (cs, ce), h in zip(vh_cols, vh_headers):
        ws.merge_cells(start_row=r, start_column=cs, end_row=r, end_column=ce)
        c = ws.cell(row=r, column=cs)
        c.value = h
        _apply_style(c, font=FONT_WHITE_BOLD, fill=FILL_DARK_BLUE, border=BORDER_ALL_THIN, alignment=ALIGN_CENTER)

    vh_data = [
        ("1.0", "2026/02/15", "田中", "初版作成"),
        ("1.1", "2026/02/28", "田中", "レビュー指摘反映（検索機能仕様の詳細化）"),
        ("2.0", "2026/03/08", "田中", "顧客ステータス管理機能を追加"),
    ]
    for i, row_data in enumerate(vh_data):
        rr = r + 1 + i
        for (cs, ce), val in zip(vh_cols, row_data):
            ws.merge_cells(start_row=rr, start_column=cs, end_row=rr, end_column=ce)
            c = ws.cell(row=rr, column=cs)
            c.value = val
            _apply_style(c, font=FONT_BODY, border=BORDER_ALL_THIN,
                         alignment=ALIGN_CENTER if cs < 4 else ALIGN_LEFT)

    # --- Specification body ---
    sections = [
        ("1", "章", "概要", None),
        ("1.1", "節", "目的", "本仕様書は、顧客管理システム（CRM）の機能仕様を定義する。\n"
         "対象読者はシステム開発者、テスト担当者、および運用担当者である。"),
        ("1.2", "節", "適用範囲", "本仕様書はCRMシステムの以下の機能に適用する。\n"
         "・顧客情報管理（登録・更新・削除・検索）\n"
         "・顧客ステータス管理\n"
         "・帳票出力\n"
         "※ バッチ処理についてはBS-CRM-001を参照のこと。"),
        ("2", "章", "機能仕様", None),
        ("2.1", "節", "顧客検索機能", None),
        ("2.1.1", "項", "検索条件",
         "以下の条件による検索が可能であること。\n"
         "① 顧客番号（完全一致）\n"
         "② 顧客名（部分一致）\n"
         "③ 電話番号（前方一致）\n"
         "④ 顧客ステータス（選択式）\n"
         "⑤ 登録日（範囲指定）"),
        ("2.1.2", "項", "検索結果表示",
         "検索結果は一覧表形式で表示する。\n"
         "1ページあたりの表示件数は20件とし、ページネーション機能を提供する。\n"
         "表示項目：顧客番号、顧客名、電話番号、ステータス、登録日、更新日"),
        ("2.2", "節", "顧客登録機能", None),
        ("2.2.1", "項", "入力項目",
         "必須項目：顧客名、顧客名カナ、電話番号、住所\n"
         "任意項目：FAX番号、メールアドレス、担当者名、備考\n"
         "自動設定：顧客番号（採番ルールはBS-CRM-001 3.2節参照）、登録日時、ステータス（初期値：活性）"),
        ("3", "章", "非機能要件", None),
        ("3.1", "節", "性能要件",
         "・検索応答時間：3秒以内（1000万件時）\n"
         "・同時接続数：100ユーザー\n"
         "・画面表示時間：2秒以内"),
    ]

    current_row = 18
    for num, level, title, body in sections:
        if level == "章":
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
            c = ws.cell(row=current_row, column=1)
            c.value = f"第{num}章　{title}"
            _apply_style(c, font=Font(name="Yu Gothic", size=14, bold=True), fill=FILL_LIGHT_BLUE,
                         border=BORDER_BOTTOM_MEDIUM, alignment=ALIGN_LEFT)
            ws.row_dimensions[current_row].height = 28
            current_row += 1
        elif level == "節":
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
            c = ws.cell(row=current_row, column=1)
            c.value = f"　{num}　{title}"
            _apply_style(c, font=FONT_HEADING, border=BORDER_BOTTOM_THIN, alignment=ALIGN_LEFT)
            ws.row_dimensions[current_row].height = 24
            current_row += 1
        elif level == "項":
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
            c = ws.cell(row=current_row, column=1)
            c.value = f"　　{num}　{title}"
            _apply_style(c, font=FONT_SUBHEADING, alignment=ALIGN_LEFT)
            current_row += 1

        if body:
            line_count = body.count("\n") + 2
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row + line_count - 1, end_column=5)
            c = ws.cell(row=current_row, column=1)
            c.value = body
            _apply_style(c, font=FONT_BODY, alignment=ALIGN_LEFT_TOP)
            current_row += line_count

        current_row += 1  # spacing

    wb.save(OUTPUT_DIR / "sample_spec.xlsx")


# ---------------------------------------------------------------------------
# 5. 管理台帳 (Management ledger)
# ---------------------------------------------------------------------------

def generate_ledger():
    wb = Workbook()
    ws = wb.active
    ws.title = "課題管理台帳"

    # Start at B3 (not A1) — realistic: margin rows/cols
    CO = 1  # column offset
    RO = 2  # row offset

    # Title
    ws.merge_cells(start_row=1 + RO, start_column=1 + CO, end_row=1 + RO, end_column=14 + CO)
    c = ws.cell(row=1 + RO, column=1 + CO)
    c.value = "課題管理台帳"
    _apply_style(c, font=FONT_TITLE, alignment=ALIGN_CENTER)
    ws.row_dimensions[1 + RO].height = 30

    ws.merge_cells(start_row=2 + RO, start_column=11 + CO, end_row=2 + RO, end_column=14 + CO)
    c = ws.cell(row=2 + RO, column=11 + CO)
    c.value = "最終更新：2026年3月10日"
    _apply_style(c, font=FONT_SMALL, alignment=ALIGN_RIGHT)

    # Category headers (row 3 + RO = 5)
    cat_headers = [
        (1, 1, "基本情報"),
        (2, 5, "課題詳細"),
        (6, 8, "対応情報"),
        (9, 11, "スケジュール"),
        (12, 14, "管理"),
    ]
    for cs, ce, label in cat_headers:
        ws.merge_cells(start_row=3 + RO, start_column=cs + CO, end_row=3 + RO, end_column=ce + CO)
        c = ws.cell(row=3 + RO, column=cs + CO)
        c.value = label
        _apply_style(c, font=FONT_WHITE_BOLD, fill=FILL_DARK_BLUE, border=BORDER_ALL_THIN, alignment=ALIGN_CENTER)

    # Sub-headers (row 4 + RO = 6)
    sub_headers = [
        "No.", "カテゴリ", "タイトル", "詳細", "影響度",
        "対応者", "対応内容", "ステータス",
        "起票日", "期限", "完了日",
        "優先度", "関連チケット", "備考",
    ]
    widths = [5, 10, 25, 40, 8, 10, 35, 10, 11, 11, 11, 8, 12, 20]
    for i, (h, w) in enumerate(zip(sub_headers, widths)):
        c = i + 1 + CO
        ws.column_dimensions[get_column_letter(c)].width = w
        cell = ws.cell(row=4 + RO, column=c)
        cell.value = h
        _apply_style(cell, font=FONT_BOLD, fill=FILL_LIGHT_BLUE, border=BORDER_ALL_THIN, alignment=ALIGN_CENTER)

    # Data
    data = [
        (1, "機能", "検索画面のレスポンス改善", "顧客数が10万件を超えるとタイムアウトする", "高",
         "田中", "インデックス追加とクエリ最適化", "対応済",
         "02/15", "03/01", "02/28", "高", "JIRA-1234", ""),
        (2, "機能", "CSV出力の文字化け", "Shift_JISで出力するとUnicode文字が化ける", "中",
         "佐藤", "UTF-8(BOM付き)に変更", "対応済",
         "02/18", "03/05", "03/03", "高", "JIRA-1235", ""),
        (3, "画面", "入力フォームのバリデーション追加", "電話番号・メールアドレスの形式チェックが不足", "中",
         "高橋", "正規表現によるバリデーション追加", "対応中",
         "02/20", "03/10", "", "中", "JIRA-1240", "フロント側のみ先行対応"),
        (4, "性能", "同時接続100ユーザー時の性能確認", "負荷テストで80ユーザー超で遅延発生", "高",
         "田中", "コネクションプール設定の調整", "対応中",
         "02/25", "03/15", "", "高", "JIRA-1250", ""),
        (5, "環境", "ステージング環境の構築", "本番同等のステージング環境が未整備", "中",
         "鈴木", "Terraformで環境構築", "対応中",
         "03/01", "03/20", "", "中", "", ""),
        (6, "仕様", "顧客ステータスの遷移ルール", "ステータス遷移の業務ルールが未確定", "高",
         "山田", "業務部門とルール策定中", "検討中",
         "03/03", "03/15", "", "高", "", "3/12に業務部門MTG予定"),
        (7, "機能", "帳票出力のレイアウト調整", "印刷時に右端が切れる", "低",
         "高橋", "", "未着手",
         "03/05", "03/25", "", "低", "JIRA-1260", ""),
        (8, "セキュリティ", "SQLインジェクション対策の確認", "外部診断で指摘あり", "高",
         "田中", "プリペアドステートメントへの全面切替", "対応中",
         "03/08", "03/12", "", "緊急", "JIRA-1265", "脆弱性診断報告書参照"),
    ]

    status_fills = {"対応済": FILL_PASS, "対応中": FILL_PENDING, "未着手": FILL_LIGHT_GRAY, "検討中": FILL_LIGHT_YELLOW}
    priority_fills = {"緊急": FILL_RED, "高": FILL_ORANGE, "中": FILL_LIGHT_YELLOW, "低": FILL_LIGHT_GREEN}

    for i, row_data in enumerate(data):
        r = 5 + RO + i
        for j, val in enumerate(row_data):
            cell = ws.cell(row=r, column=j + 1 + CO)
            cell.value = val
            _apply_style(cell, font=FONT_BODY, border=BORDER_ALL_THIN,
                         alignment=ALIGN_CENTER if j in (0, 4, 7, 11) else ALIGN_LEFT)
        ws.cell(row=r, column=8 + CO).fill = status_fills.get(row_data[7], FILL_WHITE)
        ws.cell(row=r, column=12 + CO).fill = priority_fills.get(row_data[11], FILL_WHITE)

    # Freeze panes
    ws.freeze_panes = ws.cell(row=5 + RO, column=3 + CO).coordinate

    # Auto-filter
    fc = get_column_letter(1 + CO)
    lc = get_column_letter(14 + CO)
    ws.auto_filter.ref = f"{fc}{4 + RO}:{lc}{12 + RO}"

    # Data validation for status column
    dv = DataValidation(type="list", formula1='"未着手,検討中,対応中,対応済,保留"', allow_blank=True)
    dv.error = "リストから選択してください"
    dv.errorTitle = "入力エラー"
    ws.add_data_validation(dv)
    sc = get_column_letter(8 + CO)
    dv.add(f"{sc}{5 + RO}:{sc}100")

    # Data validation for priority
    dv2 = DataValidation(type="list", formula1='"緊急,高,中,低"', allow_blank=True)
    ws.add_data_validation(dv2)
    pc = get_column_letter(12 + CO)
    dv2.add(f"{pc}{5 + RO}:{pc}100")

    wb.save(OUTPUT_DIR / "sample_ledger.xlsx")


# ---------------------------------------------------------------------------
# 6. 議事録 (Meeting minutes)
# ---------------------------------------------------------------------------

def generate_minutes():
    wb = Workbook()
    ws = wb.active
    ws.title = "議事録"

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15

    # Start at row 3 (not A1) — realistic: margin rows at top
    # Title
    ws.merge_cells("A3:D3")
    ws["A3"].value = "議 事 録"
    _apply_style(ws["A3"], font=FONT_TITLE, alignment=ALIGN_CENTER)
    ws.row_dimensions[3].height = 36

    # Key-value header
    kv = [
        ("会議名", "顧客管理システム刷新　第12回定例会議"),
        ("日時", "2026年3月9日（月）14:00〜16:00"),
        ("場所", "本社5階　会議室A（Zoom併用）"),
        ("参加者", "【システム開発部】山田部長、佐藤課長、田中、高橋、鈴木\n【業務部】中村課長、伊藤\n【外部】ABC株式会社　木村PM、渡辺SE"),
        ("記録者", "鈴木　花子"),
        ("配布先", "参加者全員、品質管理部　小林"),
    ]
    r = 5
    for k, v in kv:
        ws.cell(row=r, column=1).value = k
        _apply_style(ws.cell(row=r, column=1), font=FONT_BOLD, fill=FILL_LIGHT_BLUE,
                     border=BORDER_ALL_THIN, alignment=ALIGN_CENTER)
        line_count = max(1, v.count("\n") + 1)
        if line_count > 1:
            ws.merge_cells(start_row=r, start_column=2, end_row=r + line_count - 1, end_column=4)
            ws.merge_cells(start_row=r, start_column=1, end_row=r + line_count - 1, end_column=1)
            ws.row_dimensions[r].height = 20 * line_count
        else:
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        c = ws.cell(row=r, column=2)
        c.value = v
        _apply_style(c, font=FONT_BODY, border=BORDER_ALL_THIN, alignment=ALIGN_LEFT_TOP)
        # Borders for merged area
        for rr in range(r, r + line_count):
            for cc in range(1, 5):
                ws.cell(row=rr, column=cc).border = BORDER_ALL_THIN
        r += line_count

    # Agenda
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    ws.cell(row=r, column=1).value = "アジェンダ"
    _apply_style(ws.cell(row=r, column=1), font=FONT_HEADING, fill=FILL_LIGHT_GRAY, alignment=ALIGN_CENTER)
    r += 1

    agenda_items = [
        "1. 前回アクションアイテムの確認",
        "2. 開発進捗報告",
        "3. 課題・リスクの共有",
        "4. 次フェーズのスケジュール確認",
        "5. その他",
    ]
    ws.merge_cells(start_row=r, start_column=1, end_row=r + len(agenda_items) - 1, end_column=4)
    ws.cell(row=r, column=1).value = "\n".join(agenda_items)
    _apply_style(ws.cell(row=r, column=1), font=FONT_BODY, alignment=ALIGN_LEFT_TOP, border=BORDER_ALL_THIN)
    r += len(agenda_items)

    # Discussion sections
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    ws.cell(row=r, column=1).value = "議事内容"
    _apply_style(ws.cell(row=r, column=1), font=FONT_HEADING, fill=FILL_LIGHT_GRAY, alignment=ALIGN_CENTER)
    r += 1

    discussions = [
        ("1. 前回アクションアイテムの確認",
         "・AI-011（田中）：検索画面のインデックス追加 → 完了。レスポンスが5秒→1秒に改善。\n"
         "・AI-012（佐藤）：CSV出力の文字化け修正 → 完了。UTF-8(BOM付き)で対応。\n"
         "・AI-013（高橋）：バリデーション仕様の確定 → 対応中。3/10完了予定。"),
        ("2. 開発進捗報告",
         "【田中報告】\n"
         "・バックエンドAPI開発：計画比95%。残りは顧客ステータス更新APIのみ。\n"
         "・パフォーマンスチューニング：コネクションプール設定の見直しで改善傾向。\n\n"
         "【高橋報告】\n"
         "・フロントエンド開発：計画比85%。検索画面・登録画面は完了。\n"
         "・入力バリデーションの実装が残タスク。\n\n"
         "【木村PM（ABC社）】\n"
         "・結合テスト準備を並行で進めている。テストケースは80%作成済み。"),
        ("3. 課題・リスクの共有",
         "・（中村）顧客ステータスの遷移ルールについて、業務部門内で意見が分かれている。\n"
         "　→ 3/12に業務部門内で最終決定会議を実施予定。\n"
         "・（田中）外部脆弱性診断でSQLインジェクションの指摘あり。緊急対応中。\n"
         "　→ 3/12までに全面的にプリペアドステートメントへ切替完了見込み。\n"
         "・（山田部長）スケジュールへの影響を注視。遅延の場合は早めにエスカレーションすること。"),
        ("4. 次フェーズのスケジュール確認",
         "・基本設計レビュー完了目標：3月末\n"
         "・詳細設計開始：4月第1週\n"
         "・結合テスト開始：6月第2週（予定通り）"),
    ]

    for title, content in discussions:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        ws.cell(row=r, column=1).value = title
        _apply_style(ws.cell(row=r, column=1), font=FONT_SUBHEADING, fill=FILL_LIGHT_YELLOW, alignment=ALIGN_LEFT)
        r += 1
        line_count = content.count("\n") + 2
        ws.merge_cells(start_row=r, start_column=1, end_row=r + line_count - 1, end_column=4)
        ws.cell(row=r, column=1).value = content
        _apply_style(ws.cell(row=r, column=1), font=FONT_BODY, alignment=ALIGN_LEFT_TOP, border=BORDER_ALL_THIN)
        r += line_count + 1

    # Action items table
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    ws.cell(row=r, column=1).value = "アクションアイテム"
    _apply_style(ws.cell(row=r, column=1), font=FONT_HEADING, fill=FILL_LIGHT_GRAY, alignment=ALIGN_CENTER)
    r += 1

    ai_headers = ["AI番号", "内容", "担当", "期限"]
    for i, h in enumerate(ai_headers):
        c = ws.cell(row=r, column=i + 1)
        c.value = h
        _apply_style(c, font=FONT_WHITE_BOLD, fill=FILL_DARK_BLUE, border=BORDER_ALL_THIN, alignment=ALIGN_CENTER)
    r += 1

    action_items = [
        ("AI-014", "バリデーション仕様の確定と実装", "高橋", "3/13"),
        ("AI-015", "SQLインジェクション対策の全面切替", "田中", "3/12"),
        ("AI-016", "顧客ステータス遷移ルールの業務部門決定結果の共有", "中村", "3/13"),
        ("AI-017", "ステージング環境の構築完了", "鈴木", "3/20"),
        ("AI-018", "結合テストケースのレビュー依頼", "木村", "3/17"),
    ]
    for ai in action_items:
        for j, val in enumerate(ai):
            c = ws.cell(row=r, column=j + 1)
            c.value = val
            _apply_style(c, font=FONT_BODY, border=BORDER_ALL_THIN,
                         alignment=ALIGN_CENTER if j != 1 else ALIGN_LEFT)
        r += 1

    # Next meeting
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    ws.cell(row=r, column=1).value = "次回会議"
    _apply_style(ws.cell(row=r, column=1), font=FONT_HEADING, fill=FILL_LIGHT_GRAY, alignment=ALIGN_CENTER)
    r += 1
    next_info = [
        ("日時", "2026年3月16日（月）14:00〜16:00"),
        ("場所", "本社5階　会議室A（Zoom併用）"),
        ("議題", "開発進捗報告、課題対応状況確認、結合テスト計画レビュー"),
    ]
    for k, v in next_info:
        ws.cell(row=r, column=1).value = k
        _apply_style(ws.cell(row=r, column=1), font=FONT_BOLD, fill=FILL_LIGHT_BLUE,
                     border=BORDER_ALL_THIN, alignment=ALIGN_CENTER)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        ws.cell(row=r, column=2).value = v
        _apply_style(ws.cell(row=r, column=2), font=FONT_BODY, border=BORDER_ALL_THIN, alignment=ALIGN_LEFT)
        r += 1

    wb.save(OUTPUT_DIR / "sample_minutes.xlsx")


# ---------------------------------------------------------------------------
# 7. テスト仕様書 (Test specification)
# ---------------------------------------------------------------------------

def generate_test_spec():
    wb = Workbook()

    # --- Sheet 1: 単体テスト ---
    ws1 = wb.active
    ws1.title = "単体テスト"

    headers = ["テストID", "大分類", "中分類", "テスト項目", "テスト手順", "期待結果", "結果", "実施日", "実施者", "備考"]
    widths = [10, 12, 14, 30, 35, 30, 7, 11, 8, 20]
    for i, (h, w) in enumerate(zip(headers, widths)):
        c = i + 1
        ws1.column_dimensions[get_column_letter(c)].width = w
        cell = ws1.cell(row=1, column=c)
        cell.value = h
        _apply_style(cell, font=FONT_WHITE_BOLD, fill=FILL_DARK_BLUE, border=BORDER_ALL_THIN, alignment=ALIGN_CENTER)

    test_data = [
        # (id, cat1, cat2, item, procedure, expected, result, date, person, note)
        ("UT-001", "顧客検索", "検索条件", "顧客番号で完全一致検索", "顧客番号「C00001」を入力し検索ボタン押下",
         "顧客番号C00001の顧客情報が1件表示される", "OK", "03/05", "高橋", ""),
        ("UT-002", "顧客検索", "検索条件", "顧客名で部分一致検索", "顧客名「田中」を入力し検索ボタン押下",
         "顧客名に「田中」を含むレコードが全件表示される", "OK", "03/05", "高橋", ""),
        ("UT-003", "顧客検索", "検索条件", "電話番号で前方一致検索", "電話番号「03-」を入力し検索ボタン押下",
         "電話番号が03-で始まるレコードが全件表示される", "OK", "03/05", "高橋", ""),
        ("UT-004", "顧客検索", "検索条件", "ステータスで絞り込み検索", "ステータス「活性」を選択し検索ボタン押下",
         "ステータスが活性のレコードが全件表示される", "OK", "03/06", "高橋", ""),
        ("UT-005", "顧客検索", "検索条件", "複合条件での検索", "顧客名「株式会社」かつステータス「活性」で検索",
         "両条件を満たすレコードのみ表示される", "NG", "03/06", "高橋", "AND条件が正しく動作しない→JIRA-1270"),
        ("UT-006", "顧客検索", "検索結果", "ページネーション（次ページ）", "検索結果が30件の状態で「次へ」押下",
         "21〜30件目が表示される", "OK", "03/06", "高橋", ""),
        ("UT-007", "顧客検索", "検索結果", "ソート（昇順）", "顧客番号カラムヘッダをクリック",
         "顧客番号の昇順で表示される", "OK", "03/06", "高橋", ""),
        ("UT-008", "顧客検索", "検索結果", "検索結果0件", "存在しない顧客番号「Z99999」で検索",
         "「該当するデータがありません」メッセージ表示", "OK", "03/06", "高橋", ""),
        ("UT-009", "顧客登録", "入力チェック", "必須項目未入力", "顧客名を空にして登録ボタン押下",
         "「顧客名は必須です」エラーメッセージ表示", "OK", "03/07", "田中", ""),
        ("UT-010", "顧客登録", "入力チェック", "電話番号形式チェック", "電話番号「abc」を入力して登録",
         "「電話番号の形式が正しくありません」エラー表示", "NG", "03/07", "田中", "バリデーション未実装→AI-014"),
        ("UT-011", "顧客登録", "入力チェック", "メールアドレス形式チェック", "メール「test」を入力して登録",
         "「メールアドレスの形式が正しくありません」エラー表示", "NG", "03/07", "田中", "バリデーション未実装→AI-014"),
        ("UT-012", "顧客登録", "正常登録", "全項目入力での登録", "全項目に正しい値を入力し登録",
         "登録成功メッセージ表示。一覧に反映される", "OK", "03/07", "田中", ""),
        ("UT-013", "顧客登録", "正常登録", "必須項目のみでの登録", "必須項目のみ入力し登録",
         "登録成功。任意項目はNULLで保存", "OK", "03/07", "田中", ""),
        ("UT-014", "CSV出力", "出力機能", "検索結果のCSV出力", "検索結果表示中にCSV出力ボタン押下",
         "UTF-8(BOM付き)のCSVがダウンロードされる", "OK", "03/08", "佐藤", ""),
        ("UT-015", "CSV出力", "出力機能", "0件時のCSV出力", "検索結果0件の状態でCSV出力ボタン押下",
         "「出力対象データがありません」メッセージ表示", "未実施", "—", "—", ""),
    ]

    # Merge cells for same category groups
    merge_ranges_cat1 = []
    merge_ranges_cat2 = []
    prev_cat1 = prev_cat2 = None
    start_cat1 = start_cat2 = 2

    for i, td in enumerate(test_data):
        r = i + 2
        if td[1] != prev_cat1:
            if prev_cat1 is not None and r - 1 > start_cat1:
                merge_ranges_cat1.append((start_cat1, r - 1))
            start_cat1 = r
            prev_cat1 = td[1]
        if td[2] != prev_cat2:
            if prev_cat2 is not None and r - 1 > start_cat2:
                merge_ranges_cat2.append((start_cat2, r - 1))
            start_cat2 = r
            prev_cat2 = td[2]
    # Final groups
    last_r = len(test_data) + 1
    if last_r > start_cat1:
        merge_ranges_cat1.append((start_cat1, last_r))
    if last_r > start_cat2:
        merge_ranges_cat2.append((start_cat2, last_r))

    # Write data first
    for i, td in enumerate(test_data):
        r = i + 2
        for j, val in enumerate(td):
            cell = ws1.cell(row=r, column=j + 1)
            cell.value = val
            _apply_style(cell, font=FONT_BODY, border=BORDER_ALL_THIN,
                         alignment=ALIGN_CENTER if j in (0, 6, 7, 8) else ALIGN_LEFT)
        # Color result column
        result = td[6]
        fill = {"OK": FILL_PASS, "NG": FILL_FAIL, "未実施": FILL_PENDING}.get(result)
        if fill:
            ws1.cell(row=r, column=7).fill = fill

    # Apply merges
    for start_r, end_r in merge_ranges_cat1:
        ws1.merge_cells(start_row=start_r, start_column=2, end_row=end_r, end_column=2)
    for start_r, end_r in merge_ranges_cat2:
        ws1.merge_cells(start_row=start_r, start_column=3, end_row=end_r, end_column=3)

    # --- Sheet 2: 結合テスト ---
    ws2 = wb.create_sheet("結合テスト")
    it_headers = ["テストID", "テストシナリオ", "前提条件", "手順", "期待結果", "結果", "実施日", "実施者"]
    it_widths = [10, 25, 25, 40, 30, 7, 11, 8]
    for i, (h, w) in enumerate(zip(it_headers, it_widths)):
        c = i + 1
        ws2.column_dimensions[get_column_letter(c)].width = w
        cell = ws2.cell(row=1, column=c)
        cell.value = h
        _apply_style(cell, font=FONT_WHITE_BOLD, fill=FILL_DARK_GREEN, border=BORDER_ALL_THIN, alignment=ALIGN_CENTER)

    it_data = [
        ("IT-001", "顧客登録→検索→表示", "テストDBに初期データ投入済み",
         "1.顧客登録画面で新規登録\n2.検索画面で登録した顧客名で検索\n3.検索結果から詳細表示",
         "登録内容が正しく検索・表示される", "未実施", "—", "—"),
        ("IT-002", "顧客登録→更新→履歴確認", "テストDBに初期データ投入済み",
         "1.顧客を新規登録\n2.登録した顧客の電話番号を更新\n3.更新履歴を確認",
         "更新内容と更新履歴が正しく反映される", "未実施", "—", "—"),
        ("IT-003", "検索→CSV出力→内容検証", "検索対象データ50件以上",
         "1.条件を指定して検索\n2.CSV出力\n3.CSVファイルの内容と画面表示を比較",
         "CSV内容が画面表示と一致する", "未実施", "—", "—"),
    ]
    for i, td in enumerate(it_data):
        r = i + 2
        for j, val in enumerate(td):
            cell = ws2.cell(row=r, column=j + 1)
            cell.value = val
            _apply_style(cell, font=FONT_BODY, border=BORDER_ALL_THIN,
                         alignment=ALIGN_CENTER if j in (0, 5, 6, 7) else ALIGN_LEFT)
            if j == 5:
                cell.fill = FILL_PENDING

    wb.save(OUTPUT_DIR / "sample_test_spec.xlsx")


# ---------------------------------------------------------------------------
# 8. 予算管理表 (Budget management)
# ---------------------------------------------------------------------------

def generate_budget():
    wb = Workbook()
    ws = wb.active
    ws.title = "予算管理表"

    # Start at row 3 (not A1) — realistic: margin rows at top
    RO = 2  # row offset

    # Title
    ws.merge_cells(f"A{1+RO}:N{1+RO}")
    ws.cell(row=1+RO, column=1).value = "2026年度　予算管理表（情報システム部）"
    _apply_style(ws.cell(row=1+RO, column=1), font=FONT_TITLE, alignment=ALIGN_CENTER)
    ws.row_dimensions[1+RO].height = 32

    ws.merge_cells(f"L{2+RO}:N{2+RO}")
    ws.cell(row=2+RO, column=12).value = "単位：千円"
    _apply_style(ws.cell(row=2+RO, column=12), font=FONT_SMALL, alignment=ALIGN_RIGHT)

    # Column structure:
    # A=大分類, B=中分類, C=科目
    # D=年間予算
    # E,F,G=Q1(4-6月) 予算/実績/差異
    # H,I,J=Q2(7-9月) 予算/実績/差異
    # K,L,M=Q3(10-12月) 予算/実績/差異
    # N=通期実績累計

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 11
    for col_letter in "EFGHIJKLMN":
        ws.column_dimensions[col_letter].width = 10

    # Category header row (row 3)
    cat_headers = [
        (1, 3, "勘定科目"),
        (4, 4, "年間"),
        (5, 7, "第1四半期（4-6月）"),
        (8, 10, "第2四半期（7-9月）"),
        (11, 13, "第3四半期（10-12月）"),
        (14, 14, "累計"),
    ]
    for cs, ce, label in cat_headers:
        ws.merge_cells(start_row=3+RO, start_column=cs, end_row=3+RO, end_column=ce)
        c = ws.cell(row=3+RO, column=cs)
        c.value = label
        _apply_style(c, font=FONT_WHITE_BOLD, fill=FILL_DARK_BLUE, border=BORDER_ALL_THIN, alignment=ALIGN_CENTER)

    # Sub-header row (row 4+RO)
    sub_h = ["大分類", "中分類", "科目", "予算", "予算", "実績", "差異", "予算", "実績", "差異", "予算", "実績", "差異", "実績累計"]
    for i, h in enumerate(sub_h):
        c = ws.cell(row=4+RO, column=i + 1)
        c.value = h
        _apply_style(c, font=FONT_BOLD, fill=FILL_LIGHT_BLUE, border=BORDER_ALL_THIN, alignment=ALIGN_CENTER)

    # Budget data (cat1, cat2, item, annual, q1_b, q1_a, q2_b, q2_a, q3_b, q3_a)
    budget_data = [
        ("人件費", "正社員", "基本給", 180000, 45000, 45000, 45000, 45000, 45000, None),
        ("人件費", "正社員", "残業手当", 12000, 3000, 3500, 3000, 2800, 3000, None),
        ("人件費", "正社員", "賞与", 30000, 15000, 15000, 0, 0, 15000, None),
        ("人件費", "派遣・委託", "派遣費用", 48000, 12000, 12000, 12000, 11500, 12000, None),
        ("人件費", "派遣・委託", "業務委託費", 36000, 9000, 9200, 9000, 8800, 9000, None),
        ("物件費", "機器", "サーバー購入", 8000, 4000, 3800, 2000, 2000, 2000, None),
        ("物件費", "機器", "PC・周辺機器", 3000, 1000, 900, 1000, 1100, 1000, None),
        ("物件費", "ソフトウェア", "ライセンス費", 6000, 2000, 2000, 2000, 2000, 2000, None),
        ("物件費", "ソフトウェア", "クラウドサービス", 9600, 2400, 2500, 2400, 2600, 2400, None),
        ("物件費", "通信", "回線費用", 2400, 600, 600, 600, 600, 600, None),
        ("経費", "教育", "研修費", 2000, 500, 400, 500, 600, 500, None),
        ("経費", "教育", "資格取得支援", 800, 200, 150, 200, 250, 200, None),
        ("経費", "旅費", "出張旅費", 1200, 300, 280, 300, 320, 300, None),
        ("経費", "その他", "消耗品費", 600, 150, 140, 150, 160, 150, None),
        ("経費", "その他", "雑費", 400, 100, 90, 100, 110, 100, None),
    ]

    # Track merge ranges for cat1 and cat2
    r = 5 + RO
    prev_cat1 = prev_cat2 = None
    cat1_start = cat2_start = r
    merge_info_cat1 = []
    merge_info_cat2 = []

    for i, (cat1, cat2, item, annual, q1b, q1a, q2b, q2a, q3b, q3a) in enumerate(budget_data):
        row = r + i
        if cat1 != prev_cat1:
            if prev_cat1 is not None:
                merge_info_cat1.append((cat1_start, row - 1, prev_cat1))
            cat1_start = row
            prev_cat1 = cat1
        if cat2 != prev_cat2:
            if prev_cat2 is not None:
                merge_info_cat2.append((cat2_start, row - 1, prev_cat2))
            cat2_start = row
            prev_cat2 = cat2

        ws.cell(row=row, column=1).value = cat1
        ws.cell(row=row, column=2).value = cat2
        ws.cell(row=row, column=3).value = item
        ws.cell(row=row, column=4).value = annual

        ws.cell(row=row, column=5).value = q1b
        ws.cell(row=row, column=6).value = q1a
        ws.cell(row=row, column=7).value = (q1a - q1b) if q1a is not None else None

        ws.cell(row=row, column=8).value = q2b
        ws.cell(row=row, column=9).value = q2a
        ws.cell(row=row, column=10).value = (q2a - q2b) if q2a is not None else None

        ws.cell(row=row, column=11).value = q3b
        ws.cell(row=row, column=12).value = None  # Q3 not yet
        ws.cell(row=row, column=13).value = None

        cumulative = (q1a or 0) + (q2a or 0)
        ws.cell(row=row, column=14).value = cumulative

        # Style all cells in this row
        for c in range(1, 15):
            cell = ws.cell(row=row, column=c)
            _apply_style(cell, font=FONT_BODY, border=BORDER_ALL_THIN,
                         alignment=ALIGN_CENTER if c <= 3 else ALIGN_RIGHT)
            if c >= 4 and cell.value is not None:
                cell.number_format = "#,##0"
            # Color negative differences red
            if c in (7, 10, 13) and cell.value is not None and cell.value > 0:
                cell.font = Font(name="Yu Gothic", size=10, color="CC0000")

    # Finalize merges
    last_data_row = r + len(budget_data) - 1
    merge_info_cat1.append((cat1_start, last_data_row, prev_cat1))
    merge_info_cat2.append((cat2_start, last_data_row, prev_cat2))

    for start_r, end_r, _ in merge_info_cat1:
        if end_r > start_r:
            ws.merge_cells(start_row=start_r, start_column=1, end_row=end_r, end_column=1)
    for start_r, end_r, _ in merge_info_cat2:
        if end_r > start_r:
            ws.merge_cells(start_row=start_r, start_column=2, end_row=end_r, end_column=2)

    # --- Subtotal rows ---
    subtotals = [
        ("人件費　小計", [306000, 84000, 84700, 69000, 68100, 84000, None, None]),
        ("物件費　小計", [29000, 10000, 9800, 8000, 8300, 8000, None, None]),
        ("経費　小計", [5000, 1250, 1060, 1250, 1440, 1250, None, None]),
    ]
    subtotal_row = last_data_row + 1
    for label, vals in subtotals:
        ws.merge_cells(start_row=subtotal_row, start_column=1, end_row=subtotal_row, end_column=3)
        ws.cell(row=subtotal_row, column=1).value = label
        _apply_style(ws.cell(row=subtotal_row, column=1), font=FONT_BOLD, fill=FILL_LIGHT_YELLOW,
                     border=BORDER_ALL_MEDIUM, alignment=ALIGN_CENTER)
        col_idx = 4
        # annual, q1b, q1a, q2b, q2a, q3b, q3a_placeholder, cumulative_placeholder
        values_to_write = [
            vals[0],  # annual
            vals[1], vals[2], vals[2] - vals[1],  # Q1
            vals[3], vals[4], vals[4] - vals[3],  # Q2
            vals[5], vals[6], None,  # Q3
            (vals[2] or 0) + (vals[4] or 0),  # cumulative
        ]
        for v in values_to_write:
            cell = ws.cell(row=subtotal_row, column=col_idx)
            cell.value = v
            _apply_style(cell, font=FONT_BOLD, fill=FILL_LIGHT_YELLOW,
                         border=BORDER_ALL_MEDIUM, alignment=ALIGN_RIGHT)
            if v is not None:
                cell.number_format = "#,##0"
            col_idx += 1
        subtotal_row += 1

    # Grand total
    ws.merge_cells(start_row=subtotal_row, start_column=1, end_row=subtotal_row, end_column=3)
    ws.cell(row=subtotal_row, column=1).value = "合　　計"
    _apply_style(ws.cell(row=subtotal_row, column=1), font=Font(name="Yu Gothic", size=11, bold=True),
                 fill=FILL_DARK_BLUE, border=BORDER_ALL_MEDIUM, alignment=ALIGN_CENTER)
    ws.cell(row=subtotal_row, column=1).font = FONT_WHITE_BOLD

    grand_vals = [340000, 95250, 95560, 78250, 77840, 93250, None, None]
    gv_write = [
        grand_vals[0],
        grand_vals[1], grand_vals[2], grand_vals[2] - grand_vals[1],
        grand_vals[3], grand_vals[4], grand_vals[4] - grand_vals[3],
        grand_vals[5], grand_vals[6], None,
        (grand_vals[2] or 0) + (grand_vals[4] or 0),
    ]
    col_idx = 4
    for v in gv_write:
        cell = ws.cell(row=subtotal_row, column=col_idx)
        cell.value = v
        _apply_style(cell, font=FONT_WHITE_BOLD, fill=FILL_DARK_BLUE,
                     border=BORDER_ALL_MEDIUM, alignment=ALIGN_RIGHT)
        if v is not None:
            cell.number_format = "#,##0"
        col_idx += 1

    # Execution rate row
    rate_row = subtotal_row + 1
    ws.merge_cells(start_row=rate_row, start_column=1, end_row=rate_row, end_column=4)
    ws.cell(row=rate_row, column=1).value = "予算執行率"
    _apply_style(ws.cell(row=rate_row, column=1), font=FONT_BOLD, fill=FILL_LIGHT_GRAY,
                 border=BORDER_ALL_THIN, alignment=ALIGN_CENTER)
    # Q1 rate
    ws.merge_cells(start_row=rate_row, start_column=5, end_row=rate_row, end_column=7)
    ws.cell(row=rate_row, column=5).value = 1.003
    _apply_style(ws.cell(row=rate_row, column=5), font=FONT_BOLD, border=BORDER_ALL_THIN, alignment=ALIGN_CENTER)
    ws.cell(row=rate_row, column=5).number_format = "0.0%"
    # Q2 rate
    ws.merge_cells(start_row=rate_row, start_column=8, end_row=rate_row, end_column=10)
    ws.cell(row=rate_row, column=8).value = 0.995
    _apply_style(ws.cell(row=rate_row, column=8), font=FONT_BOLD, border=BORDER_ALL_THIN, alignment=ALIGN_CENTER)
    ws.cell(row=rate_row, column=8).number_format = "0.0%"

    wb.save(OUTPUT_DIR / "sample_budget.xlsx")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


# ---------------------------------------------------------------------------
# 9. ただ文章を書いただけ (Free text document)
# ---------------------------------------------------------------------------

def generate_freetext():
    wb = Workbook()
    ws = wb.active
    ws.title = "社内通知"

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 4
    ws.column_dimensions["C"].width = 4
    ws.column_dimensions["D"].width = 80
    ws.column_dimensions["E"].width = 10

    # Start at row 3 (not A1) — realistic: letterhead space at top
    row = 3

    # Title in column A - large bold
    ws.cell(row=row, column=1).value = "社内通知"
    _apply_style(ws.cell(row=row, column=1), font=Font(name="Yu Gothic", size=18, bold=True))
    ws.row_dimensions[row].height = 30
    row += 1

    # Empty separator
    row += 1

    # Date and document number - right side feel, but just in column A
    ws.cell(row=row, column=1).value = "2026年3月10日"
    _apply_style(ws.cell(row=row, column=1), font=FONT_BODY)
    row += 1
    ws.cell(row=row, column=1).value = "通知番号：NT-2026-0087"
    _apply_style(ws.cell(row=row, column=1), font=FONT_BODY)
    row += 1

    row += 1  # empty separator

    # Addressee
    ws.cell(row=row, column=1).value = "全社員各位"
    _apply_style(ws.cell(row=row, column=1), font=Font(name="Yu Gothic", size=12, bold=True))
    row += 1

    row += 1  # empty separator

    # Sender - indented to column B
    ws.cell(row=row, column=2).value = "総務部長　山本　健一"
    _apply_style(ws.cell(row=row, column=2), font=FONT_BODY)
    row += 1

    row += 1  # empty separator

    # Subject heading
    ws.cell(row=row, column=1).value = "件名：オフィス移転に伴う業務対応について"
    _apply_style(ws.cell(row=row, column=1), font=Font(name="Yu Gothic", size=13, bold=True))
    ws.row_dimensions[row].height = 24
    row += 1

    row += 1  # empty separator

    # Body paragraphs - line by line in column A
    body_lines = [
        "平素より業務にご尽力いただき、誠にありがとうございます。",
        "さて、かねてよりご案内しておりましたオフィス移転について、下記のとおり",
        "詳細が決定いたしましたのでお知らせいたします。",
    ]
    for line in body_lines:
        ws.cell(row=row, column=1).value = line
        _apply_style(ws.cell(row=row, column=1), font=FONT_BODY)
        row += 1

    row += 1  # empty separator

    # Section heading - bold
    ws.cell(row=row, column=1).value = "1. 移転先情報"
    _apply_style(ws.cell(row=row, column=1), font=Font(name="Yu Gothic", size=11, bold=True))
    row += 1

    # Indented key-value pairs using column B
    kv_items = [
        ("新住所：", "東京都千代田区丸の内1-2-3　丸の内ビルディング15階"),
        ("最寄駅：", "東京メトロ丸ノ内線　東京駅　徒歩3分"),
        ("電話番号：", "03-XXXX-XXXX（変更なし）"),
        ("FAX番号：", "03-XXXX-XXXY（変更あり）"),
    ]
    for key, val in kv_items:
        ws.cell(row=row, column=2).value = key + val
        _apply_style(ws.cell(row=row, column=2), font=FONT_BODY)
        row += 1

    row += 1  # empty separator

    # Another section heading
    ws.cell(row=row, column=1).value = "2. 移転スケジュール"
    _apply_style(ws.cell(row=row, column=1), font=Font(name="Yu Gothic", size=11, bold=True))
    row += 1

    schedule_lines = [
        "3月20日（金）  最終営業日（現オフィス）",
        "3月21日（土）〜22日（日）  引越作業（立入禁止）",
        "3月23日（月）  新オフィスでの業務開始",
    ]
    for line in schedule_lines:
        ws.cell(row=row, column=2).value = line
        _apply_style(ws.cell(row=row, column=2), font=FONT_BODY)
        row += 1

    row += 1

    # Another section
    ws.cell(row=row, column=1).value = "3. 社員の皆様へのお願い"
    _apply_style(ws.cell(row=row, column=1), font=Font(name="Yu Gothic", size=11, bold=True))
    row += 1

    requests_lines = [
        "・私物の整理は3月18日（水）までにお願いいたします。",
        "・段ボール箱は総務部にて配布いたします（1人3箱まで）。",
        "・PC・モニターは専門業者が運搬しますので、そのままにしてください。",
        "・新オフィスの座席表は別途メールにてご案内いたします。",
    ]
    for line in requests_lines:
        # Deeper indent - column C
        ws.cell(row=row, column=3).value = line
        _apply_style(ws.cell(row=row, column=3), font=FONT_BODY)
        row += 1

    row += 1

    # Closing
    ws.cell(row=row, column=1).value = "ご不明な点がございましたら、総務部（内線：1234）までお問い合わせください。"
    _apply_style(ws.cell(row=row, column=1), font=FONT_BODY)
    row += 1

    row += 1
    ws.cell(row=row, column=1).value = "以上"
    _apply_style(ws.cell(row=row, column=1), font=FONT_BODY)

    wb.save(OUTPUT_DIR / "sample_freetext.xlsx")


# ---------------------------------------------------------------------------
# 10. お絵描き帳スタイルの報告書 (Freeform report / sketchbook style)
# ---------------------------------------------------------------------------

def generate_freeform_report():
    from openpyxl.chart import BarChart, Reference

    wb = Workbook()
    ws = wb.active
    ws.title = "月次業績報告"

    # Wide columns for freeform layout
    for col in range(1, 12):
        ws.column_dimensions[get_column_letter(col)].width = 12

    # --- Large title at top ---
    ws.merge_cells("A1:H1")
    ws["A1"].value = "月次業績報告書"
    _apply_style(ws["A1"], font=Font(name="Yu Gothic", size=20, bold=True))
    ws.row_dimensions[1].height = 36

    # --- Subtitle indented at column B ---
    ws.merge_cells("B3:G3")
    ws["B3"].value = "2026年2月度　営業部実績まとめ"
    _apply_style(ws["B3"], font=Font(name="Yu Gothic", size=14, bold=True))
    ws.row_dimensions[3].height = 28

    # --- Key-value pairs at arbitrary positions ---
    ws.cell(row=5, column=7).value = "担当者："
    _apply_style(ws.cell(row=5, column=7), font=FONT_BOLD)
    ws.cell(row=5, column=8).value = "田中太郎"
    _apply_style(ws.cell(row=5, column=8), font=FONT_BODY)

    ws.cell(row=6, column=7).value = "日付："
    _apply_style(ws.cell(row=6, column=7), font=FONT_BOLD)
    ws.cell(row=6, column=8).value = "2026/03/15"
    _apply_style(ws.cell(row=6, column=8), font=FONT_BODY)

    ws.cell(row=7, column=7).value = "部署："
    _apply_style(ws.cell(row=7, column=7), font=FONT_BOLD)
    ws.cell(row=7, column=8).value = "第一営業部"
    _apply_style(ws.cell(row=7, column=8), font=FONT_BODY)

    # --- Free text paragraphs ---
    ws.cell(row=5, column=1).value = "【概要】"
    _apply_style(ws.cell(row=5, column=1), font=FONT_HEADING)

    text_lines = [
        "2月度の営業実績について報告いたします。全体として目標達成率は92%と",
        "なり、前月比で5ポイント改善しました。特に新規顧客開拓において顕著な",
        "成果が見られましたが、既存顧客のリピート率に課題が残ります。",
    ]
    for i, line in enumerate(text_lines):
        ws.cell(row=6 + i, column=1).value = line
        _apply_style(ws.cell(row=6 + i, column=1), font=FONT_BODY)

    # --- Heading for first table ---
    ws.cell(row=10, column=1).value = "【地域別売上実績】"
    _apply_style(ws.cell(row=10, column=1), font=FONT_HEADING)

    # --- Small table at B12-D17 (3 columns × 6 rows including header) ---
    table1_headers = ["地域", "売上（万円）", "達成率"]
    table1_data = [
        ("東京", 4500, "95%"),
        ("大阪", 2800, "88%"),
        ("名古屋", 1500, "102%"),
        ("福岡", 980, "85%"),
        ("合計", 9780, "92%"),
    ]

    for j, h in enumerate(table1_headers):
        cell = ws.cell(row=12, column=2 + j)
        cell.value = h
        _apply_style(cell, font=FONT_WHITE_BOLD, fill=FILL_DARK_BLUE,
                     border=BORDER_ALL_THIN, alignment=ALIGN_CENTER)

    for i, (region, sales, rate) in enumerate(table1_data):
        r = 13 + i
        is_total = (i == len(table1_data) - 1)
        ws.cell(row=r, column=2).value = region
        _apply_style(ws.cell(row=r, column=2), font=FONT_BOLD if is_total else FONT_BODY,
                     fill=FILL_LIGHT_YELLOW if is_total else None,
                     border=BORDER_ALL_THIN, alignment=ALIGN_CENTER)
        ws.cell(row=r, column=3).value = sales
        _apply_style(ws.cell(row=r, column=3), font=FONT_BOLD if is_total else FONT_BODY,
                     fill=FILL_LIGHT_YELLOW if is_total else None,
                     border=BORDER_ALL_THIN, alignment=ALIGN_RIGHT)
        ws.cell(row=r, column=4).value = rate
        _apply_style(ws.cell(row=r, column=4), font=FONT_BOLD if is_total else FONT_BODY,
                     fill=FILL_LIGHT_YELLOW if is_total else None,
                     border=BORDER_ALL_THIN, alignment=ALIGN_CENTER)

    # --- Text below first table ---
    ws.cell(row=19, column=1).value = "名古屋地域が唯一目標を上回りました。新規大口案件（中部電機様）の"
    _apply_style(ws.cell(row=19, column=1), font=FONT_BODY)
    ws.cell(row=20, column=1).value = "受注が大きく貢献しています。福岡は担当者交代の影響が出ています。"
    _apply_style(ws.cell(row=20, column=1), font=FONT_BODY)

    # --- Second table at a different position (F12-H16) ---
    ws.cell(row=10, column=6).value = "【商品カテゴリ別】"
    _apply_style(ws.cell(row=10, column=6), font=FONT_HEADING)

    table2_headers = ["カテゴリ", "件数", "金額（万円）"]
    table2_data = [
        ("ソフトウェア", 45, 5200),
        ("ハードウェア", 23, 2800),
        ("サービス", 67, 1780),
    ]
    for j, h in enumerate(table2_headers):
        cell = ws.cell(row=12, column=6 + j)
        cell.value = h
        _apply_style(cell, font=FONT_WHITE_BOLD, fill=FILL_DARK_GREEN,
                     border=BORDER_ALL_THIN, alignment=ALIGN_CENTER)

    for i, (cat, cnt, amt) in enumerate(table2_data):
        r = 13 + i
        ws.cell(row=r, column=6).value = cat
        _apply_style(ws.cell(row=r, column=6), font=FONT_BODY, border=BORDER_ALL_THIN, alignment=ALIGN_CENTER)
        ws.cell(row=r, column=7).value = cnt
        _apply_style(ws.cell(row=r, column=7), font=FONT_BODY, border=BORDER_ALL_THIN, alignment=ALIGN_RIGHT)
        ws.cell(row=r, column=8).value = amt
        _apply_style(ws.cell(row=r, column=8), font=FONT_BODY, border=BORDER_ALL_THIN, alignment=ALIGN_RIGHT)

    # --- Bar chart from first table data ---
    chart = BarChart()
    chart.type = "col"
    chart.title = "地域別売上（万円）"
    chart.y_axis.title = "万円"
    chart.x_axis.title = "地域"
    chart.style = 10
    chart.width = 15
    chart.height = 10

    data_ref = Reference(ws, min_col=3, min_row=12, max_row=16, max_col=3)
    cats_ref = Reference(ws, min_col=2, min_row=13, max_row=16)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws.add_chart(chart, "B22")

    # --- More text after chart ---
    ws.cell(row=38, column=1).value = "【今後のアクション】"
    _apply_style(ws.cell(row=38, column=1), font=FONT_HEADING)

    action_lines = [
        "1. 福岡地域の新担当者への引き継ぎ強化（3月末まで）",
        "2. 東京・大阪の既存顧客フォロー訪問の実施（毎週）",
        "3. 名古屋の成功事例を他地域に横展開（4月度営業会議にて共有）",
        "4. 新規顧客向けキャンペーン企画の立案（3月20日締切）",
    ]
    for i, line in enumerate(action_lines):
        ws.cell(row=39 + i, column=2).value = line
        _apply_style(ws.cell(row=39 + i, column=2), font=FONT_BODY)

    wb.save(OUTPUT_DIR / "sample_freeform_report.xlsx")


# ---------------------------------------------------------------------------
# 11. 完全自由配置 (Sketchpad / free-layout)
# ---------------------------------------------------------------------------

def generate_sketchpad():
    wb = Workbook()
    ws = wb.active
    ws.title = "オリエンテーション"

    for col in range(1, 14):
        ws.column_dimensions[get_column_letter(col)].width = 12

    FILL_ISLAND1 = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    FILL_ISLAND2 = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
    FILL_ISLAND3 = PatternFill(start_color="F5EEF8", end_color="F5EEF8", fill_type="solid")
    FILL_ISLAND5 = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")

    # --- Island 1 (A1-D5): タイトルブロック ---
    for r in range(1, 6):
        for c in range(1, 5):
            ws.cell(row=r, column=c).fill = FILL_ISLAND1

    ws.merge_cells("A2:D2")
    ws["A2"].value = "新人向け"
    _apply_style(ws["A2"], font=Font(name="Yu Gothic", size=18, bold=True, color="FFFFFF"),
                 fill=FILL_ISLAND1, alignment=ALIGN_CENTER)

    ws.merge_cells("A3:D3")
    ws["A3"].value = "オリエンテーション資料"
    _apply_style(ws["A3"], font=Font(name="Yu Gothic", size=16, bold=True, color="FFFFFF"),
                 fill=FILL_ISLAND1, alignment=ALIGN_CENTER)

    ws.merge_cells("A5:D5")
    ws["A5"].value = "2026年度　入社時研修"
    _apply_style(ws["A5"], font=Font(name="Yu Gothic", size=10, color="AAAAAA"),
                 fill=FILL_ISLAND1, alignment=ALIGN_CENTER)

    # --- Island 2 (F2-I8): 連絡先情報 ---
    ws.cell(row=2, column=6).value = "【連絡先情報】"
    _apply_style(ws.cell(row=2, column=6), font=Font(name="Yu Gothic", size=12, bold=True, color="1A5276"))

    contact_kv = [
        ("人事部", "内線 2001"),
        ("総務部", "内線 2002"),
        ("IT部門", "内線 2003"),
        ("研修担当", "佐藤花子"),
        ("メール", "kenshu@example.co.jp"),
    ]
    for i, (key, val) in enumerate(contact_kv):
        r = 3 + i
        ws.cell(row=r, column=6).value = key
        _apply_style(ws.cell(row=r, column=6), font=FONT_BOLD, fill=FILL_ISLAND2,
                     border=BORDER_ALL_THIN, alignment=ALIGN_LEFT)
        ws.merge_cells(start_row=r, start_column=7, end_row=r, end_column=9)
        ws.cell(row=r, column=7).value = val
        _apply_style(ws.cell(row=r, column=7), font=FONT_BODY, fill=FILL_ISLAND2,
                     border=BORDER_ALL_THIN, alignment=ALIGN_LEFT)

    # --- Island 3 (B10-E18): 作業手順 numbered list ---
    ws.cell(row=10, column=2).value = "【初日の流れ】"
    _apply_style(ws.cell(row=10, column=2), font=Font(name="Yu Gothic", size=13, bold=True, color="6C3483"))

    steps = [
        "1. 受付にて入館証を受け取る（9:00）",
        "2. 5階大会議室に集合（9:15）",
        "3. オリエンテーション開始（9:30〜12:00）",
        "4. 昼食休憩　※社員食堂3階（12:00〜13:00）",
        "5. 部門別研修（13:00〜15:00）",
        "6. IT機器セットアップ（15:00〜16:30）",
        "7. 質疑応答・初日終了（16:30〜17:00）",
    ]
    for i, step in enumerate(steps):
        r = 11 + i
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        ws.cell(row=r, column=2).value = step
        _apply_style(ws.cell(row=r, column=2), font=FONT_BODY, fill=FILL_ISLAND3, alignment=ALIGN_LEFT)

    # --- Island 4 (G12-K16): Small data table with totals ---
    ws.cell(row=11, column=7).value = "【研修スケジュール概要】"
    _apply_style(ws.cell(row=11, column=7), font=Font(name="Yu Gothic", size=11, bold=True))

    tbl_headers = ["研修項目", "日数", "場所"]
    for j, h in enumerate(tbl_headers):
        cell = ws.cell(row=12, column=7 + j)
        cell.value = h
        _apply_style(cell, font=FONT_WHITE_BOLD, fill=FILL_DARK_BLUE,
                     border=BORDER_ALL_THIN, alignment=ALIGN_CENTER)

    tbl_data = [
        ("全体研修", 3, "大会議室"),
        ("部門別研修", 5, "各部門"),
        ("OJT", 10, "配属先"),
        ("合計", 18, "―"),
    ]
    for i, (item, days, place) in enumerate(tbl_data):
        r = 13 + i
        is_total = (i == len(tbl_data) - 1)
        ws.cell(row=r, column=7).value = item
        _apply_style(ws.cell(row=r, column=7), font=FONT_BOLD if is_total else FONT_BODY,
                     fill=FILL_LIGHT_YELLOW if is_total else None,
                     border=BORDER_ALL_THIN, alignment=ALIGN_LEFT)
        ws.cell(row=r, column=8).value = days
        _apply_style(ws.cell(row=r, column=8), font=FONT_BOLD if is_total else FONT_BODY,
                     fill=FILL_LIGHT_YELLOW if is_total else None,
                     border=BORDER_ALL_THIN, alignment=ALIGN_CENTER)
        ws.cell(row=r, column=9).value = place
        _apply_style(ws.cell(row=r, column=9), font=FONT_BOLD if is_total else FONT_BODY,
                     fill=FILL_LIGHT_YELLOW if is_total else None,
                     border=BORDER_ALL_THIN, alignment=ALIGN_CENTER)

    # --- Island 5 (A22-H28): 注意事項 in bordered box with yellow background ---
    ws.cell(row=21, column=1).value = "【注意事項】"
    _apply_style(ws.cell(row=21, column=1), font=Font(name="Yu Gothic", size=13, bold=True, color="B7950B"))

    cautions = [
        "● 初日は必ずスーツ着用でお越しください。",
        "● 入館証は常時携帯してください。紛失時は総務部へ連絡。",
        "● 社員食堂は社員証（仮）で利用可能です。",
        "● 研修中の携帯電話はマナーモードに設定してください。",
        "● 体調不良の場合は研修担当（佐藤）まで連絡をお願いします。",
        "● 駐車場の利用は事前申請が必要です。",
    ]
    for i, caution in enumerate(cautions):
        r = 22 + i
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        cell = ws.cell(row=r, column=1)
        cell.value = caution
        _apply_style(cell, font=FONT_BODY, fill=FILL_ISLAND5,
                     border=BORDER_ALL_THIN, alignment=ALIGN_LEFT)
        # Apply border to all cells in the merged range
        for c in range(1, 9):
            ws.cell(row=r, column=c).border = BORDER_ALL_THIN

    wb.save(OUTPUT_DIR / "sample_sketchpad.xlsx")


# ---------------------------------------------------------------------------
# 12. 画像入りドキュメント (Document with images)
# ---------------------------------------------------------------------------

def generate_with_images():
    from openpyxl.drawing.image import Image as XlImage

    wb = Workbook()
    ws = wb.active
    ws.title = "作業手順書"

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20

    # Title
    ws.merge_cells("A1:E1")
    ws["A1"].value = "作業手順書：サーバー再起動手順"
    _apply_style(ws["A1"], font=Font(name="Yu Gothic", size=16, bold=True))
    ws.row_dimensions[1].height = 30

    # Document info
    doc_info = [
        ("文書番号", "OP-SRV-003"),
        ("作成日", "2026年3月1日"),
        ("作成者", "インフラ部　鈴木一郎"),
        ("対象", "本番Webサーバー（srv-web-01〜03）"),
    ]
    for i, (k, v) in enumerate(doc_info):
        r = 3 + i
        ws.cell(row=r, column=1).value = k
        _apply_style(ws.cell(row=r, column=1), font=FONT_BOLD, fill=FILL_LIGHT_GRAY,
                     border=BORDER_ALL_THIN, alignment=ALIGN_CENTER)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        ws.cell(row=r, column=2).value = v
        _apply_style(ws.cell(row=r, column=2), font=FONT_BODY, border=BORDER_ALL_THIN, alignment=ALIGN_LEFT)

    # Step 1
    ws.cell(row=8, column=1).value = "手順1"
    _apply_style(ws.cell(row=8, column=1), font=Font(name="Yu Gothic", size=12, bold=True))
    ws.cell(row=9, column=1).value = "監視ツールにログインし、対象サーバーのステータスを確認する。"
    _apply_style(ws.cell(row=9, column=1), font=FONT_BODY)
    ws.cell(row=10, column=1).value = "URL: https://monitor.example.co.jp/dashboard"
    _apply_style(ws.cell(row=10, column=1), font=FONT_BODY)

    # Try to generate placeholder images with Pillow
    images_created = False
    try:
        from PIL import Image as PILImage, ImageDraw, ImageFont
        import io

        def _create_placeholder(width, height, bg_color, text, text_color="white"):
            img = PILImage.new("RGB", (width, height), bg_color)
            draw = ImageDraw.Draw(img)
            # Use default font
            try:
                font = ImageFont.truetype("/System/Library/Fonts/Helvetica.ttc", 14)
            except (OSError, IOError):
                font = ImageFont.load_default()
            bbox = draw.textbbox((0, 0), text, font=font)
            tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]
            x = (width - tw) // 2
            y = (height - th) // 2
            draw.text((x, y), text, fill=text_color, font=font)
            # Draw border
            draw.rectangle([0, 0, width - 1, height - 1], outline="gray")
            buf = io.BytesIO()
            img.save(buf, format="PNG")
            buf.seek(0)
            return buf

        img1_buf = _create_placeholder(300, 150, "#3498DB", "Screenshot 1:\nDashboard")
        img2_buf = _create_placeholder(300, 150, "#2ECC71", "Screenshot 2:\nServer Status")
        img3_buf = _create_placeholder(300, 150, "#E74C3C", "Screenshot 3:\nRestart Button")

        ws.cell(row=11, column=1).value = "【画面キャプチャ：ダッシュボード画面】"
        _apply_style(ws.cell(row=11, column=1), font=FONT_SMALL)
        xl_img1 = XlImage(img1_buf)
        ws.add_image(xl_img1, "B12")

        # Step 2
        ws.cell(row=22, column=1).value = "手順2"
        _apply_style(ws.cell(row=22, column=1), font=Font(name="Yu Gothic", size=12, bold=True))
        ws.cell(row=23, column=1).value = "対象サーバーを選択し、「再起動」ボタンをクリックする。"
        _apply_style(ws.cell(row=23, column=1), font=FONT_BODY)
        ws.cell(row=24, column=1).value = "【画面キャプチャ：サーバー選択画面】"
        _apply_style(ws.cell(row=24, column=1), font=FONT_SMALL)
        xl_img2 = XlImage(img2_buf)
        ws.add_image(xl_img2, "B25")

        # Step 3
        ws.cell(row=35, column=1).value = "手順3"
        _apply_style(ws.cell(row=35, column=1), font=Font(name="Yu Gothic", size=12, bold=True))
        ws.cell(row=36, column=1).value = "確認ダイアログで「OK」を押下し、再起動を実行する。"
        _apply_style(ws.cell(row=36, column=1), font=FONT_BODY)
        ws.cell(row=37, column=1).value = "【画面キャプチャ：確認ダイアログ】"
        _apply_style(ws.cell(row=37, column=1), font=FONT_SMALL)
        xl_img3 = XlImage(img3_buf)
        ws.add_image(xl_img3, "B38")

        images_created = True
    except ImportError:
        # Pillow not available - create without images
        ws.cell(row=11, column=1).value = "【画像省略：Pillowライブラリが利用できないため】"
        _apply_style(ws.cell(row=11, column=1), font=Font(name="Yu Gothic", size=9, italic=True, color="999999"))

        ws.cell(row=14, column=1).value = "手順2"
        _apply_style(ws.cell(row=14, column=1), font=Font(name="Yu Gothic", size=12, bold=True))
        ws.cell(row=15, column=1).value = "対象サーバーを選択し、「再起動」ボタンをクリックする。"
        _apply_style(ws.cell(row=15, column=1), font=FONT_BODY)

        ws.cell(row=18, column=1).value = "手順3"
        _apply_style(ws.cell(row=18, column=1), font=Font(name="Yu Gothic", size=12, bold=True))
        ws.cell(row=19, column=1).value = "確認ダイアログで「OK」を押下し、再起動を実行する。"
        _apply_style(ws.cell(row=19, column=1), font=FONT_BODY)

    # Small reference table near the bottom
    tbl_row = 48 if images_created else 22
    ws.cell(row=tbl_row, column=1).value = "【対象サーバー一覧】"
    _apply_style(ws.cell(row=tbl_row, column=1), font=FONT_HEADING)

    srv_headers = ["サーバー名", "IPアドレス", "用途"]
    for j, h in enumerate(srv_headers):
        cell = ws.cell(row=tbl_row + 1, column=2 + j)
        cell.value = h
        _apply_style(cell, font=FONT_WHITE_BOLD, fill=FILL_DARK_BLUE,
                     border=BORDER_ALL_THIN, alignment=ALIGN_CENTER)

    servers = [
        ("srv-web-01", "192.168.1.101", "Webサーバー（主系）"),
        ("srv-web-02", "192.168.1.102", "Webサーバー（副系）"),
        ("srv-web-03", "192.168.1.103", "Webサーバー（待機系）"),
    ]
    for i, (name, ip, role) in enumerate(servers):
        r = tbl_row + 2 + i
        ws.cell(row=r, column=2).value = name
        _apply_style(ws.cell(row=r, column=2), font=FONT_BODY, border=BORDER_ALL_THIN, alignment=ALIGN_CENTER)
        ws.cell(row=r, column=3).value = ip
        _apply_style(ws.cell(row=r, column=3), font=FONT_BODY, border=BORDER_ALL_THIN, alignment=ALIGN_CENTER)
        ws.cell(row=r, column=4).value = role
        _apply_style(ws.cell(row=r, column=4), font=FONT_BODY, border=BORDER_ALL_THIN, alignment=ALIGN_LEFT)

    if not images_created:
        print("(Pillow not available - images skipped)", end=" ")

    wb.save(OUTPUT_DIR / "sample_with_images.xlsx")


# ---------------------------------------------------------------------------
# 13. 文章の合間にテーブルとグラフ (Mixed document with tables and charts)
# ---------------------------------------------------------------------------

def generate_mixed_document():
    from openpyxl.chart import BarChart, LineChart, Reference

    wb = Workbook()
    ws = wb.active
    ws.title = "営業実績分析"

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16

    row = 1

    # --- Title ---
    ws.merge_cells("A1:F1")
    ws["A1"].value = "営業実績分析レポート"
    _apply_style(ws["A1"], font=Font(name="Yu Gothic", size=18, bold=True))
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:F2")
    ws["A2"].value = "2025年度　通期実績　営業本部"
    _apply_style(ws["A2"], font=Font(name="Yu Gothic", size=12, bold=True, color="555555"))
    row = 4

    # --- Introduction paragraphs ---
    intro_lines = [
        "本レポートは、2025年度の営業実績を地域別・四半期別に分析し、",
        "次年度の戦略立案に資するデータを提供することを目的とする。",
        "",
        "2025年度の売上高合計は38,200万円となり、年間目標40,000万円に対して",
        "達成率95.5%であった。前年度比では103%と微増であるが、下期の減速傾向が",
        "見られるため、2026年度は上期の勢いを維持する施策が求められる。",
    ]
    for line in intro_lines:
        if line:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            ws.cell(row=row, column=1).value = line
            _apply_style(ws.cell(row=row, column=1), font=FONT_BODY)
        row += 1

    row += 1

    # --- Section 1: 地域別売上 ---
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(row=row, column=1).value = "1. 地域別売上実績"
    _apply_style(ws.cell(row=row, column=1), font=Font(name="Yu Gothic", size=13, bold=True),
                 border=BORDER_BOTTOM_MEDIUM)
    row += 2

    # Table 1: 2 columns x 5 rows (Region, Sales)
    tbl1_start = row
    tbl1_headers = ["地域", "売上高（万円）"]
    for j, h in enumerate(tbl1_headers):
        cell = ws.cell(row=row, column=2 + j)
        cell.value = h
        _apply_style(cell, font=FONT_WHITE_BOLD, fill=FILL_DARK_BLUE,
                     border=BORDER_ALL_THIN, alignment=ALIGN_CENTER)
    row += 1

    region_data = [
        ("東京", 15200),
        ("大阪", 9800),
        ("名古屋", 6500),
        ("福岡", 3800),
        ("その他", 2900),
    ]
    for region, sales in region_data:
        ws.cell(row=row, column=2).value = region
        _apply_style(ws.cell(row=row, column=2), font=FONT_BODY,
                     border=BORDER_ALL_THIN, alignment=ALIGN_CENTER)
        ws.cell(row=row, column=3).value = sales
        _apply_style(ws.cell(row=row, column=3), font=FONT_BODY,
                     border=BORDER_ALL_THIN, alignment=ALIGN_RIGHT)
        row += 1

    tbl1_end = row - 1

    # Bar chart from table 1
    chart1 = BarChart()
    chart1.type = "col"
    chart1.title = "地域別売上高（万円）"
    chart1.y_axis.title = "万円"
    chart1.style = 10
    chart1.width = 14
    chart1.height = 10

    data1 = Reference(ws, min_col=3, min_row=tbl1_start, max_row=tbl1_end)
    cats1 = Reference(ws, min_col=2, min_row=tbl1_start + 1, max_row=tbl1_end)
    chart1.add_data(data1, titles_from_data=True)
    chart1.set_categories(cats1)
    ws.add_chart(chart1, f"B{row + 1}")

    row += 17  # space for chart

    # --- Analysis text ---
    analysis1 = [
        "東京が全体の39.8%を占め、引き続き最大の売上拠点である。",
        "大阪は前年度比110%と大幅な伸びを示しており、新規開拓チームの",
        "活動が成果を上げている。一方、福岡は前年度比92%と減少しており、",
        "主力顧客の予算削減が影響している。",
    ]
    for line in analysis1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1).value = line
        _apply_style(ws.cell(row=row, column=1), font=FONT_BODY)
        row += 1

    row += 2

    # --- Section 2: 四半期別推移 ---
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(row=row, column=1).value = "2. 四半期別売上推移"
    _apply_style(ws.cell(row=row, column=1), font=Font(name="Yu Gothic", size=13, bold=True),
                 border=BORDER_BOTTOM_MEDIUM)
    row += 2

    # Table 2: 3 columns x 4 rows (Quarter, Sales, Target)
    tbl2_start = row
    tbl2_headers = ["四半期", "売上高（万円）", "目標（万円）", "達成率"]
    for j, h in enumerate(tbl2_headers):
        cell = ws.cell(row=row, column=2 + j)
        cell.value = h
        _apply_style(cell, font=FONT_WHITE_BOLD, fill=FILL_DARK_GREEN,
                     border=BORDER_ALL_THIN, alignment=ALIGN_CENTER)
    row += 1

    q_data = [
        ("Q1（4-6月）", 10500, 10000, "105%"),
        ("Q2（7-9月）", 10200, 10000, "102%"),
        ("Q3（10-12月）", 9200, 10000, "92%"),
        ("Q4（1-3月）", 8300, 10000, "83%"),
    ]
    for q, sales, target, rate in q_data:
        ws.cell(row=row, column=2).value = q
        _apply_style(ws.cell(row=row, column=2), font=FONT_BODY,
                     border=BORDER_ALL_THIN, alignment=ALIGN_CENTER)
        ws.cell(row=row, column=3).value = sales
        _apply_style(ws.cell(row=row, column=3), font=FONT_BODY,
                     border=BORDER_ALL_THIN, alignment=ALIGN_RIGHT)
        ws.cell(row=row, column=4).value = target
        _apply_style(ws.cell(row=row, column=4), font=FONT_BODY,
                     border=BORDER_ALL_THIN, alignment=ALIGN_RIGHT)
        ws.cell(row=row, column=5).value = rate
        _apply_style(ws.cell(row=row, column=5), font=FONT_BODY,
                     border=BORDER_ALL_THIN, alignment=ALIGN_CENTER)
        row += 1

    tbl2_end = row - 1

    # Line chart from table 2
    chart2 = LineChart()
    chart2.title = "四半期別売上推移"
    chart2.y_axis.title = "万円"
    chart2.style = 10
    chart2.width = 14
    chart2.height = 10

    data2_sales = Reference(ws, min_col=3, min_row=tbl2_start, max_row=tbl2_end)
    data2_target = Reference(ws, min_col=4, min_row=tbl2_start, max_row=tbl2_end)
    cats2 = Reference(ws, min_col=2, min_row=tbl2_start + 1, max_row=tbl2_end)
    chart2.add_data(data2_sales, titles_from_data=True)
    chart2.add_data(data2_target, titles_from_data=True)
    chart2.set_categories(cats2)
    ws.add_chart(chart2, f"B{row + 1}")

    row += 17  # space for chart

    # --- Conclusion ---
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(row=row, column=1).value = "3. まとめ・次年度への提言"
    _apply_style(ws.cell(row=row, column=1), font=Font(name="Yu Gothic", size=13, bold=True),
                 border=BORDER_BOTTOM_MEDIUM)
    row += 2

    conclusion_lines = [
        "2025年度は上期に堅調な推移を見せたものの、下期に減速が見られた。",
        "主な要因は以下の通りである。",
        "",
        "  ・Q3以降の大口案件の失注（3件、計2,500万円相当）",
        "  ・福岡拠点の主力顧客における予算凍結の影響",
        "  ・競合他社の価格攻勢による既存顧客の一部流出",
        "",
        "2026年度に向けた重点施策として以下を提言する。",
        "",
        "  1. 上期の勢いを持続するためのパイプライン管理の強化",
        "  2. 福岡拠点の立て直し（新規顧客開拓チームの増員）",
        "  3. 既存顧客向けクロスセル・アップセル施策の展開",
        "  4. 競合対策としてのソリューション提案力の強化",
        "",
        "以上の施策を実行することで、2026年度は売上高42,000万円（前年度比110%）を",
        "目標とする。",
    ]
    for line in conclusion_lines:
        if line:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            ws.cell(row=row, column=1).value = line
            _apply_style(ws.cell(row=row, column=1), font=FONT_BODY)
        row += 1

    wb.save(OUTPUT_DIR / "sample_mixed_document.xlsx")


# ---------------------------------------------------------------------------
# 14. ワード風の階層構造テキスト (Word-style hierarchical text)
# ---------------------------------------------------------------------------

def generate_hierarchical_text():
    wb = Workbook()
    ws = wb.active
    ws.title = "設計方針書"

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 6
    ws.column_dimensions["D"].width = 70
    ws.column_dimensions["E"].width = 15

    # Start at row 4 (not A1) — realistic: margin rows at top
    # Title
    ws.merge_cells("A4:E4")
    ws["A4"].value = "システム設計方針書"
    _apply_style(ws["A4"], font=Font(name="Yu Gothic", size=18, bold=True))
    ws.row_dimensions[4].height = 36

    ws.merge_cells("A5:E5")
    ws["A5"].value = "顧客管理システム刷新プロジェクト"
    _apply_style(ws["A5"], font=Font(name="Yu Gothic", size=12, bold=True, color="555555"))

    # Document info
    ws.cell(row=6, column=4).value = "版数：1.0　作成日：2026年2月20日　作成者：設計チーム"
    _apply_style(ws.cell(row=6, column=4), font=FONT_SMALL)

    row = 8

    # Hierarchical content structure
    # Level definitions:
    #   chapter: col A, font 16 bold
    #   section: col B, font 13 bold
    #   subsection: col C, font 11 bold
    #   body: col D, font 10 normal

    def _chapter(ws, row, text):
        ws.cell(row=row, column=1).value = text
        _apply_style(ws.cell(row=row, column=1), font=Font(name="Yu Gothic", size=16, bold=True),
                     border=BORDER_BOTTOM_MEDIUM)
        ws.row_dimensions[row].height = 28
        return row + 1

    def _section(ws, row, text):
        ws.cell(row=row, column=2).value = text
        _apply_style(ws.cell(row=row, column=2), font=Font(name="Yu Gothic", size=13, bold=True),
                     border=BORDER_BOTTOM_THIN)
        ws.row_dimensions[row].height = 24
        return row + 1

    def _subsection(ws, row, text):
        ws.cell(row=row, column=3).value = text
        _apply_style(ws.cell(row=row, column=3), font=Font(name="Yu Gothic", size=11, bold=True))
        return row + 1

    def _body(ws, row, text):
        ws.cell(row=row, column=4).value = text
        _apply_style(ws.cell(row=row, column=4), font=FONT_BODY)
        return row + 1

    def _bullet(ws, row, text):
        ws.cell(row=row, column=4).value = f"・{text}"
        _apply_style(ws.cell(row=row, column=4), font=FONT_BODY)
        return row + 1

    def _numbered(ws, row, num, text):
        ws.cell(row=row, column=4).value = f"{num}. {text}"
        _apply_style(ws.cell(row=row, column=4), font=FONT_BODY)
        return row + 1

    # --- Chapter 1 ---
    row = _chapter(ws, row, "第1章　設計基本方針")
    row += 1

    row = _section(ws, row, "1.1 目的")
    row = _body(ws, row, "本書は、顧客管理システム刷新プロジェクトにおける設計の基本方針を定める。")
    row = _body(ws, row, "開発チーム全員が統一された方針に基づいて設計を行うことで、")
    row = _body(ws, row, "システム全体の一貫性と品質を確保することを目的とする。")
    row += 1

    row = _section(ws, row, "1.2 適用範囲")
    row = _body(ws, row, "本方針書は以下のサブシステムに適用する。")
    row = _bullet(ws, row, "顧客情報管理サブシステム")
    row = _bullet(ws, row, "営業支援サブシステム")
    row = _bullet(ws, row, "帳票出力サブシステム")
    row = _bullet(ws, row, "バッチ処理サブシステム")
    row += 1

    row = _section(ws, row, "1.3 設計原則")
    row = _body(ws, row, "以下の原則に基づいて設計を行う。")
    row = _numbered(ws, row, 1, "疎結合・高凝集の実現")
    row = _numbered(ws, row, 2, "テスタビリティの確保")
    row = _numbered(ws, row, 3, "拡張性を考慮した設計")
    row = _numbered(ws, row, 4, "セキュリティ・バイ・デザイン")
    row += 1

    # --- Chapter 2 ---
    row = _chapter(ws, row, "第2章　アーキテクチャ方針")
    row += 1

    row = _section(ws, row, "2.1 全体構成")
    row = _body(ws, row, "マイクロサービスアーキテクチャを採用し、各サブシステムを独立した")
    row = _body(ws, row, "サービスとして構築する。サービス間通信にはREST APIを基本とし、")
    row = _body(ws, row, "非同期処理が必要な場合はメッセージキューを利用する。")
    row += 1

    row = _subsection(ws, row, "2.1.1 サービス構成")
    row = _body(ws, row, "各サービスの責務と技術スタックを以下に示す。")
    row += 1

    # Inline table for service configuration
    svc_headers = ["サービス名", "責務", "技術スタック"]
    for j, h in enumerate(svc_headers):
        cell = ws.cell(row=row, column=2 + j)
        cell.value = h
        _apply_style(cell, font=FONT_WHITE_BOLD, fill=FILL_DARK_BLUE,
                     border=BORDER_ALL_THIN, alignment=ALIGN_CENTER)
    row += 1

    svc_data = [
        ("customer-api", "顧客情報CRUD", "Python / FastAPI"),
        ("sales-api", "営業データ管理", "Python / FastAPI"),
        ("report-service", "帳票生成", "Java / Spring Boot"),
        ("batch-service", "夜間バッチ", "Python / Celery"),
        ("api-gateway", "認証・ルーティング", "Kong / Nginx"),
    ]
    for svc, role, stack in svc_data:
        ws.cell(row=row, column=2).value = svc
        _apply_style(ws.cell(row=row, column=2), font=FONT_BODY, border=BORDER_ALL_THIN, alignment=ALIGN_LEFT)
        ws.cell(row=row, column=3).value = role
        _apply_style(ws.cell(row=row, column=3), font=FONT_BODY, border=BORDER_ALL_THIN, alignment=ALIGN_LEFT)
        ws.cell(row=row, column=4).value = stack
        _apply_style(ws.cell(row=row, column=4), font=FONT_BODY, border=BORDER_ALL_THIN, alignment=ALIGN_LEFT)
        row += 1

    row += 1

    row = _subsection(ws, row, "2.1.2 データストア")
    row = _body(ws, row, "データストアは用途に応じて以下を使い分ける。")
    row = _bullet(ws, row, "RDB（PostgreSQL）：顧客マスタ、営業データ等の構造化データ")
    row = _bullet(ws, row, "Redis：セッション管理、キャッシュ")
    row = _bullet(ws, row, "S3互換ストレージ：帳票PDF、添付ファイル")
    row += 1

    # --- Chapter 3 ---
    row = _chapter(ws, row, "第3章　セキュリティ方針")
    row += 1

    row = _section(ws, row, "3.1 認証・認可")
    row = _body(ws, row, "OAuth 2.0 + OpenID Connectによる認証基盤を構築する。")
    row = _body(ws, row, "ロールベースアクセス制御（RBAC）により、機能単位の権限管理を実現する。")
    row += 1

    row = _subsection(ws, row, "3.1.1 ロール定義")

    # Inline table for roles
    role_headers = ["ロール", "権限概要"]
    for j, h in enumerate(role_headers):
        cell = ws.cell(row=row, column=2 + j)
        cell.value = h
        _apply_style(cell, font=FONT_WHITE_BOLD, fill=FILL_DARK_GREEN,
                     border=BORDER_ALL_THIN, alignment=ALIGN_CENTER)
    row += 1

    roles_data = [
        ("システム管理者", "全機能の参照・更新・削除、ユーザー管理"),
        ("営業マネージャー", "担当チームの顧客・営業データの参照・更新"),
        ("営業担当者", "自身の担当顧客・営業データの参照・更新"),
        ("閲覧者", "顧客・営業データの参照のみ"),
    ]
    for role_name, perm in roles_data:
        ws.cell(row=row, column=2).value = role_name
        _apply_style(ws.cell(row=row, column=2), font=FONT_BODY, border=BORDER_ALL_THIN, alignment=ALIGN_LEFT)
        ws.cell(row=row, column=3).value = perm
        _apply_style(ws.cell(row=row, column=3), font=FONT_BODY, border=BORDER_ALL_THIN, alignment=ALIGN_LEFT)
        row += 1

    row += 1

    row = _section(ws, row, "3.2 データ保護")
    row = _body(ws, row, "個人情報を含むデータは以下の方針に基づき保護する。")
    row = _numbered(ws, row, 1, "通信経路の暗号化（TLS 1.3）")
    row = _numbered(ws, row, 2, "保存データの暗号化（AES-256）")
    row = _numbered(ws, row, 3, "アクセスログの記録と定期監査")
    row = _numbered(ws, row, 4, "個人情報のマスキング表示")
    row += 1

    # --- Chapter 4 ---
    row = _chapter(ws, row, "第4章　非機能要件方針")
    row += 1

    row = _section(ws, row, "4.1 性能目標")
    row = _bullet(ws, row, "API応答時間：95パーセンタイルで500ms以内")
    row = _bullet(ws, row, "画面表示時間：3秒以内（初期表示）")
    row = _bullet(ws, row, "同時接続数：200ユーザー")
    row = _bullet(ws, row, "バッチ処理：日次バッチは6時間以内に完了")
    row += 1

    row = _section(ws, row, "4.2 可用性")
    row = _body(ws, row, "サービス稼働率99.9%を目標とし、以下の構成で可用性を確保する。")
    row = _bullet(ws, row, "Webサーバー：3台構成（Active-Active）")
    row = _bullet(ws, row, "DBサーバー：2台構成（Primary-Standby）")
    row = _bullet(ws, row, "自動フェイルオーバー機構の導入")
    row += 1

    row += 1
    ws.cell(row=row, column=4).value = "以上"
    _apply_style(ws.cell(row=row, column=4), font=FONT_BODY)

    wb.save(OUTPUT_DIR / "sample_hierarchical_text.xlsx")


# ---------------------------------------------------------------------------
# Generator registry
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# 15. 広幅フリーフォーム資料 (Wide freeform - 30+ columns)
# ---------------------------------------------------------------------------

def generate_wide_freeform():
    """社内研修資料 - 30列以上を使ったお絵描き帳スタイル"""
    wb = Workbook()
    ws = wb.active
    ws.title = "研修資料"

    # 列幅: 30列を3文字幅に設定（方眼紙的）
    for c in range(1, 35):
        ws.column_dimensions[get_column_letter(c)].width = 4.5

    # Start at row 3 (not A1) — realistic: margin rows at top
    row = 3

    # --- ヘッダーバー ---
    for c in range(1, 31):
        _apply_style(ws.cell(row=row, column=c), fill=FILL_DARK_BLUE)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=30)
    cell = ws.cell(row=row, column=1, value="株式会社ABCテクノロジー 2024年度 新人研修テキスト")
    _apply_style(cell, font=Font(name="Yu Gothic", size=18, bold=True, color="FFFFFF"),
                 alignment=ALIGN_CENTER)
    ws.row_dimensions[row].height = 45

    row = 5
    # --- サブタイトル（右寄せ） ---
    ws.merge_cells(start_row=row, start_column=20, end_row=row, end_column=30)
    cell = ws.cell(row=row, column=20, value="人事部 教育研修課")
    _apply_style(cell, font=Font(name="Yu Gothic", size=11, italic=True),
                 alignment=Alignment(horizontal="right", vertical="center"))
    row += 1
    ws.merge_cells(start_row=row, start_column=20, end_row=row, end_column=30)
    cell = ws.cell(row=row, column=20, value="最終更新: 2024年4月1日  Ver.3.2")
    _apply_style(cell, font=FONT_SMALL,
                 alignment=Alignment(horizontal="right", vertical="center"))

    # --- 第1章タイトル ---
    row = 6
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=15)
    cell = ws.cell(row=row, column=1, value="第1章  ビジネスマナーの基本")
    _apply_style(cell, font=Font(name="Yu Gothic", size=16, bold=True),
                 alignment=ALIGN_LEFT)
    ws.row_dimensions[row].height = 30
    # 右側に章のポイントボックス
    for r in range(row, row + 4):
        for c in range(18, 31):
            _apply_style(ws.cell(row=r, column=c), fill=FILL_LIGHT_YELLOW,
                         border=BORDER_ALL_THIN)
    ws.merge_cells(start_row=row, start_column=18, end_row=row, end_column=30)
    cell = ws.cell(row=row, column=18, value="★ この章のポイント")
    _apply_style(cell, font=FONT_BOLD, fill=FILL_ORANGE,
                 alignment=ALIGN_CENTER, border=BORDER_ALL_THIN)
    for i, point in enumerate(["① 第一印象は7秒で決まる",
                                "② 敬語は3種類を使い分ける",
                                "③ 報連相を徹底する"], start=1):
        ws.merge_cells(start_row=row + i, start_column=18, end_row=row + i, end_column=30)
        cell = ws.cell(row=row + i, column=18, value=point)
        _apply_style(cell, font=FONT_BODY, fill=FILL_LIGHT_YELLOW,
                     alignment=ALIGN_LEFT, border=BORDER_ALL_THIN)

    # --- 本文段落 ---
    row = 11
    paragraphs = [
        "社会人として最も重要なスキルの一つがビジネスマナーです。",
        "ビジネスマナーとは、仕事を円滑に進めるための共通のルールであり、",
        "相手への敬意と配慮を形にしたものです。",
        "",
        "特に新入社員の皆さんは、学生時代とは異なる振る舞いが求められます。",
        "以下の表で基本的なマナーを確認しましょう。",
    ]
    for text in paragraphs:
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=16)
        if text:
            cell = ws.cell(row=row, column=2, value=text)
            _apply_style(cell, font=FONT_BODY, alignment=ALIGN_LEFT)
        row += 1

    # --- 中サイズテーブル（B18:P23） 左半分 ---
    row += 1
    table_headers = ["場面", "正しいマナー", "NG例"]
    col_ranges = [(2, 5), (6, 11), (12, 16)]
    for (cs, ce), hdr in zip(col_ranges, table_headers):
        ws.merge_cells(start_row=row, start_column=cs, end_row=row, end_column=ce)
        cell = ws.cell(row=row, column=cs, value=hdr)
        _apply_style(cell, font=FONT_WHITE_BOLD, fill=FILL_DARK_BLUE,
                     alignment=ALIGN_CENTER, border=BORDER_ALL_THIN)
        for c in range(cs, ce + 1):
            ws.cell(row=row, column=c).border = BORDER_ALL_THIN

    table_data = [
        ("挨拶", "はっきりと「おはようございます」", "無言で通り過ぎる"),
        ("電話応対", "「お電話ありがとうございます」", "「はい」だけで出る"),
        ("メール", "件名を明確に、宛名から書く", "件名なし、いきなり本文"),
        ("名刺交換", "両手で受け取り、すぐにしまわない", "片手で受け取る"),
        ("報告", "結論→理由→詳細の順で話す", "時系列でダラダラ話す"),
    ]
    for data_row in table_data:
        row += 1
        for (cs, ce), val in zip(col_ranges, data_row):
            ws.merge_cells(start_row=row, start_column=cs, end_row=row, end_column=ce)
            cell = ws.cell(row=row, column=cs, value=val)
            _apply_style(cell, font=FONT_BODY, alignment=ALIGN_LEFT_TOP,
                         border=BORDER_ALL_THIN)
            for c in range(cs, ce + 1):
                ws.cell(row=row, column=c).border = BORDER_ALL_THIN

    # --- 右側に独立した「豆知識」ボックス ---
    box_row = row - 3
    for r in range(box_row, box_row + 5):
        for c in range(23, 31):
            _apply_style(ws.cell(row=r, column=c), fill=FILL_LIGHT_GREEN,
                         border=BORDER_ALL_THIN)
    ws.merge_cells(start_row=box_row, start_column=23, end_row=box_row, end_column=30)
    cell = ws.cell(row=box_row, column=23, value="💡 豆知識")
    _apply_style(cell, font=FONT_BOLD, fill=FILL_DARK_GREEN,
                 alignment=ALIGN_CENTER, border=BORDER_ALL_THIN)
    tips = [
        "メラビアンの法則によると、",
        "第一印象の55%は視覚情報、",
        "38%は聴覚情報、",
        "言語情報はわずか7%です。",
    ]
    for i, tip in enumerate(tips, start=1):
        ws.merge_cells(start_row=box_row + i, start_column=23, end_row=box_row + i, end_column=30)
        cell = ws.cell(row=box_row + i, column=23, value=tip)
        _apply_style(cell, font=FONT_SMALL, fill=FILL_LIGHT_GREEN,
                     alignment=ALIGN_LEFT, border=BORDER_ALL_THIN)

    # --- 第2章タイトル ---
    row += 3
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=15)
    cell = ws.cell(row=row, column=1, value="第2章  社内システムの使い方")
    _apply_style(cell, font=Font(name="Yu Gothic", size=16, bold=True),
                 alignment=ALIGN_LEFT)
    ws.row_dimensions[row].height = 30

    row += 2
    # 左側: 手順リスト
    steps = [
        ("2.1", "社内ポータルへのログイン"),
        ("2.2", "勤怠管理システム（TimePro）の操作"),
        ("2.3", "経費精算システムの申請方法"),
        ("2.4", "社内メール（Outlook）の設定"),
        ("2.5", "共有ドライブのフォルダ構成"),
    ]
    for num, title in steps:
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        cell = ws.cell(row=row, column=2, value=num)
        _apply_style(cell, font=FONT_BOLD, alignment=ALIGN_CENTER)
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=14)
        cell = ws.cell(row=row, column=4, value=title)
        _apply_style(cell, font=FONT_BODY, alignment=ALIGN_LEFT,
                     border=BORDER_BOTTOM_THIN)
        row += 1

    # 右側に同時にシステム一覧テーブル
    sys_row = row - 5
    sys_headers = ["システム名", "URL / アクセス方法", "用途"]
    sys_col_ranges = [(17, 20), (21, 27), (28, 30)]
    for (cs, ce), hdr in zip(sys_col_ranges, sys_headers):
        ws.merge_cells(start_row=sys_row, start_column=cs, end_row=sys_row, end_column=ce)
        cell = ws.cell(row=sys_row, column=cs, value=hdr)
        _apply_style(cell, font=FONT_WHITE_BOLD, fill=FILL_DARK_BLUE,
                     alignment=ALIGN_CENTER, border=BORDER_ALL_THIN)
        for c in range(cs, ce + 1):
            ws.cell(row=sys_row, column=c).border = BORDER_ALL_THIN

    sys_data = [
        ("社内ポータル", "portal.abc-tech.co.jp", "全社情報"),
        ("TimePro", "ポータル内リンク", "勤怠管理"),
        ("経費精算", "keihiseisan.abc-tech.co.jp", "経費申請"),
        ("Outlook", "デスクトップアプリ", "メール"),
    ]
    for sdata in sys_data:
        sys_row += 1
        for (cs, ce), val in zip(sys_col_ranges, sdata):
            ws.merge_cells(start_row=sys_row, start_column=cs, end_row=sys_row, end_column=ce)
            cell = ws.cell(row=sys_row, column=cs, value=val)
            _apply_style(cell, font=FONT_BODY, alignment=ALIGN_LEFT,
                         border=BORDER_ALL_THIN)
            for c in range(cs, ce + 1):
                ws.cell(row=sys_row, column=c).border = BORDER_ALL_THIN

    # --- フッター ---
    row += 3
    for c in range(1, 31):
        _apply_style(ws.cell(row=row, column=c), border=Border(top=MEDIUM))
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=15)
    cell = ws.cell(row=row, column=1, value="© 2024 株式会社ABCテクノロジー 人事部")
    _apply_style(cell, font=FONT_SMALL, alignment=ALIGN_LEFT)
    ws.merge_cells(start_row=row, start_column=20, end_row=row, end_column=30)
    cell = ws.cell(row=row, column=20, value="社外秘 - Confidential")
    _apply_style(cell, font=Font(name="Yu Gothic", size=9, bold=True, color="FF0000"),
                 alignment=Alignment(horizontal="right", vertical="center"))

    wb.save(OUTPUT_DIR / "sample_wide_freeform.xlsx")


# ---------------------------------------------------------------------------
# 16. 広幅ダッシュボード (Wide dashboard - 30+ columns)
# ---------------------------------------------------------------------------

def generate_wide_dashboard():
    """経営ダッシュボード - 30列を使ったKPIレイアウト"""
    wb = Workbook()
    ws = wb.active
    ws.title = "経営ダッシュボード"

    for c in range(1, 35):
        ws.column_dimensions[get_column_letter(c)].width = 4.5

    FILL_KPI_BG = PatternFill(start_color="F4F6F7", end_color="F4F6F7", fill_type="solid")
    FILL_CARD_BLUE = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")
    FILL_CARD_GREEN = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
    FILL_CARD_ORANGE = PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")
    FILL_CARD_RED = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")

    # 全体の背景
    for r in range(1, 45):
        for c in range(1, 31):
            ws.cell(row=r, column=c).fill = FILL_KPI_BG

    row = 1
    # --- タイトルバー ---
    ws.merge_cells("A1:AD1")
    cell = ws.cell(row=1, column=1, value="経営ダッシュボード  2024年度 第3四半期")
    _apply_style(cell, font=Font(name="Yu Gothic", size=16, bold=True, color="FFFFFF"),
                 fill=PatternFill(start_color="1B2631", end_color="1B2631", fill_type="solid"),
                 alignment=ALIGN_CENTER)
    ws.row_dimensions[1].height = 40

    # --- KPIカード4枚 横並び ---
    kpi_cards = [
        ("売上高", "¥2,340M", "前年比 +12.3%", FILL_CARD_BLUE),
        ("営業利益", "¥456M", "前年比 +8.7%", FILL_CARD_GREEN),
        ("受注残高", "¥1,890M", "前年比 -2.1%", FILL_CARD_ORANGE),
        ("従業員数", "1,247名", "前年比 +45名", FILL_CARD_RED),
    ]
    card_cols = [(1, 7), (8, 14), (15, 21), (22, 28)]
    for (cs, ce), (title, value, change, fill) in zip(card_cols, kpi_cards):
        # カードヘッダ
        for c in range(cs, ce + 1):
            ws.cell(row=3, column=c).fill = fill
            ws.cell(row=3, column=c).border = BORDER_ALL_THIN
        ws.merge_cells(start_row=3, start_column=cs, end_row=3, end_column=ce)
        cell = ws.cell(row=3, column=cs, value=title)
        _apply_style(cell, font=Font(name="Yu Gothic", size=10, bold=True, color="FFFFFF"),
                     fill=fill, alignment=ALIGN_CENTER, border=BORDER_ALL_THIN)
        # 値
        for c in range(cs, ce + 1):
            ws.cell(row=4, column=c).fill = FILL_WHITE
            ws.cell(row=4, column=c).border = BORDER_ALL_THIN
            ws.cell(row=5, column=c).fill = FILL_WHITE
            ws.cell(row=5, column=c).border = BORDER_ALL_THIN
        ws.merge_cells(start_row=4, start_column=cs, end_row=4, end_column=ce)
        cell = ws.cell(row=4, column=cs, value=value)
        _apply_style(cell, font=Font(name="Yu Gothic", size=20, bold=True),
                     fill=FILL_WHITE, alignment=ALIGN_CENTER, border=BORDER_ALL_THIN)
        ws.row_dimensions[4].height = 35
        ws.merge_cells(start_row=5, start_column=cs, end_row=5, end_column=ce)
        cell = ws.cell(row=5, column=cs, value=change)
        color = "27AE60" if "+" in change else "C0392B"
        _apply_style(cell, font=Font(name="Yu Gothic", size=9, color=color),
                     fill=FILL_WHITE, alignment=ALIGN_CENTER, border=BORDER_ALL_THIN)

    # --- 左側: 月別売上推移テーブル（A7:N18） ---
    row = 7
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
    cell = ws.cell(row=row, column=1, value="月別売上推移")
    _apply_style(cell, font=FONT_HEADING, fill=FILL_KPI_BG, alignment=ALIGN_LEFT)

    row = 8
    months_headers = ["", "4月", "5月", "6月", "7月", "8月", "9月",
                      "10月", "11月", "12月", "1月", "2月", "3月"]
    col_start = 1
    for i, h in enumerate(months_headers):
        ws.cell(row=row, column=col_start + i, value=h)
        _apply_style(ws.cell(row=row, column=col_start + i),
                     font=FONT_WHITE_BOLD, fill=FILL_DARK_BLUE,
                     alignment=ALIGN_CENTER, border=BORDER_ALL_THIN)

    sales_data = {
        "製品A": [180, 195, 210, 188, 205, 220, 215, 230, 198, 210, 225, 240],
        "製品B": [120, 115, 130, 125, 140, 135, 128, 145, 150, 138, 142, 155],
        "製品C": [85, 90, 78, 92, 88, 95, 100, 105, 98, 102, 108, 112],
        "サービス": [60, 65, 70, 68, 72, 75, 78, 80, 82, 85, 88, 90],
    }
    for category, values in sales_data.items():
        row += 1
        ws.cell(row=row, column=1, value=category)
        _apply_style(ws.cell(row=row, column=1), font=FONT_BOLD,
                     alignment=ALIGN_LEFT, border=BORDER_ALL_THIN)
        for i, v in enumerate(values):
            ws.cell(row=row, column=2 + i, value=v)
            _apply_style(ws.cell(row=row, column=2 + i), font=FONT_BODY,
                         alignment=ALIGN_RIGHT, border=BORDER_ALL_THIN)

    row += 1
    ws.cell(row=row, column=1, value="合計")
    _apply_style(ws.cell(row=row, column=1), font=FONT_BOLD,
                 fill=FILL_LIGHT_GRAY, alignment=ALIGN_LEFT, border=BORDER_ALL_THIN)
    for i in range(12):
        total = sum(sales_data[k][i] for k in sales_data)
        ws.cell(row=row, column=2 + i, value=total)
        _apply_style(ws.cell(row=row, column=2 + i), font=FONT_BOLD,
                     fill=FILL_LIGHT_GRAY, alignment=ALIGN_RIGHT, border=BORDER_ALL_THIN)

    # --- 右側: 部門別構成比（P7:AD14） ---
    dept_row = 7
    ws.merge_cells(start_row=dept_row, start_column=16, end_row=dept_row, end_column=30)
    cell = ws.cell(row=dept_row, column=16, value="部門別売上構成比")
    _apply_style(cell, font=FONT_HEADING, fill=FILL_KPI_BG, alignment=ALIGN_LEFT)

    dept_row = 8
    dept_headers = ["部門", "売上(M)", "構成比", "目標", "達成率"]
    dept_col_ranges = [(16, 18), (19, 21), (22, 23), (24, 26), (27, 29)]
    for (cs, ce), h in zip(dept_col_ranges, dept_headers):
        ws.merge_cells(start_row=dept_row, start_column=cs, end_row=dept_row, end_column=ce)
        cell = ws.cell(row=dept_row, column=cs, value=h)
        _apply_style(cell, font=FONT_WHITE_BOLD, fill=FILL_DARK_BLUE,
                     alignment=ALIGN_CENTER, border=BORDER_ALL_THIN)
        for c in range(cs, ce + 1):
            ws.cell(row=dept_row, column=c).border = BORDER_ALL_THIN

    dept_data = [
        ("東日本営業", 890, "38.0%", 850, "104.7%"),
        ("西日本営業", 720, "30.8%", 750, "96.0%"),
        ("海外事業", 430, "18.4%", 400, "107.5%"),
        ("ソリューション", 300, "12.8%", 320, "93.8%"),
    ]
    for ddata in dept_data:
        dept_row += 1
        for (cs, ce), val in zip(dept_col_ranges, ddata):
            ws.merge_cells(start_row=dept_row, start_column=cs, end_row=dept_row, end_column=ce)
            cell = ws.cell(row=dept_row, column=cs, value=val)
            font = FONT_BODY
            if isinstance(val, str) and "%" in val:
                try:
                    pct = float(val.replace("%", ""))
                    if pct >= 100:
                        font = Font(name="Yu Gothic", size=10, color="27AE60")
                    else:
                        font = Font(name="Yu Gothic", size=10, color="C0392B")
                except ValueError:
                    pass
            _apply_style(cell, font=font, alignment=ALIGN_CENTER, border=BORDER_ALL_THIN)
            for c in range(cs, ce + 1):
                ws.cell(row=dept_row, column=c).border = BORDER_ALL_THIN

    # --- 下段左: 注目トピック（A15:N22） ---
    topic_row = row + 2
    ws.merge_cells(start_row=topic_row, start_column=1, end_row=topic_row, end_column=14)
    cell = ws.cell(row=topic_row, column=1, value="今月の注目トピック")
    _apply_style(cell, font=FONT_HEADING, fill=FILL_KPI_BG, alignment=ALIGN_LEFT)

    topics = [
        "• 新製品「CloudManager Pro」の受注が好調（前月比+35%）",
        "• 西日本エリアで大型案件（¥120M）の提案中、来月クロージング予定",
        "• 海外事業部がASEAN市場で新規パートナー3社と契約締結",
        "• DX推進プロジェクトの第1フェーズが予定通り完了",
        "• 人材採用: エンジニア15名の中途採用が完了（充足率94%）",
    ]
    for topic in topics:
        topic_row += 1
        ws.merge_cells(start_row=topic_row, start_column=2, end_row=topic_row, end_column=14)
        cell = ws.cell(row=topic_row, column=2, value=topic)
        _apply_style(cell, font=FONT_BODY, fill=FILL_KPI_BG, alignment=ALIGN_LEFT)

    # --- 下段右: リスク・課題（P15:AD22） ---
    risk_row = topic_row - len(topics)
    ws.merge_cells(start_row=risk_row, start_column=16, end_row=risk_row, end_column=30)
    cell = ws.cell(row=risk_row, column=16, value="リスク・課題")
    _apply_style(cell, font=Font(name="Yu Gothic", size=12, bold=True, color="C0392B"),
                 fill=FILL_KPI_BG, alignment=ALIGN_LEFT)

    risk_headers = ["項目", "影響度", "対策状況"]
    risk_col_ranges = [(16, 21), (22, 24), (25, 30)]
    risk_row += 1
    for (cs, ce), h in zip(risk_col_ranges, risk_headers):
        ws.merge_cells(start_row=risk_row, start_column=cs, end_row=risk_row, end_column=ce)
        cell = ws.cell(row=risk_row, column=cs, value=h)
        _apply_style(cell, font=FONT_WHITE_BOLD, fill=FILL_CARD_RED,
                     alignment=ALIGN_CENTER, border=BORDER_ALL_THIN)
        for c in range(cs, ce + 1):
            ws.cell(row=risk_row, column=c).border = BORDER_ALL_THIN

    risks = [
        ("半導体不足による納期遅延", "高", "代替サプライヤー選定中"),
        ("為替変動（円安）", "中", "ヘッジ比率80%に引上げ"),
        ("競合X社の価格攻勢", "中", "付加価値提案を強化"),
        ("基幹システム老朽化", "高", "2025年度刷新予算承認済"),
    ]
    for rdata in risks:
        risk_row += 1
        for (cs, ce), val in zip(risk_col_ranges, rdata):
            ws.merge_cells(start_row=risk_row, start_column=cs, end_row=risk_row, end_column=ce)
            cell = ws.cell(row=risk_row, column=cs, value=val)
            fill = FILL_KPI_BG
            if val == "高":
                fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
            elif val == "中":
                fill = FILL_LIGHT_YELLOW
            _apply_style(cell, font=FONT_BODY, fill=fill,
                         alignment=ALIGN_LEFT, border=BORDER_ALL_THIN)
            for c in range(cs, ce + 1):
                ws.cell(row=risk_row, column=c).border = BORDER_ALL_THIN
                ws.cell(row=risk_row, column=c).fill = fill

    wb.save(OUTPUT_DIR / "sample_wide_dashboard.xlsx")


# ---------------------------------------------------------------------------
# 17. 広幅レイアウト文書 (Wide layout doc - 30+ columns, mixed content)
# ---------------------------------------------------------------------------

def generate_wide_proposal():
    """提案書 - 30列使った本格的な提案資料レイアウト"""
    wb = Workbook()
    ws = wb.active
    ws.title = "提案書"

    for c in range(1, 35):
        ws.column_dimensions[get_column_letter(c)].width = 4.0

    row = 1
    # --- 表紙風ヘッダー ---
    for r in range(1, 6):
        for c in range(1, 31):
            ws.cell(row=r, column=c).fill = PatternFill(
                start_color="1A5276", end_color="1A5276", fill_type="solid")

    ws.merge_cells("A2:AD2")
    cell = ws.cell(row=2, column=1, value="社内DX推進に関するご提案")
    _apply_style(cell, font=Font(name="Yu Gothic", size=22, bold=True, color="FFFFFF"),
                 alignment=ALIGN_CENTER)
    ws.row_dimensions[2].height = 45

    ws.merge_cells("A4:AD4")
    cell = ws.cell(row=4, column=1, value="株式会社ABCテクノロジー  ソリューション事業部")
    _apply_style(cell, font=Font(name="Yu Gothic", size=12, color="FFFFFF"),
                 alignment=ALIGN_CENTER)

    # --- 目次 ---
    row = 7
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    cell = ws.cell(row=row, column=1, value="目次")
    _apply_style(cell, font=Font(name="Yu Gothic", size=14, bold=True),
                 alignment=ALIGN_LEFT)

    toc = [
        "1. 背景と課題認識 .......................... P.2",
        "2. 提案概要 .................................... P.3",
        "3. 導入スケジュール ...................... P.5",
        "4. 費用見積 .................................... P.6",
        "5. 期待効果 .................................... P.7",
    ]
    for item in toc:
        row += 1
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=15)
        cell = ws.cell(row=row, column=3, value=item)
        _apply_style(cell, font=FONT_BODY, alignment=ALIGN_LEFT)

    # 目次の右側に提案サマリーボックス
    for r in range(7, 13):
        for c in range(17, 31):
            ws.cell(row=r, column=c).fill = FILL_LIGHT_BLUE
            ws.cell(row=r, column=c).border = BORDER_ALL_THIN
    ws.merge_cells(start_row=7, start_column=17, end_row=7, end_column=30)
    cell = ws.cell(row=7, column=17, value="提案サマリー")
    _apply_style(cell, font=FONT_WHITE_BOLD, fill=FILL_DARK_BLUE,
                 alignment=ALIGN_CENTER, border=BORDER_ALL_THIN)
    summary_items = [
        ("提案内容", "業務プロセスのデジタル化"),
        ("対象範囲", "営業部門・管理部門"),
        ("導入期間", "6ヶ月（2024年7月〜12月）"),
        ("概算費用", "¥48,000,000（税別）"),
        ("期待ROI", "初年度 150%"),
    ]
    for i, (key, val) in enumerate(summary_items):
        r = 8 + i
        ws.merge_cells(start_row=r, start_column=17, end_row=r, end_column=20)
        cell = ws.cell(row=r, column=17, value=key)
        _apply_style(cell, font=FONT_BOLD, fill=FILL_LIGHT_BLUE,
                     alignment=ALIGN_LEFT, border=BORDER_ALL_THIN)
        ws.merge_cells(start_row=r, start_column=21, end_row=r, end_column=30)
        cell = ws.cell(row=r, column=21, value=val)
        _apply_style(cell, font=FONT_BODY, fill=FILL_LIGHT_BLUE,
                     alignment=ALIGN_LEFT, border=BORDER_ALL_THIN)

    # --- セクション1: 背景と課題認識 ---
    row = 15
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=30)
    cell = ws.cell(row=row, column=1, value="1. 背景と課題認識")
    _apply_style(cell, font=Font(name="Yu Gothic", size=14, bold=True),
                 alignment=ALIGN_LEFT,
                 border=Border(bottom=Side(style="medium", color="1A5276")))
    for c in range(1, 31):
        ws.cell(row=row, column=c).border = Border(
            bottom=Side(style="medium", color="1A5276"))

    row += 2
    bg_text = [
        "貴社における現状の業務プロセスを調査した結果、以下の課題が確認されました。",
        "特にExcelベースの業務フローが全体の約70%を占めており、データの二重入力や",
        "バージョン管理の困難さが業務効率を大きく低下させています。",
    ]
    for text in bg_text:
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=29)
        cell = ws.cell(row=row, column=2, value=text)
        _apply_style(cell, font=FONT_BODY, alignment=ALIGN_LEFT)
        row += 1

    # 課題テーブル（幅広）
    row += 1
    issue_headers = ["No.", "課題カテゴリ", "現状の問題点", "影響範囲", "緊急度"]
    issue_col_ranges = [(2, 3), (4, 9), (10, 20), (21, 25), (26, 29)]
    for (cs, ce), h in zip(issue_col_ranges, issue_headers):
        ws.merge_cells(start_row=row, start_column=cs, end_row=row, end_column=ce)
        cell = ws.cell(row=row, column=cs, value=h)
        _apply_style(cell, font=FONT_WHITE_BOLD,
                     fill=PatternFill(start_color="1A5276", end_color="1A5276", fill_type="solid"),
                     alignment=ALIGN_CENTER, border=BORDER_ALL_THIN)
        for c in range(cs, ce + 1):
            ws.cell(row=row, column=c).border = BORDER_ALL_THIN

    issues = [
        ("1", "データ管理", "Excelファイルが部門ごとに散在し、最新版の特定が困難",
         "全部門", "高"),
        ("2", "承認プロセス", "紙の回覧による承認で平均5営業日を要する",
         "管理部門", "高"),
        ("3", "レポート作成", "月次報告書の作成に各部門で延べ40時間/月を消費",
         "営業・経理", "中"),
        ("4", "情報共有", "ナレッジが個人のPC内に属人化している",
         "全部門", "中"),
        ("5", "セキュリティ", "USB持ち出しによる情報漏洩リスクが存在",
         "全部門", "高"),
    ]
    for idata in issues:
        row += 1
        for (cs, ce), val in zip(issue_col_ranges, idata):
            ws.merge_cells(start_row=row, start_column=cs, end_row=row, end_column=ce)
            cell = ws.cell(row=row, column=cs, value=val)
            fill = FILL_WHITE
            if val == "高":
                fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
            elif val == "中":
                fill = FILL_LIGHT_YELLOW
            _apply_style(cell, font=FONT_BODY, fill=fill,
                         alignment=ALIGN_CENTER if cs > 9 else ALIGN_LEFT,
                         border=BORDER_ALL_THIN)
            for c in range(cs, ce + 1):
                ws.cell(row=row, column=c).border = BORDER_ALL_THIN
                if fill != FILL_WHITE:
                    ws.cell(row=row, column=c).fill = fill

    # --- セクション2: 費用見積 ---
    row += 3
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=30)
    cell = ws.cell(row=row, column=1, value="4. 費用見積（税別）")
    _apply_style(cell, font=Font(name="Yu Gothic", size=14, bold=True),
                 alignment=ALIGN_LEFT,
                 border=Border(bottom=Side(style="medium", color="1A5276")))
    for c in range(1, 31):
        ws.cell(row=row, column=c).border = Border(
            bottom=Side(style="medium", color="1A5276"))

    row += 2
    cost_headers = ["費目", "内容", "数量", "単価（千円）", "金額（千円）", "備考"]
    cost_col_ranges = [(2, 5), (6, 13), (14, 16), (17, 20), (21, 24), (25, 29)]
    for (cs, ce), h in zip(cost_col_ranges, cost_headers):
        ws.merge_cells(start_row=row, start_column=cs, end_row=row, end_column=ce)
        cell = ws.cell(row=row, column=cs, value=h)
        _apply_style(cell, font=FONT_WHITE_BOLD,
                     fill=PatternFill(start_color="1A5276", end_color="1A5276", fill_type="solid"),
                     alignment=ALIGN_CENTER, border=BORDER_ALL_THIN)
        for c in range(cs, ce + 1):
            ws.cell(row=row, column=c).border = BORDER_ALL_THIN

    costs = [
        ("ライセンス", "クラウドサービス利用料（年間）", "200ユーザー", 24, 4800, "初年度"),
        ("構築費", "システム設計・開発", "一式", None, 18000, ""),
        ("移行費", "データ移行・クレンジング", "一式", None, 8000, ""),
        ("教育費", "ユーザートレーニング", "10回", 200, 2000, "各部門"),
        ("PM費", "プロジェクト管理", "6ヶ月", 2500, 15000, ""),
    ]
    for cdata in costs:
        row += 1
        for (cs, ce), val in zip(cost_col_ranges, cdata):
            ws.merge_cells(start_row=row, start_column=cs, end_row=row, end_column=ce)
            display = val if val is not None else "—"
            cell = ws.cell(row=row, column=cs, value=display)
            align = ALIGN_RIGHT if isinstance(val, (int, float)) else ALIGN_LEFT
            _apply_style(cell, font=FONT_BODY, alignment=align, border=BORDER_ALL_THIN)
            for c in range(cs, ce + 1):
                ws.cell(row=row, column=c).border = BORDER_ALL_THIN

    # 合計行
    row += 1
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=20)
    cell = ws.cell(row=row, column=2, value="合計")
    _apply_style(cell, font=FONT_BOLD, fill=FILL_LIGHT_GRAY,
                 alignment=Alignment(horizontal="right", vertical="center"),
                 border=BORDER_ALL_THIN)
    for c in range(2, 21):
        ws.cell(row=row, column=c).fill = FILL_LIGHT_GRAY
        ws.cell(row=row, column=c).border = BORDER_ALL_THIN
    ws.merge_cells(start_row=row, start_column=21, end_row=row, end_column=24)
    cell = ws.cell(row=row, column=21, value=47800)
    _apply_style(cell, font=Font(name="Yu Gothic", size=12, bold=True),
                 fill=FILL_LIGHT_GRAY, alignment=ALIGN_RIGHT, border=BORDER_ALL_THIN)
    for c in range(21, 25):
        ws.cell(row=row, column=c).fill = FILL_LIGHT_GRAY
        ws.cell(row=row, column=c).border = BORDER_ALL_THIN
    ws.merge_cells(start_row=row, start_column=25, end_row=row, end_column=29)
    cell = ws.cell(row=row, column=25, value="税別")
    _apply_style(cell, font=FONT_BOLD, fill=FILL_LIGHT_GRAY,
                 alignment=ALIGN_LEFT, border=BORDER_ALL_THIN)
    for c in range(25, 30):
        ws.cell(row=row, column=c).fill = FILL_LIGHT_GRAY
        ws.cell(row=row, column=c).border = BORDER_ALL_THIN

    wb.save(OUTPUT_DIR / "sample_wide_proposal.xlsx")


# ---------------------------------------------------------------------------
# Phase 2a samples
# ---------------------------------------------------------------------------

def generate_bullet_list():
    """Generate sample with bullet/numbered lists."""
    wb = Workbook()
    ws = wb.active
    ws.title = "箇条書きテスト"

    RO = 2  # row offset

    # Title
    ws.merge_cells(f"B{RO}:F{RO}")
    c = ws[f"B{RO}"]
    c.value = "プロジェクト活動報告"
    _apply_style(c, font=FONT_TITLE)

    row = RO + 2
    # Section heading
    ws.cell(row=row, column=2, value="1. 今週の成果").font = Font(name="Yu Gothic", size=11, bold=True)
    row += 1

    # Bullet list with ・
    items_bullet = [
        "・新規APIエンドポイントの設計完了",
        "・データベーススキーマのレビュー実施",
        "・フロントエンド画面のプロトタイプ作成",
        "・セキュリティ監査チェックリストの更新",
    ]
    for item in items_bullet:
        ws.cell(row=row, column=3, value=item).font = FONT_BODY
        row += 1

    row += 1
    ws.cell(row=row, column=2, value="2. 来週の予定").font = Font(name="Yu Gothic", size=11, bold=True)
    row += 1

    # Numbered list
    items_numbered = [
        "1) 結合テスト環境の構築",
        "2) パフォーマンステストの実施",
        "3) ドキュメントの最終レビュー",
    ]
    for item in items_numbered:
        ws.cell(row=row, column=3, value=item).font = FONT_BODY
        row += 1

    row += 1
    ws.cell(row=row, column=2, value="3. 課題・リスク").font = Font(name="Yu Gothic", size=11, bold=True)
    row += 1

    # Mixed markers
    items_mixed = [
        "● サーバーリソースの追加が必要（見積もり中）",
        "● 外部APIの仕様変更への対応が遅延",
        "※ 予算超過の可能性あり。上長承認が必要",
    ]
    for item in items_mixed:
        ws.cell(row=row, column=3, value=item).font = FONT_BODY
        row += 1

    row += 1
    ws.cell(row=row, column=2, value="4. 参加者への依頼事項").font = Font(name="Yu Gothic", size=11, bold=True)
    row += 1

    items_parenthesized = [
        "(1) テスト仕様書のレビューをお願いします",
        "(2) 来週月曜までにフィードバックをお送りください",
        "(3) 環境構築手順書を確認してください",
    ]
    for item in items_parenthesized:
        ws.cell(row=row, column=3, value=item).font = FONT_BODY
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 50

    wb.save(OUTPUT_DIR / "sample_bullet_list.xlsx")


def generate_borderless_table():
    """Generate sample with borderless tables (no borders, but tabular data)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "罫線なし表"

    RO = 2

    # Title
    ws.merge_cells(f"A{RO}:E{RO}")
    c = ws[f"A{RO}"]
    c.value = "社員名簿（簡易版）"
    _apply_style(c, font=FONT_TITLE)

    row = RO + 2
    # Borderless table — header row (no borders, just bold)
    headers = ["社員番号", "氏名", "部署", "入社年", "年収（万円）"]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col_idx, value=h)
        c.font = FONT_BOLD
    row += 1

    # Data rows — no borders at all
    data = [
        ("E001", "田中太郎", "開発部", "2018", "650"),
        ("E002", "佐藤花子", "営業部", "2019", "580"),
        ("E003", "鈴木一郎", "人事部", "2020", "520"),
        ("E004", "高橋美咲", "開発部", "2017", "720"),
        ("E005", "伊藤健太", "経理部", "2021", "480"),
        ("E006", "渡辺由美", "営業部", "2018", "610"),
        ("E007", "山本拓也", "開発部", "2016", "780"),
    ]
    for d in data:
        for col_idx, v in enumerate(d, start=1):
            ws.cell(row=row, column=col_idx, value=v).font = FONT_BODY
        row += 1

    # Column widths
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 15

    # Second borderless table on another sheet
    ws2 = wb.create_sheet("売上サマリ")

    ws2.merge_cells("B2:E2")
    c = ws2["B2"]
    c.value = "月別売上サマリ"
    _apply_style(c, font=FONT_TITLE)

    row = 4
    headers2 = ["月", "売上（百万円）", "前年比", "達成率"]
    for col_idx, h in enumerate(headers2, start=2):
        c = ws2.cell(row=row, column=col_idx, value=h)
        c.font = FONT_BOLD
    row += 1

    monthly = [
        ("4月", "125.3", "102%", "98%"),
        ("5月", "132.1", "105%", "103%"),
        ("6月", "118.7", "96%", "92%"),
        ("7月", "145.2", "110%", "113%"),
        ("8月", "138.9", "108%", "108%"),
        ("9月", "151.4", "112%", "118%"),
    ]
    for m in monthly:
        for col_idx, v in enumerate(m, start=2):
            ws2.cell(row=row, column=col_idx, value=v).font = FONT_BODY
        row += 1

    for col in range(2, 6):
        ws2.column_dimensions[get_column_letter(col)].width = 18

    wb.save(OUTPUT_DIR / "sample_borderless_table.xlsx")


def generate_implicit_index():
    """Generate sample with implicit row indices (grouped rows)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "グループ化データ"

    RO = 2

    # Title
    ws.merge_cells(f"A{RO}:E{RO}")
    c = ws[f"A{RO}"]
    c.value = "部署別プロジェクト一覧"
    _apply_style(c, font=FONT_TITLE)

    row = RO + 2
    # Header
    headers = ["部署", "プロジェクト名", "担当者", "進捗", "備考"]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col_idx, value=h)
        _apply_style(c, font=FONT_BOLD, fill=FILL_LIGHT_BLUE, border=BORDER_ALL_THIN)
    row += 1

    # Data with implicit row indices
    # "開発部" only appears once, the next rows have blank first column
    data = [
        ("開発部", "新規ECサイト", "田中", "80%", ""),
        ("", "顧客管理CRM", "佐藤", "45%", "要件変更あり"),
        ("", "社内ポータル", "鈴木", "100%", "完了"),
        ("営業部", "販促キャンペーン", "高橋", "60%", ""),
        ("", "代理店開拓", "伊藤", "30%", "遅延"),
        ("人事部", "採用管理システム", "渡辺", "90%", ""),
        ("", "研修プログラム", "山本", "70%", ""),
        ("", "人事評価制度", "中村", "50%", ""),
    ]
    for d in data:
        for col_idx, v in enumerate(d, start=1):
            c = ws.cell(row=row, column=col_idx, value=v)
            _apply_style(c, font=FONT_BODY, border=BORDER_ALL_THIN)
        row += 1

    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 20

    wb.save(OUTPUT_DIR / "sample_implicit_index.xlsx")


def generate_table_with_footer():
    """Generate sample with table footer rows (合計, 小計)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "売上集計"

    RO = 2

    ws.merge_cells(f"A{RO}:E{RO}")
    c = ws[f"A{RO}"]
    c.value = "2026年度 四半期売上集計"
    _apply_style(c, font=FONT_TITLE)

    row = RO + 2
    headers = ["カテゴリ", "Q1", "Q2", "Q3", "Q4"]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col_idx, value=h)
        _apply_style(c, font=FONT_BOLD, fill=FILL_LIGHT_BLUE, border=BORDER_ALL_THIN)
    row += 1

    data = [
        ("ソフトウェア", 1200, 1350, 1180, 1420),
        ("ハードウェア", 800, 750, 820, 900),
        ("サービス", 450, 520, 480, 550),
        ("コンサルティング", 300, 280, 350, 320),
    ]
    for d in data:
        for col_idx, v in enumerate(d, start=1):
            c = ws.cell(row=row, column=col_idx, value=v)
            _apply_style(c, font=FONT_BODY, border=BORDER_ALL_THIN)
            if col_idx >= 2:
                c.number_format = '#,##0'
        row += 1

    # Footer row — 合計 (with borders and bold)
    totals = ("合計", 2750, 2900, 2830, 3190)
    for col_idx, v in enumerate(totals, start=1):
        c = ws.cell(row=row, column=col_idx, value=v)
        _apply_style(c, font=FONT_BOLD, border=BORDER_ALL_THIN, fill=FILL_LIGHT_YELLOW)
        if col_idx >= 2:
            c.number_format = '#,##0'

    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 18

    wb.save(OUTPUT_DIR / "sample_table_footer.xlsx")


def generate_number_formats():
    """Generate sample with various number formats."""
    wb = Workbook()
    ws = wb.active
    ws.title = "数値フォーマット"

    # Title
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "各種数値フォーマットのテスト"
    _apply_style(c, font=FONT_TITLE)

    row = 3
    headers = ["項目", "値", "フォーマット種別"]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col_idx, value=h)
        _apply_style(c, font=FONT_BOLD, fill=FILL_LIGHT_BLUE, border=BORDER_ALL_THIN)
    row += 1

    formats = [
        ("通常数値", 12345, '#,##0', "number"),
        ("小数点", 3.14159, '0.00', "decimal"),
        ("パーセント", 0.856, '0.0%', "percentage"),
        ("日本円", 1500000, '¥#,##0', "currency"),
        ("USドル", 299.99, '$#,##0.00', "currency"),
        ("日付", 46068, 'yyyy/mm/dd', "date"),
        ("日時", 46068.75, 'yyyy/mm/dd hh:mm', "datetime"),
        ("時刻", 0.75, 'hh:mm:ss', "time"),
        ("和暦", 46068, 'ge.m.d', "date"),
        ("万円単位", 35000, '#,##0"万円"', "currency"),
    ]
    for label, val, fmt, type_name in formats:
        c1 = ws.cell(row=row, column=1, value=label)
        _apply_style(c1, font=FONT_BODY, border=BORDER_ALL_THIN)
        c2 = ws.cell(row=row, column=2, value=val)
        _apply_style(c2, font=FONT_BODY, border=BORDER_ALL_THIN, number_format=fmt)
        c3 = ws.cell(row=row, column=3, value=type_name)
        _apply_style(c3, font=FONT_BODY, border=BORDER_ALL_THIN)
        row += 1

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20

    wb.save(OUTPUT_DIR / "sample_number_formats.xlsx")


GENERATORS = [
    ("sample_houganshi.xlsx", "Excel方眼紙（稟議書）", generate_houganshi),
    ("sample_review.xlsx", "レビュー資料", generate_review),
    ("sample_schedule.xlsx", "工程表（ガントチャート）", generate_schedule),
    ("sample_spec.xlsx", "仕様書", generate_spec),
    ("sample_ledger.xlsx", "管理台帳", generate_ledger),
    ("sample_minutes.xlsx", "議事録", generate_minutes),
    ("sample_test_spec.xlsx", "テスト仕様書", generate_test_spec),
    ("sample_budget.xlsx", "予算管理表", generate_budget),
    ("sample_freetext.xlsx", "ただ文章を書いただけ（社内通知）", generate_freetext),
    ("sample_freeform_report.xlsx", "お絵描き帳スタイルの報告書", generate_freeform_report),
    ("sample_sketchpad.xlsx", "完全自由配置（オリエンテーション資料）", generate_sketchpad),
    ("sample_with_images.xlsx", "画像入りドキュメント（作業手順書）", generate_with_images),
    ("sample_mixed_document.xlsx", "文章+テーブル+グラフ（営業実績分析）", generate_mixed_document),
    ("sample_hierarchical_text.xlsx", "ワード風階層構造テキスト（設計方針書）", generate_hierarchical_text),
    ("sample_wide_freeform.xlsx", "広幅フリーフォーム（新人研修資料・30列）", generate_wide_freeform),
    ("sample_wide_dashboard.xlsx", "広幅ダッシュボード（経営KPI・30列）", generate_wide_dashboard),
    ("sample_wide_proposal.xlsx", "広幅提案書（DX提案資料・30列）", generate_wide_proposal),
    ("sample_bullet_list.xlsx", "箇条書きリスト（活動報告）", generate_bullet_list),
    ("sample_borderless_table.xlsx", "罫線なしテーブル（社員名簿）", generate_borderless_table),
    ("sample_implicit_index.xlsx", "暗黙的行インデックス（グループ化）", generate_implicit_index),
    ("sample_table_footer.xlsx", "テーブルフッター（売上集計）", generate_table_with_footer),
    ("sample_number_formats.xlsx", "数値フォーマット各種", generate_number_formats),
]


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    print(f"Output directory: {OUTPUT_DIR}")
    print(f"Generating {len(GENERATORS)} sample Excel files...\n")

    for filename, label, gen_func in GENERATORS:
        print(f"  Generating {filename} ({label}) ...", end=" ", flush=True)
        gen_func()
        print("done")

    print(f"\nAll files generated in {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
