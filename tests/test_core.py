"""Tests for xlmelt core modules."""

import json
import tempfile
from pathlib import Path

import openpyxl
import pytest

from xlmelt.core.analyzer import StructureAnalyzer
from xlmelt.core.model import (
    CellInfo,
    DocumentModel,
    Region,
    Section,
    SectionType,
    SheetModel,
)
from xlmelt.core.parser import ExcelParser
from xlmelt.output.html_writer import HtmlWriter
from xlmelt.output.json_writer import JsonWriter


@pytest.fixture
def simple_xlsx(tmp_path: Path) -> Path:
    """Create a simple Excel file for testing."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Title row - bold, large, merged
    ws.merge_cells("A1:D1")
    cell = ws["A1"]
    cell.value = "テスト文書"
    cell.font = openpyxl.styles.Font(size=18, bold=True)

    # Key-value pairs
    for row, (key, val) in enumerate([
        ("プロジェクト名", "テストプロジェクト"),
        ("作成日", "2026-03-14"),
    ], start=3):
        k_cell = ws.cell(row=row, column=1, value=key)
        k_cell.font = openpyxl.styles.Font(bold=True)
        ws.cell(row=row, column=2, value=val)

    # Table with borders
    headers = ["No.", "項目", "値"]
    thin = openpyxl.styles.Side(style="thin")
    border = openpyxl.styles.Border(top=thin, bottom=thin, left=thin, right=thin)

    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=6, column=col, value=h)
        c.font = openpyxl.styles.Font(bold=True)
        c.border = border

    for row, (no, item, val) in enumerate([
        ("1", "テスト項目A", "100"),
        ("2", "テスト項目B", "200"),
    ], start=7):
        for col, v in enumerate([no, item, val], start=1):
            c = ws.cell(row=row, column=col, value=v)
            c.border = border

    path = tmp_path / "test.xlsx"
    wb.save(str(path))
    return path


@pytest.fixture
def multi_sheet_xlsx(tmp_path: Path) -> Path:
    """Create a multi-sheet Excel file."""
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "概要"
    ws1["A1"] = "概要シート"
    ws1["A1"].font = openpyxl.styles.Font(size=16, bold=True)

    ws2 = wb.create_sheet("データ")
    ws2["A1"] = "データシート"
    ws2["A1"].font = openpyxl.styles.Font(size=16, bold=True)

    path = tmp_path / "multi.xlsx"
    wb.save(str(path))
    return path


class TestModel:
    def test_section_to_dict(self):
        s = Section(
            type=SectionType.HEADING,
            level=1,
            title="テスト見出し",
            source_region=Region(1, 1, 1, 4),
        )
        d = s.to_dict()
        assert d["type"] == "heading"
        assert d["level"] == 1
        assert d["title"] == "テスト見出し"

    def test_document_to_dict(self):
        doc = DocumentModel(
            title="テスト",
            source_file="test.xlsx",
            sheets=[
                SheetModel(name="Sheet1", sections=[
                    Section(type=SectionType.TEXT, content="テスト内容"),
                ]),
            ],
        )
        d = doc.to_dict()
        assert d["document"]["title"] == "テスト"
        assert len(d["document"]["sheets"]) == 1
        assert d["document"]["sheets"][0]["sections"][0]["type"] == "text"

    def test_region_contains(self):
        r = Region(1, 1, 5, 5)
        assert r.contains(3, 3)
        assert not r.contains(6, 3)
        assert r.width == 5
        assert r.height == 5


class TestParser:
    def test_parse_simple(self, simple_xlsx: Path):
        with ExcelParser(simple_xlsx) as parser:
            assert "Sheet1" in parser.sheet_names
            grid, rows, cols = parser.parse_sheet("Sheet1")
            assert rows > 0
            assert cols > 0
            # Check title cell
            title_cell = grid[1][1]
            assert title_cell is not None
            assert title_cell.value == "テスト文書"
            assert title_cell.font_bold
            assert title_cell.font_size == 18
            assert title_cell.is_merged_origin
            assert title_cell.merge_width == 4

    def test_parse_multi_sheet(self, multi_sheet_xlsx: Path):
        with ExcelParser(multi_sheet_xlsx) as parser:
            assert len(parser.sheet_names) == 2
            assert "概要" in parser.sheet_names
            assert "データ" in parser.sheet_names

    def test_column_widths(self, simple_xlsx: Path):
        with ExcelParser(simple_xlsx) as parser:
            widths = parser.get_column_widths("Sheet1")
            assert len(widths) > 0
            assert all(w > 0 for w in widths)


class TestAnalyzer:
    def test_analyze_simple(self, simple_xlsx: Path):
        analyzer = StructureAnalyzer()
        doc = analyzer.analyze(simple_xlsx)
        assert doc.title == "test"
        assert len(doc.sheets) == 1

        sheet = doc.sheets[0]
        types = [s.type for s in sheet.sections]

        # Should detect heading, key-value, and table
        assert SectionType.HEADING in types
        assert SectionType.KEY_VALUE in types
        assert SectionType.TABLE in types

    def test_heading_detection(self, simple_xlsx: Path):
        analyzer = StructureAnalyzer()
        doc = analyzer.analyze(simple_xlsx)
        headings = [s for s in doc.sheets[0].sections if s.type == SectionType.HEADING]
        assert len(headings) >= 1
        assert headings[0].title == "テスト文書"
        assert headings[0].level == 1  # font_size=18

    def test_table_detection(self, simple_xlsx: Path):
        analyzer = StructureAnalyzer()
        doc = analyzer.analyze(simple_xlsx)
        tables = [s for s in doc.sheets[0].sections if s.type == SectionType.TABLE]
        assert len(tables) >= 1
        table = tables[0]
        assert "headers" in table.content
        assert "rows" in table.content
        assert len(table.content["rows"]) == 2

    def test_kv_detection(self, simple_xlsx: Path):
        analyzer = StructureAnalyzer()
        doc = analyzer.analyze(simple_xlsx)
        kvs = [s for s in doc.sheets[0].sections if s.type == SectionType.KEY_VALUE]
        assert len(kvs) >= 1
        kv = kvs[0]
        assert isinstance(kv.content, dict)
        assert "プロジェクト名" in kv.content

    def test_analyze_multi_sheet(self, multi_sheet_xlsx: Path):
        analyzer = StructureAnalyzer()
        doc = analyzer.analyze(multi_sheet_xlsx)
        assert len(doc.sheets) == 2


class TestJsonWriter:
    def test_write(self, simple_xlsx: Path, tmp_path: Path):
        analyzer = StructureAnalyzer()
        doc = analyzer.analyze(simple_xlsx)
        writer = JsonWriter()
        out = tmp_path / "output.json"
        writer.write(doc, out)
        assert out.exists()

        with open(out, encoding="utf-8") as f:
            data = json.load(f)
        assert "document" in data
        assert data["document"]["title"] == "test"

    def test_to_string(self, simple_xlsx: Path):
        analyzer = StructureAnalyzer()
        doc = analyzer.analyze(simple_xlsx)
        writer = JsonWriter()
        s = writer.to_string(doc)
        data = json.loads(s)
        assert "document" in data


class TestHtmlWriter:
    def test_write(self, simple_xlsx: Path, tmp_path: Path):
        analyzer = StructureAnalyzer()
        doc = analyzer.analyze(simple_xlsx)
        writer = HtmlWriter()
        out = tmp_path / "output.html"
        writer.write(doc, out)
        assert out.exists()

        content = out.read_text(encoding="utf-8")
        assert "<!DOCTYPE html>" in content
        assert "テスト文書" in content

    def test_no_style(self, simple_xlsx: Path):
        analyzer = StructureAnalyzer()
        doc = analyzer.analyze(simple_xlsx)
        writer = HtmlWriter(include_style=False)
        html = writer.to_string(doc)
        assert "<style>" not in html
        assert "<!DOCTYPE html>" in html


class TestListDetection:
    """Tests for bullet/numbered list detection."""

    def test_list_markers(self, tmp_path: Path):
        """Various list markers should be detected as LIST sections."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        ws["A1"] = "タイトル"
        ws["A1"].font = openpyxl.styles.Font(size=16, bold=True)

        markers = [
            "・項目A", "● 項目B", "1) 項目C", "(1) 項目D",
            "- 項目E", "※ 注意事項",
        ]
        for i, item in enumerate(markers, start=3):
            ws.cell(row=i, column=1, value=item)

        path = tmp_path / "list_test.xlsx"
        wb.save(str(path))

        analyzer = StructureAnalyzer()
        doc = analyzer.analyze(path)
        lists = [s for s in doc.sheets[0].sections if s.type == SectionType.LIST]
        assert len(lists) >= 1
        content = lists[0].content
        assert isinstance(content, dict)
        assert "items" in content
        assert "ordered" in content
        items = content["items"]
        assert len(items) == 6
        # Markers should be stripped
        assert not items[0].startswith("・")
        assert not items[2].startswith("1)")

    def test_non_list_text(self, tmp_path: Path):
        """Regular text should not be detected as list."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "これは通常のテキストです"
        ws["A2"] = "リストマーカーはありません"

        path = tmp_path / "nonlist.xlsx"
        wb.save(str(path))

        analyzer = StructureAnalyzer()
        doc = analyzer.analyze(path)
        lists = [s for s in doc.sheets[0].sections if s.type == SectionType.LIST]
        assert len(lists) == 0


class TestImplicitRowIndex:
    """Tests for implicit row index filling."""

    def test_fill_blank_first_column(self, tmp_path: Path):
        """Blank first column should be filled from the row above."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        thin = openpyxl.styles.Side(style="thin")
        border = openpyxl.styles.Border(top=thin, bottom=thin, left=thin, right=thin)

        # Header
        for col, h in enumerate(["部署", "名前", "役割"], start=1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = openpyxl.styles.Font(bold=True)
            c.border = border

        # Data with implicit index
        data = [
            ("開発部", "田中", "リーダー"),
            ("", "佐藤", "メンバー"),
            ("営業部", "鈴木", "部長"),
            ("", "高橋", "メンバー"),
        ]
        for i, (dept, name, role) in enumerate(data, start=2):
            ws.cell(row=i, column=1, value=dept).border = border
            ws.cell(row=i, column=2, value=name).border = border
            ws.cell(row=i, column=3, value=role).border = border

        path = tmp_path / "implicit.xlsx"
        wb.save(str(path))

        analyzer = StructureAnalyzer()
        doc = analyzer.analyze(path)
        tables = [s for s in doc.sheets[0].sections if s.type == SectionType.TABLE]
        assert len(tables) == 1
        rows = tables[0].content["rows"]
        # Second row should have "開発部" filled in
        assert rows[1][0] == "開発部"
        # Fourth row should have "営業部" filled in
        assert rows[3][0] == "営業部"


class TestNumberFormat:
    """Tests for number format type inference."""

    def test_format_inference(self):
        from xlmelt.core.analyzer import _infer_format_type
        assert _infer_format_type("General") is None
        assert _infer_format_type(None) is None
        assert _infer_format_type("#,##0") == "number"
        assert _infer_format_type("0.00") == "decimal"
        assert _infer_format_type("0.0%") == "percentage"
        assert _infer_format_type("¥#,##0") == "currency"
        assert _infer_format_type("$#,##0.00") == "currency"
        assert _infer_format_type("yyyy/mm/dd") == "date"
        assert _infer_format_type("yyyy/mm/dd hh:mm") == "datetime"
        assert _infer_format_type("hh:mm:ss") == "time"


class TestSampleFiles:
    """Test with actual generated sample files (if available)."""

    SAMPLES_DIR = Path(__file__).parent.parent / "samples" / "output"

    @pytest.mark.skipif(
        not (Path(__file__).parent.parent / "samples" / "output").exists(),
        reason="Sample files not generated",
    )
    def test_all_samples_convert(self):
        """Ensure all sample Excel files can be converted without errors."""
        analyzer = StructureAnalyzer()
        xlsx_files = list(self.SAMPLES_DIR.glob("*.xlsx"))
        assert len(xlsx_files) > 0, "No sample files found"

        for xlsx_file in xlsx_files:
            doc = analyzer.analyze(xlsx_file)
            assert len(doc.sheets) > 0, f"{xlsx_file.name} has no sheets"
            for sheet in doc.sheets:
                assert len(sheet.sections) > 0, (
                    f"{xlsx_file.name}/{sheet.name} has no sections"
                )

            # Verify JSON serialization
            writer = JsonWriter()
            json_str = writer.to_string(doc)
            data = json.loads(json_str)
            assert "document" in data

            # Verify HTML generation
            html_writer = HtmlWriter()
            html = html_writer.to_string(doc)
            assert "<!DOCTYPE html>" in html


class TestVerify:
    """Tests for JSON↔HTML verification."""

    def test_render_html_from_json_matches_original(self):
        """JSON→HTML re-render should exactly match original HTML."""
        from xlmelt.verify import render_html_from_json

        doc = DocumentModel(
            title="Test",
            source_file="test.xlsx",
            sheets=[
                SheetModel(
                    name="Sheet1",
                    sections=[
                        Section(type=SectionType.HEADING, level=2, title="Heading"),
                        Section(type=SectionType.KEY_VALUE, content={"Key": "Val"}),
                        Section(type=SectionType.TABLE, content={
                            "headers": ["A", "B"],
                            "rows": [["1", "2"]],
                        }),
                        Section(type=SectionType.LIST, content={
                            "ordered": False,
                            "items": ["item1", "item2"],
                        }),
                        Section(type=SectionType.TEXT, content="Hello\nWorld"),
                    ],
                )
            ],
        )

        writer = HtmlWriter()
        original_html = writer.to_string(doc)

        json_writer = JsonWriter()
        json_str = json_writer.to_string(doc)
        json_data = json.loads(json_str)

        rerendered = render_html_from_json(json_data)

        orig_lines = [l.strip() for l in original_html.splitlines() if l.strip()]
        new_lines = [l.strip() for l in rerendered.splitlines() if l.strip()]
        assert orig_lines == new_lines

    def test_verify_json_html_passes(self):
        """verify_json_html should pass for consistent data."""
        from xlmelt.verify import verify_json_html

        doc = DocumentModel(
            title="Test",
            source_file="test.xlsx",
            sheets=[
                SheetModel(
                    name="Sheet1",
                    sections=[
                        Section(type=SectionType.HEADING, level=3, title="H3"),
                        Section(type=SectionType.TABLE, content={
                            "headers": ["X"], "rows": [["v"]],
                        }),
                    ],
                )
            ],
        )

        writer = HtmlWriter()
        html = writer.to_string(doc)
        json_data = json.loads(JsonWriter().to_string(doc))

        result = verify_json_html(json_data, html)
        assert result.ok
        assert len(result.failed) == 0

    def test_verify_detects_missing_title(self):
        """verify should flag a heading with no title."""
        from xlmelt.verify import verify_json_html

        json_data = {
            "document": {
                "title": "T",
                "source": "t.xlsx",
                "sheets": [{
                    "name": "S",
                    "sections": [{"type": "heading", "level": 2}],
                }],
            }
        }
        html = "<html></html>"
        result = verify_json_html(json_data, html)
        assert any("no title" in f for f in result.failed)

    SAMPLES_DIR = Path(__file__).parent.parent / "samples" / "output"

    @pytest.mark.skipif(
        not (Path(__file__).parent.parent / "samples" / "output").exists(),
        reason="Sample files not found",
    )
    def test_all_samples_verify(self):
        """All sample files should pass JSON↔HTML verification."""
        from xlmelt.verify import verify_json_html

        analyzer = StructureAnalyzer()
        xlsx_files = list(self.SAMPLES_DIR.glob("*.xlsx"))
        assert len(xlsx_files) > 0

        for xlsx_file in xlsx_files:
            doc = analyzer.analyze(xlsx_file)
            html = HtmlWriter().to_string(doc)
            json_data = json.loads(JsonWriter().to_string(doc))
            result = verify_json_html(json_data, html)
            assert result.ok, f"{xlsx_file.name}: {result.summary()}"
