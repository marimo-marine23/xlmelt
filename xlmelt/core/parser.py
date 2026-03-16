"""Excel file parser using openpyxl."""

from __future__ import annotations

import zipfile
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .model import CellInfo, ImageInfo

# Safety limits for grid allocation to prevent memory exhaustion
MAX_GRID_ROWS = 10_000
MAX_GRID_COLS = 1_000
MAX_IMAGE_SIZE = 50 * 1024 * 1024  # 50 MB per image


def _sanitize_filename(name: str) -> str:
    """Sanitize a string for use as a filename component."""
    import re as _re
    # Replace problematic characters with underscore
    return _re.sub(r'[\\/:*?"<>|\s]', '_', name).strip('_')


def _color_to_hex(color: Any) -> str | None:
    """Convert openpyxl color to hex string."""
    if color is None:
        return None
    if hasattr(color, "rgb") and color.rgb and color.rgb != "00000000":
        rgb = str(color.rgb)
        # openpyxl returns ARGB format
        if len(rgb) == 8:
            return f"#{rgb[2:]}"
        if len(rgb) == 6:
            return f"#{rgb}"
    return None


def _has_border(side: Any) -> bool:
    """Check if a border side is present."""
    if side is None:
        return False
    return side.style is not None and side.style != "none"


class ExcelParser:
    """Parse Excel files and extract cell information."""

    def __init__(self, file_path: str | Path):
        self.file_path = Path(file_path)
        self._wb: openpyxl.Workbook | None = None

    def open(self) -> None:
        """Open the workbook."""
        self._wb = openpyxl.load_workbook(
            str(self.file_path), data_only=True, read_only=False
        )

    def close(self) -> None:
        """Close the workbook."""
        if self._wb:
            self._wb.close()
            self._wb = None

    def __enter__(self) -> ExcelParser:
        self.open()
        return self

    def __exit__(self, *args: Any) -> None:
        self.close()

    @property
    def wb(self) -> openpyxl.Workbook:
        if self._wb is None:
            raise RuntimeError("Workbook not opened. Call open() first.")
        return self._wb

    @property
    def sheet_names(self) -> list[str]:
        return self.wb.sheetnames

    def _build_merge_map(self, ws: Worksheet) -> dict[tuple[int, int], tuple[int, int, int, int]]:
        """Build a map of merged cells: (row, col) -> (min_row, min_col, max_row, max_col)."""
        merge_map: dict[tuple[int, int], tuple[int, int, int, int]] = {}
        for merge_range in ws.merged_cells.ranges:
            for row in range(merge_range.min_row, merge_range.max_row + 1):
                for col in range(merge_range.min_col, merge_range.max_col + 1):
                    merge_map[(row, col)] = (
                        merge_range.min_row,
                        merge_range.min_col,
                        merge_range.max_row,
                        merge_range.max_col,
                    )
        return merge_map

    def parse_sheet(self, sheet_name: str) -> tuple[list[list[CellInfo | None]], int, int]:
        """Parse a sheet into a 2D grid of CellInfo.

        Returns (grid, row_count, col_count).
        Grid is 1-indexed: grid[row][col], with row/col starting at 1.
        """
        ws = self.wb[sheet_name]
        merge_map = self._build_merge_map(ws)

        max_row = min(ws.max_row or 0, MAX_GRID_ROWS)
        max_col = min(ws.max_column or 0, MAX_GRID_COLS)

        # Find actual used range (skip trailing empty rows/cols)
        actual_max_row = 0
        actual_max_col = 0
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None or (row, col) in merge_map:
                    actual_max_row = max(actual_max_row, row)
                    actual_max_col = max(actual_max_col, col)

        # Also account for merge ranges extending beyond content
        for merge_range in ws.merged_cells.ranges:
            actual_max_row = max(actual_max_row, merge_range.max_row)
            actual_max_col = max(actual_max_col, merge_range.max_col)

        if actual_max_row == 0:
            return [], 0, 0

        # Safety: cap grid dimensions to prevent memory exhaustion
        if actual_max_row > MAX_GRID_ROWS or actual_max_col > MAX_GRID_COLS:
            import warnings
            warnings.warn(
                f"Sheet '{sheet_name}' dimensions ({actual_max_row}x{actual_max_col}) "
                f"exceed safety limits ({MAX_GRID_ROWS}x{MAX_GRID_COLS}), truncating"
            )
            actual_max_row = min(actual_max_row, MAX_GRID_ROWS)
            actual_max_col = min(actual_max_col, MAX_GRID_COLS)

        # Build grid (1-indexed, so index 0 is unused)
        grid: list[list[CellInfo | None]] = [
            [None] * (actual_max_col + 1) for _ in range(actual_max_row + 1)
        ]

        for row in range(1, actual_max_row + 1):
            for col in range(1, actual_max_col + 1):
                cell = ws.cell(row=row, column=col)
                info = self._parse_cell(cell, row, col, merge_map)
                grid[row][col] = info

        return grid, actual_max_row, actual_max_col

    def _parse_cell(
        self,
        cell: Cell | MergedCell,
        row: int,
        col: int,
        merge_map: dict[tuple[int, int], tuple[int, int, int, int]],
    ) -> CellInfo:
        """Extract CellInfo from an openpyxl cell."""
        merge_info = merge_map.get((row, col))
        is_merged = merge_info is not None
        is_origin = False
        merge_width = 1
        merge_height = 1

        if merge_info:
            min_r, min_c, max_r, max_c = merge_info
            is_origin = row == min_r and col == min_c
            if is_origin:
                merge_width = max_c - min_c + 1
                merge_height = max_r - min_r + 1

        # For merged cells that aren't the origin, get value from origin
        value = None
        font_size = None
        font_bold = False
        font_color = None
        fill_color = None
        border_top = False
        border_bottom = False
        border_left = False
        border_right = False
        number_format = None
        h_align = None
        v_align = None
        indent = 0

        if isinstance(cell, MergedCell):
            # MergedCell has limited attributes; mark as merged non-origin
            pass
        else:
            value = cell.value
            if value is not None:
                value = str(value)

            font = cell.font
            if font:
                font_size = font.size
                font_bold = bool(font.bold)
                font_color = _color_to_hex(font.color)

            fill = cell.fill
            if fill and fill.fgColor:
                fill_color = _color_to_hex(fill.fgColor)

            border = cell.border
            if border:
                border_top = _has_border(border.top)
                border_bottom = _has_border(border.bottom)
                border_left = _has_border(border.left)
                border_right = _has_border(border.right)

            number_format = cell.number_format
            alignment = cell.alignment
            if alignment:
                h_align = alignment.horizontal
                v_align = alignment.vertical
                indent = alignment.indent or 0

        return CellInfo(
            row=row,
            col=col,
            value=value,
            font_size=font_size,
            font_bold=font_bold,
            font_color=font_color,
            fill_color=fill_color,
            border_top=border_top,
            border_bottom=border_bottom,
            border_left=border_left,
            border_right=border_right,
            merge_width=merge_width,
            merge_height=merge_height,
            is_merged_origin=is_origin,
            is_merged_cell=is_merged,
            number_format=number_format,
            alignment_horizontal=h_align,
            alignment_vertical=v_align,
            indent=indent,
        )

    def extract_images(self, output_dir: Path) -> list[ImageInfo]:
        """Extract embedded images and chart placeholders from the Excel file.

        Parses relationship files to map images to sheets and anchor cells.
        Also detects chart objects and creates placeholder entries.
        """
        images: list[ImageInfo] = []
        images_dir = output_dir / "images"

        try:
            with zipfile.ZipFile(str(self.file_path), "r") as zf:
                image_files = [
                    f for f in zf.namelist()
                    if f.startswith("xl/media/")
                ]

                # Build media filename → sheet mapping from relationships
                media_to_sheet = self._build_image_sheet_map(zf)
                # Build media filename → anchor cell mapping from drawings
                media_to_anchor = self._build_image_anchor_map(zf)

                if image_files:
                    images_dir.mkdir(parents=True, exist_ok=True)
                    for i, img_path in enumerate(image_files):
                        ext = Path(img_path).suffix.lower()
                        if ext not in (".png", ".jpg", ".jpeg", ".gif", ".bmp", ".tiff", ".emf", ".wmf"):
                            continue

                        media_name = Path(img_path).name
                        sheet_name = media_to_sheet.get(media_name, "")
                        anchor_cell = media_to_anchor.get(media_name)

                        safe_sheet = _sanitize_filename(sheet_name) if sheet_name else ""
                        if safe_sheet:
                            out_name = f"{safe_sheet}_img{i + 1:03d}{ext}"
                        else:
                            out_name = f"image_{i + 1:03d}{ext}"

                        out_path = images_dir / out_name
                        with zf.open(img_path) as src, open(out_path, "wb") as dst:
                            data = src.read(MAX_IMAGE_SIZE + 1)
                            if len(data) > MAX_IMAGE_SIZE:
                                import warnings
                                warnings.warn(
                                    f"Image '{img_path}' exceeds {MAX_IMAGE_SIZE // (1024*1024)}MB limit, skipping"
                                )
                                continue
                            dst.write(data)

                        images.append(ImageInfo(
                            path=f"images/{out_name}",
                            format=ext.lstrip("."),
                            sheet_name=sheet_name,
                            anchor_cell=anchor_cell,
                        ))

                # Try to render charts as PNG, fallback to placeholders
                try:
                    from .chart_renderer import extract_and_render_charts
                    rendered = extract_and_render_charts(
                        self.file_path, output_dir, wb=self._wb,
                    )
                    for rc in rendered:
                        images.append(ImageInfo(
                            path=rc["chart_path"],
                            format="png",
                            sheet_name=self._resolve_sheet_name(rc.get("sheet_name", "")),
                            anchor_cell=rc.get("anchor_cell"),
                            alt_text=rc.get("title", "[Chart]"),
                        ))
                except ImportError:
                    # matplotlib not available — use placeholders
                    chart_infos = self._extract_chart_info(zf)
                    for ci in chart_infos:
                        images.append(ci)

        except zipfile.BadZipFile:
            pass

        return images

    def _extract_chart_info(self, zf: zipfile.ZipFile) -> list[ImageInfo]:
        """Extract chart metadata (title, anchor) from drawing XML."""
        import xml.etree.ElementTree as ET

        charts: list[ImageInfo] = []
        xdr_ns = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
        c_ns = "http://schemas.openxmlformats.org/drawingml/2006/chart"
        a_ns = "http://schemas.openxmlformats.org/drawingml/2006/main"
        r_ns = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
        rel_ns = "http://schemas.openxmlformats.org/package/2006/relationships"

        # Build sheet mapping for drawings
        sheet_for_drawing = self._build_drawing_to_sheet_map(zf)

        for name in zf.namelist():
            if not name.startswith("xl/drawings/") or not name.endswith(".xml"):
                continue
            if "_rels" in name:
                continue

            drawing_name = Path(name).name
            sheet_name = sheet_for_drawing.get(drawing_name, "")

            # Load drawing rels to find chart targets
            rels_path = name.replace("xl/drawings/", "xl/drawings/_rels/") + ".rels"
            rid_to_chart: dict[str, str] = {}
            if rels_path in zf.namelist():
                try:
                    rels_tree = ET.parse(zf.open(rels_path))
                    for rel in rels_tree.getroot().findall(f"{{{rel_ns}}}Relationship"):
                        rid = rel.get("Id", "")
                        target = rel.get("Target", "")
                        if "chart" in target.lower():
                            rid_to_chart[rid] = target
                except Exception:
                    continue

            if not rid_to_chart:
                continue

            try:
                tree = ET.parse(zf.open(name))
                root = tree.getroot()

                for anchor_tag in (f"{{{xdr_ns}}}twoCellAnchor", f"{{{xdr_ns}}}oneCellAnchor"):
                    for anchor in root.findall(anchor_tag):
                        # Get from cell position
                        from_el = anchor.find(f"{{{xdr_ns}}}from")
                        cell_ref = None
                        if from_el is not None:
                            col_el = from_el.find(f"{{{xdr_ns}}}col")
                            row_el = from_el.find(f"{{{xdr_ns}}}row")
                            if col_el is not None and row_el is not None:
                                col = int(col_el.text or "0")
                                row = int(row_el.text or "0")
                                cell_ref = f"{get_column_letter(col + 1)}{row + 1}"

                        # Check for chart reference in graphicFrame
                        for chart_el in anchor.iter(f"{{{c_ns}}}chart"):
                            rid = chart_el.get(f"{{{r_ns}}}id", "")
                            if rid in rid_to_chart:
                                chart_path = rid_to_chart[rid]
                                # Try to get chart title
                                title = self._get_chart_title(zf, chart_path, name)
                                charts.append(ImageInfo(
                                    path="",  # No image file for charts
                                    format="chart",
                                    sheet_name=sheet_name,
                                    anchor_cell=cell_ref,
                                    alt_text=title or "[Chart]",
                                ))
            except Exception:
                continue

        return charts

    def _get_chart_title(self, zf: zipfile.ZipFile, chart_target: str,
                         drawing_path: str) -> str | None:
        """Try to extract chart title from chart XML."""
        import xml.etree.ElementTree as ET

        # Resolve relative path from drawing
        if chart_target.startswith("../"):
            chart_path = "xl/" + chart_target.lstrip("../")
        elif chart_target.startswith("/"):
            chart_path = chart_target.lstrip("/")
        else:
            chart_path = str(Path(drawing_path).parent / chart_target)

        c_ns = "http://schemas.openxmlformats.org/drawingml/2006/chart"
        a_ns = "http://schemas.openxmlformats.org/drawingml/2006/main"

        try:
            if chart_path not in zf.namelist():
                return None
            tree = ET.parse(zf.open(chart_path))
            root = tree.getroot()
            # Look for c:title → c:tx → c:rich → a:p → a:r → a:t
            for title_el in root.iter(f"{{{c_ns}}}title"):
                for t_el in title_el.iter(f"{{{a_ns}}}t"):
                    if t_el.text:
                        return t_el.text.strip()
        except Exception:
            pass
        return None

    def _build_drawing_to_sheet_map(self, zf: zipfile.ZipFile) -> dict[str, str]:
        """Map drawing filenames to sheet names."""
        import xml.etree.ElementTree as ET

        drawing_to_sheet: dict[str, str] = {}
        rel_ns = "http://schemas.openxmlformats.org/package/2006/relationships"

        for name in zf.namelist():
            if not name.startswith("xl/worksheets/_rels/") or not name.endswith(".xml.rels"):
                continue
            sheet_file = Path(name).stem
            sheet_name = self._sheet_file_to_name(sheet_file)
            if not sheet_name:
                continue
            try:
                tree = ET.parse(zf.open(name))
                for rel in tree.getroot().findall(f"{{{rel_ns}}}Relationship"):
                    target = rel.get("Target", "")
                    if "drawing" in target.lower():
                        drawing_to_sheet[Path(target).name] = sheet_name
            except Exception:
                continue

        return drawing_to_sheet

    def _build_image_sheet_map(self, zf: zipfile.ZipFile) -> dict[str, str]:
        """Map media filenames to sheet names via worksheet → drawing → media chain."""
        import xml.etree.ElementTree as ET

        media_to_sheet: dict[str, str] = {}
        rel_ns = "http://schemas.openxmlformats.org/package/2006/relationships"

        # Step 1: Build sheet → drawing mapping from worksheet rels
        sheet_to_drawings: dict[str, list[str]] = {}
        for name in zf.namelist():
            if not name.startswith("xl/worksheets/_rels/") or not name.endswith(".xml.rels"):
                continue
            sheet_file = Path(name).stem  # "sheet1.xml"
            sheet_name = self._sheet_file_to_name(sheet_file)
            if not sheet_name:
                continue
            try:
                tree = ET.parse(zf.open(name))
                for rel in tree.getroot().findall(f"{{{rel_ns}}}Relationship"):
                    target = rel.get("Target", "")
                    if "drawing" in target.lower():
                        drawing_name = Path(target).name
                        sheet_to_drawings.setdefault(sheet_name, []).append(drawing_name)
                    # Also handle direct media references (some Excel versions)
                    media_name = self._extract_media_name(target)
                    if media_name:
                        media_to_sheet[media_name] = sheet_name
            except Exception:
                continue

        # Step 2: Build drawing → media mapping from drawing rels
        drawing_to_media: dict[str, list[str]] = {}
        for name in zf.namelist():
            if not name.startswith("xl/drawings/_rels/") or not name.endswith(".xml.rels"):
                continue
            drawing_name = Path(name).stem  # "drawing1.xml"
            try:
                tree = ET.parse(zf.open(name))
                for rel in tree.getroot().findall(f"{{{rel_ns}}}Relationship"):
                    target = rel.get("Target", "")
                    media_name = self._extract_media_name(target)
                    if media_name:
                        drawing_to_media.setdefault(drawing_name, []).append(media_name)
            except Exception:
                continue

        # Step 3: Chain: sheet → drawing → media
        for sheet_name, drawings in sheet_to_drawings.items():
            for drawing in drawings:
                # drawing filename may be "drawing1.xml", rels stem is "drawing1.xml"
                for media_name in drawing_to_media.get(drawing, []):
                    media_to_sheet[media_name] = sheet_name

        return media_to_sheet

    @staticmethod
    def _extract_media_name(target: str) -> str | None:
        """Extract media filename from a relationship target path.

        Handles both relative (../media/image1.png) and absolute (/xl/media/image1.png).
        """
        if "/media/" in target or target.startswith("../media/"):
            return Path(target).name
        return None

    def _resolve_sheet_name(self, name: str) -> str:
        """Resolve sheet name, including __sheet_idx_N placeholders."""
        if name.startswith("__sheet_idx_") and self._wb:
            try:
                idx = int(name.split("_")[-1])
                if 0 <= idx < len(self._wb.sheetnames):
                    return self._wb.sheetnames[idx]
            except (ValueError, IndexError):
                pass
        return name

    def _sheet_file_to_name(self, sheet_file: str) -> str | None:
        """Convert sheet filename (e.g., 'sheet1.xml') to sheet name."""
        # The sheet file embedded in .rels has a format like 'sheet1.xml.rels'
        # We need to strip '.rels' to get 'sheet1.xml'
        clean = sheet_file.replace(".rels", "") if sheet_file.endswith(".rels") else sheet_file
        # Try to match by index
        import re as _re
        m = _re.search(r"sheet(\d+)", clean)
        if m and self._wb:
            idx = int(m.group(1)) - 1
            names = self._wb.sheetnames
            if 0 <= idx < len(names):
                return names[idx]
        return None

    def _build_image_anchor_map(self, zf: zipfile.ZipFile) -> dict[str, str | None]:
        """Map media filenames to anchor cell positions from drawing XML."""
        import xml.etree.ElementTree as ET

        media_to_anchor: dict[str, str | None] = {}
        xdr_ns = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
        a_ns = "http://schemas.openxmlformats.org/drawingml/2006/main"
        r_ns = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
        rel_ns = "http://schemas.openxmlformats.org/package/2006/relationships"

        for name in zf.namelist():
            if not name.startswith("xl/drawings/") or not name.endswith(".xml"):
                continue
            if "_rels" in name:
                continue

            # Load drawing rels for this drawing
            rels_path = name.replace("xl/drawings/", "xl/drawings/_rels/") + ".rels"
            rid_to_media: dict[str, str] = {}
            if rels_path in zf.namelist():
                try:
                    rels_tree = ET.parse(zf.open(rels_path))
                    for rel in rels_tree.getroot().findall(f"{{{rel_ns}}}Relationship"):
                        rid = rel.get("Id", "")
                        target = rel.get("Target", "")
                        media_name = self._extract_media_name(target)
                        if media_name:
                            rid_to_media[rid] = media_name
                except Exception:
                    continue

            if not rid_to_media:
                continue

            try:
                tree = ET.parse(zf.open(name))
                root = tree.getroot()

                # Process twoCellAnchor and oneCellAnchor elements
                for anchor_tag in (f"{{{xdr_ns}}}twoCellAnchor", f"{{{xdr_ns}}}oneCellAnchor"):
                    for anchor in root.findall(anchor_tag):
                        # Get from cell position
                        from_el = anchor.find(f"{{{xdr_ns}}}from")
                        if from_el is not None:
                            col_el = from_el.find(f"{{{xdr_ns}}}col")
                            row_el = from_el.find(f"{{{xdr_ns}}}row")
                            if col_el is not None and row_el is not None:
                                col = int(col_el.text or "0")
                                row = int(row_el.text or "0")
                                cell_ref = f"{get_column_letter(col + 1)}{row + 1}"
                            else:
                                cell_ref = None
                        else:
                            cell_ref = None

                        # Find the blipFill → blip to get rId
                        for blip in anchor.iter(f"{{{a_ns}}}blip"):
                            rid = blip.get(f"{{{r_ns}}}embed", "")
                            if rid in rid_to_media:
                                media_to_anchor[rid_to_media[rid]] = cell_ref
            except Exception:
                continue

        return media_to_anchor

    def get_column_widths(self, sheet_name: str) -> list[float]:
        """Get column widths (useful for detecting Excel方眼紙)."""
        ws = self.wb[sheet_name]
        widths: list[float] = []
        max_col = ws.max_column or 0
        for col in range(1, max_col + 1):
            letter = get_column_letter(col)
            dim = ws.column_dimensions.get(letter)
            if dim and dim.width is not None:
                widths.append(dim.width)
            else:
                widths.append(8.43)  # Excel default width
        return widths
