"""Chart renderer - extract chart data from xlsx and render as PNG using matplotlib."""

from __future__ import annotations

import re
import xml.etree.ElementTree as ET
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any


C_NS = "http://schemas.openxmlformats.org/drawingml/2006/chart"
A_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"
R_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"


@dataclass
class ChartSeries:
    """A single data series in a chart."""

    name: str = ""
    categories: list[str] = field(default_factory=list)
    values: list[float] = field(default_factory=list)
    color: str | None = None  # hex color e.g. "#4F81BD"


@dataclass
class ChartData:
    """Parsed chart data."""

    title: str = ""
    chart_type: str = ""  # bar, line, pie, area, scatter
    bar_direction: str = "col"  # col or bar (horizontal)
    grouping: str = "clustered"  # clustered, stacked, percentStacked
    series: list[ChartSeries] = field(default_factory=list)
    cat_axis_title: str = ""
    val_axis_title: str = ""


def extract_and_render_charts(
    xlsx_path: Path,
    output_dir: Path,
    wb: Any = None,
) -> list[dict]:
    """Extract charts from xlsx and render as PNG.

    Returns list of dicts: {"chart_path": str, "sheet_name": str, "anchor_cell": str, "title": str}
    """
    try:
        import matplotlib
        matplotlib.use("Agg")  # Non-interactive backend
    except ImportError:
        return []

    results: list[dict] = []

    try:
        with zipfile.ZipFile(str(xlsx_path), "r") as zf:
            chart_files = [f for f in zf.namelist() if f.startswith("xl/charts/chart") and f.endswith(".xml")]
            if not chart_files:
                return results

            # Build chart → drawing → sheet mapping
            chart_to_anchor = _build_chart_anchor_map(zf)
            chart_to_sheet = _build_chart_sheet_map(zf)

            images_dir = output_dir / "images"
            images_dir.mkdir(parents=True, exist_ok=True)

            theme_colors = _extract_theme_colors(zf)

            for chart_file in sorted(chart_files):
                chart_name = Path(chart_file).stem  # "chart1"
                chart_data = _parse_chart_xml(zf, chart_file, wb, theme_colors)
                if not chart_data or not chart_data.series:
                    continue

                # Render chart
                out_name = f"chart_{chart_name}.png"
                out_path = images_dir / out_name
                _render_chart(chart_data, out_path)

                anchor = chart_to_anchor.get(chart_name)
                sheet = chart_to_sheet.get(chart_name, "")

                results.append({
                    "chart_path": f"images/{out_name}",
                    "sheet_name": sheet,
                    "anchor_cell": anchor,
                    "title": chart_data.title,
                })

    except (zipfile.BadZipFile, Exception):
        pass

    return results


def _extract_theme_colors(zf: zipfile.ZipFile) -> list[str]:
    """Extract accent colors from the workbook theme. Returns list of hex colors."""
    for name in zf.namelist():
        if "theme" in name and name.endswith(".xml"):
            try:
                tree = ET.parse(zf.open(name))
                root = tree.getroot()
                colors: list[str] = []
                for i in range(1, 7):  # accent1..accent6
                    for acc in root.iter(f"{{{A_NS}}}accent{i}"):
                        srgb = acc.find(f"{{{A_NS}}}srgbClr")
                        if srgb is not None:
                            colors.append(f"#{srgb.get('val', '000000')}")
                        break
                if colors:
                    return colors
            except Exception:
                pass
    return []


def _parse_chart_xml(
    zf: zipfile.ZipFile,
    chart_file: str,
    wb: Any = None,
    theme_colors: list[str] | None = None,
) -> ChartData | None:
    """Parse chart XML and extract data."""
    try:
        tree = ET.parse(zf.open(chart_file))
        root = tree.getroot()
    except Exception:
        return None

    data = ChartData()

    # Extract title
    for title_el in root.iter(f"{{{C_NS}}}title"):
        for t in title_el.iter(f"{{{A_NS}}}t"):
            if t.text:
                data.title = t.text.strip()
                break
        break

    # Detect chart type
    chart_type_map = {
        "barChart": "bar",
        "bar3DChart": "bar",
        "lineChart": "line",
        "line3DChart": "line",
        "pieChart": "pie",
        "pie3DChart": "pie",
        "areaChart": "area",
        "area3DChart": "area",
        "scatterChart": "scatter",
        "doughnutChart": "doughnut",
        "radarChart": "radar",
    }
    for tag, ctype in chart_type_map.items():
        chart_el = root.find(f".//{{{C_NS}}}{tag}")
        if chart_el is not None:
            data.chart_type = ctype
            # barDir
            bar_dir = chart_el.find(f"{{{C_NS}}}barDir")
            if bar_dir is not None:
                data.bar_direction = bar_dir.get("val", "col")
            grouping = chart_el.find(f"{{{C_NS}}}grouping")
            if grouping is not None:
                data.grouping = grouping.get("val", "clustered")
            # Parse series
            for idx, ser in enumerate(chart_el.findall(f"{{{C_NS}}}ser")):
                series = _parse_series(ser, wb)
                if series:
                    # Assign color: explicit spPr > theme accent
                    if not series.color and theme_colors:
                        series.color = theme_colors[idx % len(theme_colors)]
                    data.series.append(series)
            break

    # Axis titles
    for cat_ax in root.iter(f"{{{C_NS}}}catAx"):
        for t in cat_ax.iter(f"{{{A_NS}}}t"):
            if t.text:
                data.cat_axis_title = t.text.strip()
                break
    for val_ax in root.iter(f"{{{C_NS}}}valAx"):
        for t in val_ax.iter(f"{{{A_NS}}}t"):
            if t.text:
                data.val_axis_title = t.text.strip()
                break

    return data


def _parse_series(ser_el: ET.Element, wb: Any = None) -> ChartSeries | None:
    """Parse a single chart series element."""
    series = ChartSeries()

    # Series name
    tx = ser_el.find(f"{{{C_NS}}}tx")
    if tx is not None:
        # Try cached value first
        for v in tx.iter(f"{{{C_NS}}}v"):
            if v.text:
                series.name = v.text
                break
        # If no cache, try cell reference
        if not series.name and wb:
            str_ref = tx.find(f"{{{C_NS}}}strRef")
            if str_ref is not None:
                f_el = str_ref.find(f"{{{C_NS}}}f")
                if f_el is not None and f_el.text:
                    val = _read_cell_ref(wb, f_el.text)
                    if val:
                        series.name = str(val)

    # Categories
    cat = ser_el.find(f"{{{C_NS}}}cat")
    if cat is not None:
        series.categories = _extract_ref_values(cat, wb, as_str=True)

    # Values
    val = ser_el.find(f"{{{C_NS}}}val")
    if val is not None:
        series.values = _extract_ref_values(val, wb, as_str=False)

    if not series.values:
        return None

    # Extract explicit color from spPr (solidFill or line solidFill)
    sp_pr = ser_el.find(f"{{{C_NS}}}spPr")
    if sp_pr is not None:
        color = _extract_color_from_element(sp_pr)
        if color:
            series.color = color

    return series


def _extract_color_from_element(sp_pr: ET.Element) -> str | None:
    """Extract hex color from a spPr element (solidFill or line solidFill)."""
    # Direct solidFill
    for sf in sp_pr.iter(f"{{{A_NS}}}solidFill"):
        srgb = sf.find(f"{{{A_NS}}}srgbClr")
        if srgb is not None:
            return f"#{srgb.get('val', '000000')}"
        scheme = sf.find(f"{{{A_NS}}}schemeClr")
        if scheme is not None:
            # schemeClr needs theme resolution — skip for now
            pass
        break
    return None


def _extract_ref_values(
    parent: ET.Element, wb: Any = None, as_str: bool = False
) -> list:
    """Extract values from a numRef/strRef element, using cache or workbook."""
    values: list = []

    # Try cached values first (numCache or strCache)
    for cache_tag in ("numCache", "strCache"):
        cache = parent.find(f".//{{{C_NS}}}{cache_tag}")
        if cache is not None:
            for pt in cache.findall(f"{{{C_NS}}}pt"):
                v = pt.find(f"{{{C_NS}}}v")
                if v is not None and v.text:
                    if as_str:
                        values.append(v.text)
                    else:
                        try:
                            values.append(float(v.text))
                        except ValueError:
                            values.append(0.0)
            if values:
                return values

    # No cache — read from workbook
    if wb:
        for ref_tag in ("numRef", "strRef"):
            ref = parent.find(f"{{{C_NS}}}{ref_tag}")
            if ref is not None:
                f_el = ref.find(f"{{{C_NS}}}f")
                if f_el is not None and f_el.text:
                    values = _read_range_ref(wb, f_el.text, as_str=as_str)
                    if values:
                        return values

    return values


def _read_cell_ref(wb: Any, ref: str) -> Any:
    """Read a single cell value from workbook. e.g. '月次業績報告'!C12"""
    try:
        sheet_name, cell_ref = _parse_ref(ref)
        if sheet_name and cell_ref and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            return ws[cell_ref].value
    except Exception:
        pass
    return None


def _read_range_ref(
    wb: Any, ref: str, as_str: bool = False
) -> list:
    """Read a range of cell values. e.g. '月次業績報告'!$B$13:$B$16"""
    try:
        sheet_name, range_ref = _parse_ref(ref)
        if not sheet_name or not range_ref or sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        # Remove $ signs
        range_ref = range_ref.replace("$", "")
        values = []
        for row in ws[range_ref]:
            for cell in (row if isinstance(row, tuple) else [row]):
                val = cell.value
                if as_str:
                    values.append(str(val) if val is not None else "")
                else:
                    try:
                        values.append(float(val) if val is not None else 0.0)
                    except (ValueError, TypeError):
                        values.append(0.0)
        return values
    except Exception:
        return []


def _parse_ref(ref: str) -> tuple[str | None, str | None]:
    """Parse Excel reference like \"'Sheet Name'!$A$1:$B$5\" → (sheet_name, range)."""
    ref = ref.strip()
    if "!" not in ref:
        return None, ref
    parts = ref.split("!", 1)
    sheet = parts[0].strip("'")
    cell_range = parts[1]
    return sheet, cell_range


def _render_chart(chart_data: ChartData, output_path: Path) -> None:
    """Render chart data as PNG using matplotlib."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import matplotlib.font_manager as fm

    # Try to use Japanese font
    jp_fonts = [f.name for f in fm.fontManager.ttflist
                if any(n in f.name.lower() for n in ("hiragino", "meiryo", "gothic", "noto sans cjk", "yu gothic"))]
    if jp_fonts:
        plt.rcParams["font.family"] = jp_fonts[0]
    plt.rcParams["font.size"] = 10

    fig, ax = plt.subplots(figsize=(8, 5))

    if chart_data.chart_type == "pie" or chart_data.chart_type == "doughnut":
        _render_pie(ax, chart_data)
    elif chart_data.chart_type == "line":
        _render_line(ax, chart_data)
    elif chart_data.chart_type == "scatter":
        _render_scatter(ax, chart_data)
    else:
        # Default: bar chart
        _render_bar(ax, chart_data)

    if chart_data.title:
        ax.set_title(chart_data.title, fontsize=13, fontweight="bold", pad=12)

    plt.tight_layout()
    fig.savefig(str(output_path), dpi=120, bbox_inches="tight",
                facecolor="white", edgecolor="none")
    plt.close(fig)


def _render_bar(ax, chart_data: ChartData) -> None:
    """Render bar/column chart."""
    import numpy as np
    import matplotlib.pyplot as plt

    n_series = len(chart_data.series)
    if n_series == 0:
        return

    categories = chart_data.series[0].categories
    if not categories:
        categories = [str(i + 1) for i in range(len(chart_data.series[0].values))]

    n_cats = len(categories)
    x = np.arange(n_cats)

    is_horizontal = chart_data.bar_direction == "bar"
    is_stacked = "stacked" in chart_data.grouping.lower()

    fallback_colors = plt.cm.Set2(np.linspace(0, 1, max(n_series, 3)))

    if is_stacked:
        bottom = np.zeros(n_cats)
        for i, ser in enumerate(chart_data.series):
            c = ser.color or fallback_colors[i]
            vals = ser.values[:n_cats]
            while len(vals) < n_cats:
                vals.append(0.0)
            vals_arr = np.array(vals)
            if is_horizontal:
                ax.barh(x, vals_arr, left=bottom, label=ser.name or f"Series {i+1}",
                        color=c)
            else:
                ax.bar(x, vals_arr, bottom=bottom, label=ser.name or f"Series {i+1}",
                       color=c)
            bottom += vals_arr
    else:
        width = 0.8 / n_series
        for i, ser in enumerate(chart_data.series):
            c = ser.color or fallback_colors[i]
            vals = ser.values[:n_cats]
            while len(vals) < n_cats:
                vals.append(0.0)
            offset = x - 0.4 + width * (i + 0.5)
            if is_horizontal:
                ax.barh(offset, vals, height=width, label=ser.name or f"Series {i+1}",
                        color=c)
            else:
                ax.bar(offset, vals, width=width, label=ser.name or f"Series {i+1}",
                       color=c)

    if is_horizontal:
        ax.set_yticks(x)
        ax.set_yticklabels(categories)
        if chart_data.val_axis_title:
            ax.set_xlabel(chart_data.val_axis_title)
        if chart_data.cat_axis_title:
            ax.set_ylabel(chart_data.cat_axis_title)
    else:
        ax.set_xticks(x)
        ax.set_xticklabels(categories)
        if chart_data.val_axis_title:
            ax.set_ylabel(chart_data.val_axis_title)
        if chart_data.cat_axis_title:
            ax.set_xlabel(chart_data.cat_axis_title)

    if n_series > 1:
        ax.legend()
    ax.grid(axis="y" if not is_horizontal else "x", alpha=0.3)


def _render_line(ax, chart_data: ChartData) -> None:
    """Render line chart."""
    import numpy as np
    import matplotlib.pyplot as plt

    fallback_colors = plt.cm.Set1(np.linspace(0, 1, max(len(chart_data.series), 3)))
    for i, ser in enumerate(chart_data.series):
        c = ser.color or fallback_colors[i]
        categories = ser.categories or [str(j + 1) for j in range(len(ser.values))]
        ax.plot(categories, ser.values[:len(categories)],
                marker="o", label=ser.name or f"Series {i+1}",
                color=c, linewidth=2)

    if chart_data.val_axis_title:
        ax.set_ylabel(chart_data.val_axis_title)
    if chart_data.cat_axis_title:
        ax.set_xlabel(chart_data.cat_axis_title)
    if len(chart_data.series) > 1:
        ax.legend()
    ax.grid(alpha=0.3)


def _render_pie(ax, chart_data: ChartData) -> None:
    """Render pie/doughnut chart."""
    if not chart_data.series:
        return
    ser = chart_data.series[0]
    labels = ser.categories or [str(i + 1) for i in range(len(ser.values))]
    values = ser.values[:len(labels)]

    if chart_data.chart_type == "doughnut":
        wedgeprops = {"width": 0.4}
    else:
        wedgeprops = {}

    # For pie charts, use theme colors for each wedge (each data point is a "series" color-wise)
    pie_colors = None
    if chart_data.series[0].color:
        # If colors were assigned, try to get one per data point from theme
        # Pie charts typically assign accent colors per data point, not per series
        pass

    ax.pie(values, labels=labels, autopct="%1.1f%%", startangle=90,
           wedgeprops=wedgeprops)
    ax.axis("equal")


def _render_scatter(ax, chart_data: ChartData) -> None:
    """Render scatter chart."""
    import numpy as np
    import matplotlib.pyplot as plt

    fallback_colors = plt.cm.Set1(np.linspace(0, 1, max(len(chart_data.series), 3)))
    for i, ser in enumerate(chart_data.series):
        c = ser.color or fallback_colors[i]
        x_vals = list(range(len(ser.values))) if not ser.categories else [
            float(c_val) for c_val in ser.categories
        ]
        ax.scatter(x_vals[:len(ser.values)], ser.values,
                   label=ser.name or f"Series {i+1}", color=c)

    if chart_data.val_axis_title:
        ax.set_ylabel(chart_data.val_axis_title)
    if chart_data.cat_axis_title:
        ax.set_xlabel(chart_data.cat_axis_title)
    if len(chart_data.series) > 1:
        ax.legend()
    ax.grid(alpha=0.3)


def _build_chart_anchor_map(zf: zipfile.ZipFile) -> dict[str, str | None]:
    """Map chart names (e.g. 'chart1') to anchor cell positions."""
    xdr_ns = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
    c_ns_chart = "http://schemas.openxmlformats.org/drawingml/2006/chart"

    chart_to_anchor: dict[str, str | None] = {}

    for name in zf.namelist():
        if not name.startswith("xl/drawings/") or not name.endswith(".xml") or "_rels" in name:
            continue

        rels_path = name.replace("xl/drawings/", "xl/drawings/_rels/") + ".rels"
        rid_to_chart: dict[str, str] = {}
        if rels_path in zf.namelist():
            try:
                rels_tree = ET.parse(zf.open(rels_path))
                for rel in rels_tree.getroot().findall(f"{{{REL_NS}}}Relationship"):
                    target = rel.get("Target", "")
                    if "chart" in target.lower():
                        rid = rel.get("Id", "")
                        chart_name = Path(target).stem  # "chart1"
                        rid_to_chart[rid] = chart_name
            except Exception:
                continue

        if not rid_to_chart:
            continue

        try:
            tree = ET.parse(zf.open(name))
            root = tree.getroot()
            for anchor_tag in (f"{{{xdr_ns}}}twoCellAnchor", f"{{{xdr_ns}}}oneCellAnchor"):
                for anchor in root.findall(anchor_tag):
                    from_el = anchor.find(f"{{{xdr_ns}}}from")
                    cell_ref = None
                    if from_el is not None:
                        col_el = from_el.find(f"{{{xdr_ns}}}col")
                        row_el = from_el.find(f"{{{xdr_ns}}}row")
                        if col_el is not None and row_el is not None:
                            from openpyxl.utils import get_column_letter
                            col = int(col_el.text or "0")
                            row = int(row_el.text or "0")
                            cell_ref = f"{get_column_letter(col + 1)}{row + 1}"

                    for chart_el in anchor.iter(f"{{{c_ns_chart}}}chart"):
                        rid = chart_el.get(f"{{{R_NS}}}id", "")
                        if rid in rid_to_chart:
                            chart_to_anchor[rid_to_chart[rid]] = cell_ref
        except Exception:
            continue

    return chart_to_anchor


def _build_chart_sheet_map(zf: zipfile.ZipFile) -> dict[str, str]:
    """Map chart names to sheet names via drawing → worksheet chain."""
    chart_to_sheet: dict[str, str] = {}

    # drawing → charts
    drawing_to_charts: dict[str, list[str]] = {}
    for name in zf.namelist():
        if not name.startswith("xl/drawings/_rels/") or not name.endswith(".xml.rels"):
            continue
        drawing_name = Path(name).stem  # "drawing1.xml"
        try:
            tree = ET.parse(zf.open(name))
            for rel in tree.getroot().findall(f"{{{REL_NS}}}Relationship"):
                target = rel.get("Target", "")
                if "chart" in target.lower():
                    drawing_to_charts.setdefault(drawing_name, []).append(Path(target).stem)
        except Exception:
            continue

    # worksheet → drawing
    for name in zf.namelist():
        if not name.startswith("xl/worksheets/_rels/") or not name.endswith(".xml.rels"):
            continue
        sheet_file = Path(name).stem
        try:
            tree = ET.parse(zf.open(name))
            for rel in tree.getroot().findall(f"{{{REL_NS}}}Relationship"):
                target = rel.get("Target", "")
                if "drawing" in target.lower():
                    drawing_name = Path(target).name  # "drawing1.xml"
                    # Find sheet name from workbook
                    sheet_idx_match = re.search(r"sheet(\d+)", sheet_file)
                    if sheet_idx_match:
                        # We need workbook sheetnames, but don't have wb here
                        # Store index-based name for now
                        chart_to_sheet_idx = int(sheet_idx_match.group(1)) - 1
                        for chart_name in drawing_to_charts.get(drawing_name, []):
                            chart_to_sheet[chart_name] = f"__sheet_idx_{chart_to_sheet_idx}"
        except Exception:
            continue

    return chart_to_sheet
